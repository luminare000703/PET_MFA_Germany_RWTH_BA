# ============================================
# STEP 1 – Initialize
# ============================================

import sys
import os
from pathlib import Path
from copy import deepcopy
import logging as log

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# --- ODYM imports  ---
import odym.classes as msc
import odym.functions as msf
import odym.dynamic_stock_model as dsm


# --------------------------------------------
# 1.1 Project paths
# --------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data_szenario"
RESULTS_DIR = PROJECT_DIR / "results"
LOG_DIR = PROJECT_DIR / "logs"

RESULTS_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)


# --------------------------------------------
# 1.2 Logging
# --------------------------------------------
log_verbosity = log.DEBUG
log_filename = LOG_DIR / "PET_MFA_Szenario_S1a1.log"

[Mylog, console_log, file_log] = msf.function_logger(
    str(log_filename),
    str(LOG_DIR),
    log_verbosity,
    log_verbosity
)

Mylog.info("### 1 - Initialize PET MFA ODYM model")
Mylog.info(f"SCRIPT_DIR:  {SCRIPT_DIR}")
Mylog.info(f"PROJECT_DIR: {PROJECT_DIR}")
Mylog.info(f"DATA_DIR:    {DATA_DIR}")
Mylog.info(f"RESULTS_DIR: {RESULTS_DIR}")

# ============================================
# STEP 2 – Load Config file and read control parameters
# ============================================

Mylog.info("### 2 - Load Config file and read model control parameters")

# -------------------------------------------------
# 2.1 Load project-specific config file
# -------------------------------------------------
ProjectSpecs_Name_ConFile = "ODYM_Config_S1a1.xlsx"
ConfigPath = DATA_DIR / ProjectSpecs_Name_ConFile

if not ConfigPath.exists():
    raise FileNotFoundError(f"Config file not found: {ConfigPath}")

Model_Configfile = openpyxl.load_workbook(str(ConfigPath), data_only=True)

# -------------------------------------------------
# 2.2 Read model setting (baseline / scenario name)
# -------------------------------------------------
# Convention: sheet "Config", cell (4,4)
ConfigSheet = Model_Configfile["Config"]

ScriptConfig = {}
ScriptConfig["Model Setting"] = ConfigSheet.cell(4, 4).value

Mylog.info(f"Model Setting: {ScriptConfig['Model Setting']}")

# Load the corresponding setting sheet
SettingSheetName = "Setting_" + ScriptConfig["Model Setting"]
if SettingSheetName not in Model_Configfile.sheetnames:
    raise KeyError(f"Expected sheet '{SettingSheetName}' not found in config file.")

Model_Configsheet = Model_Configfile[SettingSheetName]

# Optional: scenario name
Name_Scenario = Model_Configsheet.cell(4, 4).value
Mylog.info(f"Scenario name: {Name_Scenario}")


# -------------------------------------------------
# 2.3 Read model control parameters – General Info
# -------------------------------------------------
SCix = 0
while Model_Configsheet.cell(SCix + 1, 2).value != "General Info":
    SCix += 1

SCix += 2

while Model_Configsheet.cell(SCix + 1, 4).value is not None:
    key = Model_Configsheet.cell(SCix + 1, 3).value
    val = Model_Configsheet.cell(SCix + 1, 4).value
    ScriptConfig[key] = val
    SCix += 1


# -------------------------------------------------
# 2.4 Read software version selection
# -------------------------------------------------
SCix = 0
while Model_Configsheet.cell(SCix + 1, 2).value != "Software version selection":
    SCix += 1

SCix += 2

while Model_Configsheet.cell(SCix + 1, 4).value is not None:
    key = Model_Configsheet.cell(SCix + 1, 3).value
    val = Model_Configsheet.cell(SCix + 1, 4).value
    ScriptConfig[key] = val
    SCix += 1


# -------------------------------------------------
# 2.5 Log final ScriptConfig
# -------------------------------------------------
Mylog.info("ScriptConfig successfully read:")
for k, v in ScriptConfig.items():
    Mylog.info(f"  {k}: {v}")

# ============================================
# STEP 3 – Read classification and config blocks
# ============================================

Mylog.info("### 3 - Read classification and data")

# -------------------------------------------------
# Helper functions
# -------------------------------------------------
def find_row_by_label(ws, label: str, label_col: int = 2, start_row: int = 1, max_scan: int = 2000) -> int:

    r = start_row
    for _ in range(max_scan):
        if ws.cell(r, label_col).value == label:
            return r
        r += 1
    raise ValueError(f"Label '{label}' not found in column {label_col} within {max_scan} rows.")


def safe_liststring_to_int_list(raw):
    if raw is None:
        return []
    if isinstance(raw, str) and raw.strip() == "":
        return []
    if isinstance(raw, (list, tuple)):
        return [int(x) for x in raw]

    s = str(raw).strip()
    # allow '(1,2,3)' style
    if s.startswith("(") and s.endswith(")"):
        s = "[" + s[1:-1] + "]"
    return msf.ListStringToListNumbers(s)


# -------------------------------------------------
# 3.1 Load master classification file
# -------------------------------------------------
if "Version of master classification" not in ScriptConfig or ScriptConfig["Version of master classification"] is None:
    raise KeyError("ScriptConfig['Version of master classification'] is missing. Check Step 2 General Info block.")

ClassPath = DATA_DIR / f"{ScriptConfig['Version of master classification']}.xlsx"
if not ClassPath.exists():
    raise FileNotFoundError(f"Classification file not found: {ClassPath}")

Classfile = openpyxl.load_workbook(str(ClassPath), data_only=True)
if "MAIN_Table" not in Classfile.sheetnames:
    raise KeyError(f"'MAIN_Table' not found in classification file. Sheets: {Classfile.sheetnames}")

Classsheet = Classfile["MAIN_Table"]

# Read MAIN_Table
ci = 1
MasterClassification = {}

max_cols = Classsheet.max_column
while (ci + 1) <= max_cols:
    if Classsheet.cell(1, ci + 1).value is None:
        Mylog.info(f"End of MAIN_Table reached at column {ci}.")
        break

    ThisName = Classsheet.cell(1, ci + 1).value
    ThisDim  = Classsheet.cell(2, ci + 1).value
    ThisID   = Classsheet.cell(4, ci + 1).value
    ThisUUID = Classsheet.cell(5, ci + 1).value

    TheseItems = []
    ri = 10

    first_item = Classsheet.cell(ri + 1, ci + 1).value
    if first_item is not None:
        TheseItems.append(first_item)

    while True:
        ri += 1
        v = Classsheet.cell(ri + 1, ci + 1).value
        if v is None:
            break
        TheseItems.append(v)

    MasterClassification[ThisName] = msc.Classification(
        Name=ThisName,
        Dimension=ThisDim,
        ID=ThisID,
        UUID=ThisUUID,
        Items=TheseItems
    )
    ci += 1

Mylog.info(f"Loaded {len(MasterClassification)} master classifications from {ClassPath.name}")


# -------------------------------------------------
# 3.2 Read Index Table block
# -------------------------------------------------
Mylog.info("Read index table from model config sheet.")
idx_row = find_row_by_label(Model_Configsheet, "Index Table", label_col=2)
r = idx_row + 2  # first data row

IT_Aspects, IT_Description, IT_Dimension = [], [], []
IT_Classification, IT_Selector, IT_IndexLetter = [], [], []

while Model_Configsheet.cell(r, 3).value is not None:
    IT_Aspects.append(Model_Configsheet.cell(r, 3).value)
    IT_Description.append(Model_Configsheet.cell(r, 4).value)
    IT_Dimension.append(Model_Configsheet.cell(r, 5).value)
    IT_Classification.append(Model_Configsheet.cell(r, 6).value)
    IT_Selector.append(Model_Configsheet.cell(r, 7).value)
    IT_IndexLetter.append(Model_Configsheet.cell(r, 8).value)
    r += 1

Mylog.info(f"Index table rows read: {len(IT_Aspects)}")


# -------------------------------------------------
# 3.3 Read Model Parameters block
# -------------------------------------------------
Mylog.info("Read parameter list from model config sheet.")
pl_row = find_row_by_label(Model_Configsheet, "Model Parameters", label_col=2)
r = pl_row + 2  # first data row

PL_Names, PL_Description, PL_Version = [], [], []
PL_IndexStructure, PL_IndexMatch, PL_IndexLayer = [], [], []

while Model_Configsheet.cell(r, 3).value is not None:
    PL_Names.append(Model_Configsheet.cell(r, 3).value)
    PL_Description.append(Model_Configsheet.cell(r, 4).value)
    PL_Version.append(Model_Configsheet.cell(r, 5).value)
    PL_IndexStructure.append(Model_Configsheet.cell(r, 6).value)
    PL_IndexMatch.append(Model_Configsheet.cell(r, 7).value)

    raw_layer = Model_Configsheet.cell(r, 8).value
    PL_IndexLayer.append(safe_liststring_to_int_list(raw_layer))
    r += 1

Mylog.info(f"Parameter list rows read: {len(PL_Names)}")
if len(PL_Names) == 0:
    Mylog.warning("Parameter list rows read: 0  --> Config sheet 'Model Parameters' Checking header or the location of the block.")

# -------------------------------------------------
# 3.4 Read Process Group List block
# -------------------------------------------------
Mylog.info("Read process list from model config sheet.")
pr_row = find_row_by_label(Model_Configsheet, "Process Group List", label_col=2)
r = pr_row + 2

PrL_Number, PrL_Name, PrL_Code, PrL_Type = [], [], [], []

while Model_Configsheet.cell(r, 3).value is not None:
    v = Model_Configsheet.cell(r, 3).value
    try:
        PrL_Number.append(int(v))
    except Exception:
        PrL_Number.append(v)

    PrL_Name.append(Model_Configsheet.cell(r, 4).value)
    PrL_Code.append(Model_Configsheet.cell(r, 5).value)
    PrL_Type.append(Model_Configsheet.cell(r, 6).value)
    r += 1

Mylog.info(f"Process group list rows read: {len(PrL_Number)}")


# -------------------------------------------------
# 3.5 Read Model flow control block (append into ScriptConfig)
# -------------------------------------------------
Mylog.info("Read model run control from model config sheet.")
mf_row = find_row_by_label(Model_Configsheet, "Model flow control", label_col=2)
r = mf_row + 2

while Model_Configsheet.cell(r, 3).value is not None:
    key = Model_Configsheet.cell(r, 3).value
    val = Model_Configsheet.cell(r, 4).value
    if key is not None:
        ScriptConfig[key] = val
    r += 1

Mylog.info("Model flow control entries added to ScriptConfig.")


# -------------------------------------------------
# 3.6 Quick sanity logs
# -------------------------------------------------
Mylog.info(f"ModelClassification keys loaded (master): {list(MasterClassification.keys())[:8]} ...")
Mylog.info(f"IT_Aspects: {IT_Aspects}")
Mylog.info(f"First 5 parameters: {PL_Names[:5]}")

# ============================================
# STEP 4 – Build ModelClassification from Index Table (apply selectors)
# ============================================

Mylog.info("### 4 - Define model classifications and select items according to config file.")

ModelClassification = {}

def normalize_selector(x):

    if x is None:
        return "all"

    if isinstance(x, str):
        s = x.strip()
        if s == "":
            return "all"
        if s.lower() == "all":
            return "all"
        return s

    if isinstance(x, (int, float)):
        return str(int(x))

    return str(x).strip()


import ast

for a, cls_name, selector_raw in zip(IT_Aspects, IT_Classification, IT_Selector):

    if cls_name not in MasterClassification:
        raise KeyError(
            f"Index Table requests classification '{cls_name}' for aspect '{a}', "
            f"but it is not in MasterClassification."
        )

    ModelClassification[a] = deepcopy(MasterClassification[cls_name])

    selector = normalize_selector(selector_raw)

    items = ModelClassification[a].Items
    n_items = len(items)

    if selector == "all":
        pass

    elif (":" in selector) or ("[" in selector):
        pass

    elif selector.lstrip("-").isdigit():
        num = int(selector)
        if 0 <= num < n_items:
            selector = f"[{num}]"
        elif num in items:
            selector = f"[{items.index(num)}]"
        elif str(num) in [str(it) for it in items]:
            idx = [str(it) for it in items].index(str(num))
            selector = f"[{idx}]"
        else:
            raise ValueError(
                f"Numeric selector '{num}' for aspect '{a}' is neither a valid index "
                f"(0..{n_items - 1}) nor an existing item value. "
                f"Classification='{cls_name}', items={items}"
            )

    else:
        if selector in items:
            selector = f"[{items.index(selector)}]"
        else:
            low = [str(it).strip().lower() for it in items]
            s_low = selector.strip().lower()
            hits = [k for k, it in enumerate(low) if it == s_low]
            if len(hits) == 1:
                selector = f"[{hits[0]}]"
            else:
                hits2 = [k for k, it in enumerate(low) if s_low in it]
                if len(hits2) == 1:
                    selector = f"[{hits2[0]}]"
                else:
                    raise ValueError(
                        f"Text selector '{selector}' for aspect '{a}' could not be matched uniquely.\n"
                        f"Classification='{cls_name}', n_items={n_items}\n"
                        f"Items={items}"
                    )

    Mylog.info(
        f"Selector check: aspect={a}, raw={selector_raw} ({type(selector_raw)}), normalized={selector}"
    )

    EvalString = msf.EvalItemSelectString(
        selector,
        len(ModelClassification[a].Items)
    )

    if EvalString.find(':') > -1:
        RangeStart = int(EvalString.split(':')[0])
        RangeStop  = int(EvalString.split(':')[1])
        ModelClassification[a].Items = ModelClassification[a].Items[RangeStart:RangeStop]

    elif EvalString.find('[') > -1:
        idx_list = ast.literal_eval(EvalString)

        n_items = len(ModelClassification[a].Items)
        bad = [i for i in idx_list if (not isinstance(i, int)) or i < 0 or i >= n_items]

        if bad:
            Mylog.error(
                f"[IndexError prevent] Selector out of range!\n"
                f"  aspect={a}\n"
                f"  classification={cls_name}\n"
                f"  selector_raw={selector_raw} ({type(selector_raw)})\n"
                f"  normalized_selector={selector}\n"
                f"  EvalString={EvalString}\n"
                f"  n_items={n_items}\n"
                f"  bad_indices={bad}\n"
                f"  available_index_range=0..{n_items-1}"
            )
            raise IndexError(
                f"Selector for aspect '{a}' selects indices {bad} "
                f"but only {n_items} items exist."
            )

        ModelClassification[a].Items = [
            ModelClassification[a].Items[i] for i in idx_list
        ]

    elif EvalString == 'all':
        pass

    else:
        raise ValueError(
            f"Invalid selector '{selector}' for aspect '{a}'. "
            f"EvalItemSelectString returned '{EvalString}'."
        )

# --- Step 4 end: index sizes ---
Mylog.info(f"ModelClassification aspects: {list(ModelClassification.keys())}")
IndexSize = {a: len(ModelClassification[a].Items) for a in IT_Aspects}
Mylog.info(f"Index sizes: {IndexSize}")

# ============================================
# STEP 5 – Read parameters from Excel files
# ============================================

import inspect
import re

Mylog.info("### 5 - Read parameters from Excel files")

# ODYM uncertainty parsing ON
ParseUncertainty = True


# -------------------------------------------------
# 5.0 Build IndexTable (ODYM legacy requirement)
# -------------------------------------------------
IndexTable = pd.DataFrame({
    "Aspect": IT_Aspects,
    "Description": IT_Description,
    "Dimension": IT_Dimension,
    "Classification": [ModelClassification[a] for a in IT_Aspects],
    "ClassificationName": IT_Classification,
    "Selector": IT_Selector,
    "IndexLetter": IT_Aspects,
    "IndexSize": [len(ModelClassification[a].Items) for a in IT_Aspects],
})
IndexTable_ClassificationNames = IT_Classification

Mylog.info("IndexTable IndexLetter check: " + str(IndexTable["IndexLetter"].tolist()))
Mylog.info("IndexTable IndexSize check: " + str(IndexTable["IndexSize"].tolist()))


# -------------------------------------------------
# 5.1 ReadParameterXLSX signature (robust dispatch)
# -------------------------------------------------
sig = inspect.signature(msf.ReadParameterXLSX)
param_names = list(sig.parameters.keys())

Mylog.info(f"ReadParameterXLSX signature: {sig}")
Mylog.info(f"ReadParameterXLSX parameters: {param_names}")


# -------------------------------------------------
# 5.2 Helper: IndexMatch / LayerSel / Ix
# -------------------------------------------------
def convert_indexmatch_to_int_list_string(indexmatch_raw, indexstructure_raw):
    if indexmatch_raw is None or str(indexmatch_raw).strip() == "":
        indexmatch_raw = str(indexstructure_raw)

    s = str(indexmatch_raw).strip()

    if s.startswith("[") and s.endswith("]"):
        return s

    if "*" in s:
        parts = [x.strip() for x in s.split("*")]
    elif "," in s:
        parts = [x.strip() for x in s.split(",")]
    else:
        parts = [s]

    idx = [IT_Aspects.index(letter) for letter in parts]
    return str(idx)


def normalize_layersel(ls):
    if ls is None:
        return [[0]]
    if isinstance(ls, list):
        if len(ls) == 0:
            return [[0]]
        if isinstance(ls[0], int):
            return [ls]
        return ls
    return [[int(ls)]]


def build_thisparix_from_indexstructure(indexstructure):
    return [s.strip() for s in str(indexstructure).split("*") if s.strip()]

# -------------------------------------------------
# 5.3 Robust ODYM call (AUTO signature dispatch)
# -------------------------------------------------
import inspect

_READPAR_SIG = inspect.signature(msf.ReadParameterXLSX)
_READPAR_PARAMS = list(_READPAR_SIG.parameters.keys())
Mylog.info(f"[STEP5] ReadParameterXLSX detected params: {_READPAR_PARAMS}")

def call_ReadParameterXLSX_auto(
    *,
    p_path,
    ThisPar,
    ThisParIx,
    IndexMatch,
    ThisParLayerSel,
    MasterClassification,
    IndexTable,
    IndexTable_ClassificationNames,
    ScriptConfig,
    Mylog,
    ParseUncertainty
):

    # ODYM expects ParPath WITHOUT .xlsx
    p_path = str(p_path)
    if p_path.lower().endswith(".xlsx"):
        p_path = p_path[:-5]

    # Provide ProcMethod if ODYM signature includes it
    proc_method = "[]"
    needs_proc = ("ThisParProcMethod" in _READPAR_PARAMS)

    try:
        if needs_proc:
            return msf.ReadParameterXLSX(
                p_path,
                ThisPar=ThisPar,
                ThisParIx=ThisParIx,
                IndexMatch=IndexMatch,
                ThisParLayerSel=ThisParLayerSel,
                MasterClassification=MasterClassification,
                IndexTable=IndexTable,
                IndexTable_ClassificationNames=IndexTable_ClassificationNames,
                ScriptConfig=ScriptConfig,
                Mylog=Mylog,
                ThisParProcMethod=proc_method,
                ParseUncertainty=ParseUncertainty
            )
        else:
            return msf.ReadParameterXLSX(
                p_path,
                ThisPar=ThisPar,
                ThisParIx=ThisParIx,
                IndexMatch=IndexMatch,
                ThisParLayerSel=ThisParLayerSel,
                MasterClassification=MasterClassification,
                IndexTable=IndexTable,
                IndexTable_ClassificationNames=IndexTable_ClassificationNames,
                ScriptConfig=ScriptConfig,
                Mylog=Mylog,
                ParseUncertainty=ParseUncertainty
            )
    except TypeError as e:

        Mylog.warning(f"[STEP5] ReadParameterXLSX tutorial-style call failed -> trying positional. Reason: {e}")

    # ---- Fallback: classic positional signature ----
    ThisParType = "Parameter"
    ThisParName = ThisPar
    ThisParUnit = ""
    Elements = ["All"]

    # Try ThisParDim as list first, then as string
    try:
        return msf.ReadParameterXLSX(
            p_path,
            ThisParType,
            ThisParName,
            ThisParUnit,
            ThisParIx,
            IndexTable,
            Elements,
            proc_method,
            ParseUncertainty=ParseUncertainty
        )
    except TypeError:
        ThisParDim_str = "*".join(ThisParIx) if isinstance(ThisParIx, (list, tuple)) else str(ThisParIx)
        return msf.ReadParameterXLSX(
            p_path,
            ThisParType,
            ThisParName,
            ThisParUnit,
            ThisParDim_str,
            IndexTable,
            Elements,
            proc_method,
            ParseUncertainty=ParseUncertainty
        )

# -------------------------------------------------
# StdDev reading helpers
# -------------------------------------------------
def _to_float_or_none(x):

    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return None

    m = re.match(r"^(-?\d+(\.\d+)?)\s*%$", s)
    if m:
        return float(m.group(1)) / 100.0

    try:
        return float(s)
    except Exception:
        return None


def _find_header_row(ws, required_cols, max_scan_rows=80, max_scan_cols=50):
    req = [c.strip().lower() for c in required_cols]

    for rr in range(1, max_scan_rows + 1):
        row_vals = []
        for cc in range(1, max_scan_cols + 1):
            v = ws.cell(rr, cc).value
            row_vals.append("" if v is None else str(v).strip())

        low = [x.lower() for x in row_vals]
        if all(any(rc == lv for lv in low) for rc in req):
            col_map = {}
            for rc in req:
                col_map[rc] = low.index(rc)
            return rr, col_map

    return None, None


def _get_sheet_for_values_table(wb):
    candidates = []
    for name in ["Values_Master", "Values", "values_master", "values"]:
        if name in wb.sheetnames:
            candidates.append(name)

    if not candidates:
        for sn in wb.sheetnames:
            ws = wb[sn]
            rr, _ = _find_header_row(ws, required_cols=["value"], max_scan_rows=60, max_scan_cols=60)
            if rr is not None:
                candidates.append(sn)

    return candidates[0] if candidates else None


def _index_lookup_for_aspect(ModelClassification, aspect_letter):
    items = list(ModelClassification[aspect_letter].Items)
    lut = {}
    for idx, it in enumerate(items):
        lut[str(it).strip()] = idx
    return lut


def read_stddev_list_parameter(
    *,
    xlsx_path,
    index_structure,
    ModelClassification,
    target_shape,
):
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheet_name = _get_sheet_for_values_table(wb)
    if sheet_name is None:
        return None

    ws = wb[sheet_name]

    aspects = [s.strip() for s in str(index_structure).split("*") if s.strip()]
    required = aspects + ["value", "StdDev"]

    header_row, col_map = _find_header_row(ws, required_cols=required, max_scan_rows=120, max_scan_cols=80)
    if header_row is None:
        return None

    aspect_luts = {a: _index_lookup_for_aspect(ModelClassification, a) for a in aspects}

    SD = np.zeros(target_shape, dtype=float)
    filled = 0

    rr = header_row + 1
    while True:
        first_val = ws.cell(rr, col_map[aspects[0].lower()] + 1).value
        val_cell  = ws.cell(rr, col_map["value"] + 1).value
        sd_cell   = ws.cell(rr, col_map["stddev"] + 1).value

        if first_val is None and val_cell is None and sd_cell is None:
            break

        idxs = []
        ok = True
        for a in aspects:
            col0 = col_map[a.lower()] + 1
            v = ws.cell(rr, col0).value
            if v is None:
                ok = False
                break

            key = str(v).strip()
            lut = aspect_luts[a]
            if key not in lut:
                key2 = str(int(v)) if isinstance(v, float) and float(v).is_integer() else key
                if key2 in lut:
                    idxs.append(lut[key2])
                else:
                    ok = False
                    break
            else:
                idxs.append(lut[key])

        if not ok:
            rr += 1
            continue

        sd = _to_float_or_none(sd_cell)
        if sd is None:
            rr += 1
            continue

        SD[tuple(idxs)] = float(sd)
        filled += 1
        rr += 1

    if filled == 0:
        return None

    return SD


# -------------------------------------------------
# 5.4 Extract Values + StdDev from ANY ODYM return
# -------------------------------------------------
def extract_values_and_std(res, par_name=""):
    cover = None
    values = None
    std = None

    def _take(x):
        nonlocal cover, values, std

        if isinstance(x, np.ndarray):
            if values is None:
                values = x
            elif std is None and x.shape == values.shape:
                std = x
            return

        if isinstance(x, dict):
            if "Values" in x:
                values = np.asarray(x["Values"])
                if "StdDev" in x:
                    std = np.asarray(x["StdDev"])
            else:
                cover = x
            return

        if hasattr(x, "Values"):
            values = np.asarray(x.Values)
            if hasattr(x, "StdDev"):
                std = np.asarray(x.StdDev)
            return

    if isinstance(res, tuple):
        for x in res:
            _take(x)
    else:
        _take(res)

    if values is None:
        meta = list(cover.keys()) if isinstance(cover, dict) else None
        raise TypeError(
            f"[STEP5 FAIL] '{par_name}' returned no Values. "
            f"type={type(res)}, cover_keys={meta}"
        )

    return values, std, cover


# -------------------------------------------------
# 5.5 Read all parameters → ParameterDict
# -------------------------------------------------
Mylog.info("### 5.5 - Read parameters (xlsx) and build ParameterDict")

ParameterDict = {}

for k in range(len(PL_Names)):
    ThisPar = PL_Names[k]

    IndexStructure = PL_IndexStructure[k]
    ThisParIx = build_thisparix_from_indexstructure(IndexStructure)

    IndexMatch = convert_indexmatch_to_int_list_string(
        PL_IndexMatch[k], IndexStructure
    )

    ThisParLayerSel = normalize_layersel(PL_IndexLayer[k])

    p_file = PL_Version[k]
    if not str(p_file).lower().endswith(".xlsx"):
        p_file = str(p_file) + ".xlsx"
    p_path = DATA_DIR / p_file

    Mylog.info(
        f"[{k+1}/{len(PL_Names)}] Reading '{ThisPar}' from {p_path.name} "
        f"(Ix={IndexStructure}, Match={IndexMatch}, LayerSel={ThisParLayerSel})"
    )

    Par_raw = call_ReadParameterXLSX_auto(
        p_path=p_path,
        ThisPar=ThisPar,
        ThisParIx=ThisParIx,
        IndexMatch=IndexMatch,
        ThisParLayerSel=ThisParLayerSel,
        MasterClassification=MasterClassification,
        IndexTable=IndexTable,
        IndexTable_ClassificationNames=IndexTable_ClassificationNames,
        ScriptConfig=ScriptConfig,
        Mylog=Mylog,
        ParseUncertainty=ParseUncertainty
    )

    V, SD, COVER = extract_values_and_std(Par_raw, ThisPar)


    if SD is None:
        SD = read_stddev_list_parameter(
            xlsx_path=p_path,
            index_structure=IndexStructure,
            ModelClassification=ModelClassification,
            target_shape=V.shape
        )

    ParameterDict[ThisPar] = {
        "Values": V,
        "StdDev": SD,
        "Cover": COVER
    }

    Mylog.info(
        f"  -> OK: '{ThisPar}' Values={V.shape}, StdDev={'yes' if SD is not None else 'no'}"
    )

Mylog.info(f"### STEP 5 finished: {len(ParameterDict)} parameters loaded")

# ============================================================
# STEP 6 + STEP 7
# ============================================================

import numpy as np
import pandas as pd
import time
from pathlib import Path
import plotly.graph_objects as go


# ============================================================
# USER SETTINGS (MC)
# ============================================================
N_MC = 500
SEED = 42
LOG_EVERY = 50
rng = np.random.default_rng(SEED)

MC_DIR = RESULTS_DIR / "mc"
MC_DIR.mkdir(parents=True, exist_ok=True)

Mylog.info(f"### STEP6 settings: N_MC={N_MC}, SEED={SEED}, LOG_EVERY={LOG_EVERY}, MC_DIR={MC_DIR}")


# ============================================================
# STEP 6.0 – Create PET_MFA_System
# ============================================================
Mylog.info("### STEP6.0 - Creating PET_MFA_System (MFAsystem)")

t_items = list(ModelClassification["t"].Items)

def _to_int_year(x):
    if isinstance(x, (int, np.integer)):
        return int(x)
    s = str(x).strip()
    try:
        return int(float(s))
    except Exception:
        for i in range(len(s) - 3):
            chunk = s[i:i+4]
            if chunk.isdigit():
                return int(chunk)
        raise ValueError(f"Cannot parse time item: {x}")

Time_Start = _to_int_year(t_items[0])
Time_End   = _to_int_year(t_items[-1])

if "Chemical_Elements" in MasterClassification:
    Elements = MasterClassification["Chemical_Elements"]
else:
    Elements = msc.Classification(
        Name="Chemical_Elements",
        Dimension="Element",
        ID="E",
        UUID="",
        Items=["All"]
    )

ProcessList = []
for n, name in zip(PrL_Number, PrL_Name):
    ProcessList.append(msc.Process(ID=int(n), Name=str(name)))

PET_MFA_System = msc.MFAsystem(
    Name=f"PET_MFA_DE_2019_{ScriptConfig.get('Model Setting','Baseline')}",
    Geogr_Scope="Germany",
    Unit="kt",
    Time_Start=Time_Start,
    Time_End=Time_End,
    IndexTable=IndexTable,
    Elements=Elements,
    ProcessList=ProcessList,
    FlowDict={},
    StockDict={},
    ParameterDict={}
)

Mylog.info("### STEP6.0 - PET_MFA_System created successfully")

# ============================================================
# STEP 6.x – Monte Carlo helper functions
# ============================================================

def _get_V_SD(entry):

    if isinstance(entry, np.ndarray):
        return entry, None
    if isinstance(entry, dict):
        V = entry.get("Values", None)
        SD = entry.get("StdDev", None)
        if V is None:
            raise TypeError("Parameter entry dict without 'Values'")
        return np.asarray(V, dtype=float), (None if SD is None else np.asarray(SD, dtype=float))
    if hasattr(entry, "Values"):
        V = np.asarray(entry.Values, dtype=float)
        SD = np.asarray(entry.StdDev, dtype=float) if hasattr(entry, "StdDev") else None
        return V, SD
    raise TypeError(f"Unsupported parameter entry type: {type(entry)}")


def _clip_nonneg(A):
    return np.clip(np.asarray(A, dtype=float), 0.0, None)

def _clip_01(A):
    return np.clip(np.asarray(A, dtype=float), 0.0, 1.0)

def _safe_normal(mean, sd, rng):
    mean = np.asarray(mean, dtype=float)
    if sd is None:
        return mean.copy()
    sd = np.asarray(sd, dtype=float)
    if sd.size == 0 or np.all(sd == 0):
        return mean.copy()
    return rng.normal(loc=mean, scale=sd)

def _row_normalize_last_axis(A, eps=1e-12):
    A = np.asarray(A, dtype=float)
    s = A.sum(axis=-1, keepdims=True)
    return np.where(s > eps, A / s, A)

def _sample_mass(mean, sd, rng):
    return _clip_nonneg(_safe_normal(mean, sd, rng))

def _sample_prob01(mean, sd, rng):
    return _clip_01(_safe_normal(mean, sd, rng))

def _sample_share(mean, sd, rng):
    x = _clip_nonneg(_safe_normal(mean, sd, rng))
    return _row_normalize_last_axis(x)

def _sample_matrix(mean, sd, rng):
    # matrix (..., i, j) with j last axis: normalize rows along last axis
    x = _clip_nonneg(_safe_normal(mean, sd, rng))
    return _row_normalize_last_axis(x)

def sample_parameter(name, V, SD, rng):

    # MASS
    if name.startswith("p_mass_"):
        return _sample_mass(V, SD, rng)

    # SPECIFIC SHARES (row-normalize last axis)
    if name in ("p_share_use_waste_type_S1a1"):
        return _sample_share(V, SD, rng)

    # ROUTING MATRICES (row-normalize last axis)
    if name in (
        "p_share_collection_routing_S1a1",
        "p_share_sorting_separate_S1a1",
        "p_share_sorting_mixed_S1a1",
        "p_share_recycling_routing_S1a1",
        "p_share_recycling_stream_routing",
        "p_share_sorting_to_terminal",
    ):
        return _sample_matrix(V, SD, rng)

    # PROBABILITIES (0..1)
    if name in (
        "p_collection_efficiency_w_S1a1",
        "p_rPET_eligible_share_packaging_tr",
        "p_share_recycled_output_to_PM",
    ):
        return _sample_prob01(V, SD, rng)

    # YIELDS (0..1, no normalization)
    if name == "p_recycling_process_yields_S1a1":
        return _sample_prob01(V, SD, rng)

    # DEFAULT for other p_share_*:
    if name.startswith("p_share_"):
        # if it's vector-like (<=3 dims), treat as prob; else as share
        if np.asarray(V).ndim <= 3:
            return _sample_prob01(V, SD, rng)
        return _sample_share(V, SD, rng)

    # FALLBACK: nonnegative
    return _sample_mass(V, SD, rng)

def build_sampled_parameterdict(ParameterDict_base, rng):

    PD = {}
    for name, entry in ParameterDict_base.items():
        V, SD = _get_V_SD(entry)
        PD[name] = sample_parameter(name, V, SD, rng)
    return PD

def apply_hybrid_constraints_out7_relative_to_baseline(
    out7: dict,
    baseline_out7: dict,
    Mylog=None,
    enforce_openloop_min_share=0.10,   # 10% open-loop minimum
    enforce_no_inflation=True,
):


    import numpy as np

    def _safe_sum(x, default=0.0):
        try:
            v = float(np.nansum(x))
            if not np.isfinite(v):
                return float(default)
            return v
        except Exception:
            return float(default)

    def _sum_prefix(d, prefixes):
        tot = 0.0
        for k, v in d.items():
            if any(k.startswith(p) for p in prefixes):
                try:
                    tot += float(np.nansum(v))
                except Exception:
                    pass
        return tot

    # -------------------------
    # (1) Baseline PM total (Produktherstellung domestic after export)
    # -------------------------
    PM_domestic_base = _safe_sum(
        baseline_out7.get("PM_domestic_after_export", np.nan),
        default=np.nan
    )
    if not np.isfinite(PM_domestic_base):
        # fallback: reconstruct baseline PM_domestic = Prim + Imp + chemsec + rPET - Exp
        Prim_b  = _safe_sum(baseline_out7.get("PrimPET_to_PM_raw", 0.0), 0.0)
        Imp_b   = _safe_sum(baseline_out7.get("product_import", 0.0), 0.0)
        Chem_b  = _safe_sum(baseline_out7.get("chemsec_to_PM_raw", baseline_out7.get("chemsec_to_PM", 0.0)), 0.0)
        rPET_b  = _safe_sum(baseline_out7.get("rPET_to_PM_total", baseline_out7.get("rPET_to_PM_raw", 0.0)), 0.0)
        Exp_b   = _safe_sum(baseline_out7.get("product_export", 0.0), 0.0)
        PM_domestic_base = max(Prim_b + Imp_b + Chem_b + rPET_b - Exp_b, 0.0)

    # -------------------------
    # (2) rPET supply & open-loop constraint
    # -------------------------
    rPET_supply = _safe_sum(out7.get("rPET_total", np.nan), default=np.nan)
    if not np.isfinite(rPET_supply):
        rPET_supply = _sum_prefix(out7, prefixes=("rPET_from_MEC", "rPET_from_CHE"))
    rPET_supply = max(rPET_supply, 0.0)
    out7["rPET_total"] = rPET_supply

    rPET_open_min = enforce_openloop_min_share * rPET_supply

    # current scenario demand cap
    rPET_to_PM_cap = _safe_sum(
        out7.get("rPET_to_PM_fixed_total",
                 out7.get("rPET_to_PM_total",
                          out7.get("rPET_to_PM_raw", 0.0))),
        default=0.0
    )
    rPET_to_PM_cap = max(rPET_to_PM_cap, 0.0)

    # apply open-loop constraint
    rPET_to_PM_eff = min(rPET_to_PM_cap, max(rPET_supply - rPET_open_min, 0.0))
    rPET_open_eff  = max(rPET_supply - rPET_to_PM_eff, 0.0)

    # mirror keys (used in Sankey/export)
    out7["rPET_to_PM_fixed_total"] = float(rPET_to_PM_eff)
    out7["rPET_to_PM_total"] = float(rPET_to_PM_eff)
    out7["rPET_to_PM_raw"] = float(rPET_to_PM_eff)
    out7["rPET_to_PM"] = float(rPET_to_PM_eff)
    out7["rPET_openloop_total"] = float(rPET_open_eff)
    out7["rPET_open"] = float(rPET_open_eff)

    # -------------------------
    # (3) HARD: Fix Produktherstellung total to baseline, solve Prim as residual
    #     PM_domestic = Prim + Import + Chemsec + rPET - Export
    # -------------------------
    Imp   = _safe_sum(out7.get("product_import", 0.0), 0.0)
    Exp   = _safe_sum(out7.get("product_export", 0.0), 0.0)
    Chem  = _safe_sum(out7.get("chemsec_to_PM_raw", out7.get("chemsec_to_PM", 0.0)), 0.0)
    rPET  = float(rPET_to_PM_eff)

    Prim_new = PM_domestic_base - Imp - Chem - rPET + Exp
    Prim_new = max(float(Prim_new), 0.0)

    out7["PrimPET_to_PM_raw"] = float(Prim_new)

    for k in ["Prim_raw", "PrimPET_production", "PrimPET_prod", "PrimaryPET_production", "PrimPET_to_PM"]:
        if k in out7:
            out7[k] = float(Prim_new)

    # -------------------------
    # (4) Enforce no inflation: lock PM_domestic_after_export to baseline
    #     and recompute PM_rest_loss consistently
    # -------------------------
    if enforce_no_inflation:
        out7["PM_domestic_after_export"] = float(PM_domestic_base)

        PM_market = _safe_sum(out7.get("PM_to_U_total", 0.0), 0.0)
        PM_rest   = PM_domestic_base - PM_market
        if PM_rest < 0 and abs(PM_rest) < 1e-6:
            PM_rest = 0.0
        out7["PM_rest_loss"] = float(max(PM_rest, 0.0))

    if Mylog:
        Mylog.info(
            "### HYBRID HARD-FIX (PM constant): "
            f"PM_domestic_base={PM_domestic_base:.3f} | "
            f"Imp={Imp:.3f} Chem={Chem:.3f} Exp={Exp:.3f} | "
            f"rPET_to_PM_eff={rPET_to_PM_eff:.3f} (open={rPET_open_eff:.3f}, open_min={rPET_open_min:.3f}) | "
            f"Prim_new={Prim_new:.3f} | "
            f"PM_to_U_total={_safe_sum(out7.get('PM_to_U_total',0.0),0.0):.3f} | PM_rest_loss={_safe_sum(out7.get('PM_rest_loss',0.0),0.0):.3f}"
        )

    return out7

# ============================================================
# Hybrid injection AFTER Step7 (rPET back to PM)
# ============================================================
def apply_step7_hybrid_injection(
    PM_primary,
    rPET_supply_total,
    rPET_eligible_to_PM_share,
    rPET_target_in_PM_share,
    Mylog=None
):

    import numpy as np

    PM_demand = np.array(PM_primary, dtype=float)  # copy
    rPET_need_for_target = PM_demand * float(rPET_target_in_PM_share)
    rPET_eligible_supply = np.array(rPET_supply_total, dtype=float) * float(rPET_eligible_to_PM_share)
    rPET_to_PM = np.minimum(rPET_need_for_target, rPET_eligible_supply)
    PM_primary_new = np.maximum(PM_demand - rPET_to_PM, 0.0)
    PM_hybrid = PM_primary_new + rPET_to_PM

    if Mylog:
        Mylog.info(
            f"### HYBRID DISPLACEMENT: "
            f"PM_demand={PM_demand.sum():.3f} | "
            f"rPET_to_PM={rPET_to_PM.sum():.3f} | "
            f"PM_primary_new={PM_primary_new.sum():.3f} | "
            f"PM_hybrid={PM_hybrid.sum():.3f}"
        )

    rPET_supply_remaining = np.maximum(np.array(rPET_supply_total, dtype=float) - rPET_to_PM, 0.0)
    return PM_primary_new, rPET_to_PM, rPET_supply_remaining


def apply_hybrid_constraints_out7(
    out7: dict,
    Mylog=None,
    enforce_no_inflation=True,
    mirror_primary_keys=None,
):
    import numpy as np
    # (unchanged body)
    def _sum_prefix(d, prefixes):
        tot = 0.0
        for k, v in d.items():
            if any(k.startswith(p) for p in prefixes):
                try:
                    tot += float(np.nansum(v))
                except Exception:
                    pass
        return tot

    def _safe_float(x, default=0.0):
        try:
            val = float(np.nansum(x))
            if not np.isfinite(val):
                return float(default)
            return val
        except Exception:
            return float(default)

    fixed_keys = [
        "PM_to_U_total",
        "PM_to_U_total_primary",
        "PM_to_U_total_hybrid",
        "PM_domestic_after_export",
        "PM_rest_loss",
        "Use_to_Waste_total",
        "DeltaStock_u_total",
        "DeltaStock_u_total_hybrid",
        "P_to_PM_total_raw",
        "P_to_PM_total_raw_sum",
    ]
    fixed_before = {}
    if enforce_no_inflation:
        for k in fixed_keys:
            if k in out7:
                fixed_before[k] = out7[k]

    rPET_supply = None
    if "rPET_total" in out7:
        rPET_supply = _safe_float(out7.get("rPET_total"), default=np.nan)
        if not np.isfinite(rPET_supply):
            rPET_supply = None
    if rPET_supply is None:
        rPET_supply = _sum_prefix(out7, prefixes=("rPET_from_MEC", "rPET_from_CHE"))

    rPET_supply = max(float(rPET_supply), 0.0)
    out7["rPET_total"] = rPET_supply

    rPET_open_min = 0.10 * rPET_supply

    rPET_to_PM_param = _safe_float(
        out7.get("rPET_to_PM_fixed_total",
                 out7.get("rPET_to_PM_raw",
                          out7.get("rPET_to_PM_total", 0.0))),
        default=0.0
    )
    rPET_to_PM_param = max(rPET_to_PM_param, 0.0)

    rPET_open_min = 0.10 * rPET_supply
    rPET_to_PM_effective = min(rPET_to_PM_param, max(rPET_supply - rPET_open_min, 0.0))
    rPET_open_effective = max(rPET_supply - rPET_to_PM_effective, 0.0)

    out7["rPET_to_PM_fixed_total"] = rPET_to_PM_effective
    out7["rPET_to_PM_raw"] = rPET_to_PM_effective
    out7["rPET_to_PM_total"] = rPET_to_PM_effective
    out7["rPET_to_PM"] = rPET_to_PM_effective

    out7["rPET_open"] = rPET_open_effective
    out7["rPET_openloop_total"] = rPET_open_effective
    out7["rPET_deficit_for_PM"] = max(rPET_to_PM_param - rPET_to_PM_effective, 0.0)

    PM_total_raw_sum = _safe_float(out7.get("P_to_PM_total_raw", 0.0), default=0.0)
    chemsec_sum = _safe_float(out7.get("chemsec_to_PM_raw", 0.0), default=0.0)

    Prim_new = max(PM_total_raw_sum - chemsec_sum - rPET_to_PM_effective, 0.0)
    out7["PrimPET_to_PM_raw"] = Prim_new

    if mirror_primary_keys is None:
        mirror_primary_keys = [
            "Prim_raw",
            "PrimPET_production",
            "PrimPET_prod",
            "PrimaryPET_production",
            "PrimPET_input",
            "PrimPET_to_PM",
        ]
    for k in mirror_primary_keys:
        if k in out7:
            out7[k] = Prim_new

    if enforce_no_inflation and fixed_before:
        for k, v in fixed_before.items():
            out7[k] = v

    if Mylog:
        Mylog.info(
            "### HYBRID CONSTRAINTS: "
            f"rPET_supply={rPET_supply:.3f} | open_min(10%)={rPET_open_min:.3f} | "
            f"rPET_to_PM_param={rPET_to_PM_param:.3f} -> effective={rPET_to_PM_effective:.3f} | "
            f"rPET_open={rPET_open_effective:.3f} | "
            f"PM_total_raw={PM_total_raw_sum:.3f} | chemsec={chemsec_sum:.3f} | Prim_new={Prim_new:.3f}"
        )

    return out7


def apply_hybrid_relative_keep_production_constant(
    out7: dict,
    baseline_out7: dict,
    Mylog=None,
    openloop_min_share=0.10,
    enforce_no_inflation=True,
):
    # (unchanged body; NOT USED)
    import numpy as np

    def _safe_sum(x, default=0.0):
        try:
            v = float(np.nansum(x))
            if not np.isfinite(v):
                return float(default)
            return v
        except Exception:
            return float(default)

    def _sum_prefix(d, prefixes):
        tot = 0.0
        for k, v in d.items():
            if any(k.startswith(p) for p in prefixes):
                try:
                    tot += float(np.nansum(v))
                except Exception:
                    pass
        return tot

    fixed_keys = [
        "PM_to_U_total",
        "PM_domestic_after_export",
        "PM_rest_loss",
        "Use_to_Waste_total",
        "DeltaStock_u_total",
        "product_import",
        "product_export",
    ]
    fixed_before = {}
    if enforce_no_inflation:
        for k in fixed_keys:
            if k in out7:
                fixed_before[k] = out7[k]

    rPET_to_PM_base = _safe_sum(
        baseline_out7.get("rPET_to_PM_fixed_total",
                          baseline_out7.get("rPET_to_PM_total",
                                            baseline_out7.get("rPET_to_PM_raw", 0.0))),
        default=0.0
    )
    rPET_to_PM_scen = _safe_sum(
        out7.get("rPET_to_PM_fixed_total",
                 out7.get("rPET_to_PM_total",
                          out7.get("rPET_to_PM_raw", 0.0))),
        default=0.0
    )

    rPET_supply = _safe_sum(out7.get("rPET_total", np.nan), default=np.nan)
    if not np.isfinite(rPET_supply):
        rPET_supply = _sum_prefix(out7, prefixes=("rPET_from_MEC", "rPET_from_CHE"))
    rPET_supply = max(rPET_supply, 0.0)
    out7["rPET_total"] = rPET_supply

    rPET_open_min = openloop_min_share * rPET_supply
    rPET_to_PM_eff = min(rPET_to_PM_scen, max(rPET_supply - rPET_open_min, 0.0))
    rPET_open_eff = max(rPET_supply - rPET_to_PM_eff, 0.0)

    out7["rPET_to_PM_fixed_total"] = rPET_to_PM_eff
    out7["rPET_to_PM_total"] = rPET_to_PM_eff
    out7["rPET_to_PM_raw"] = rPET_to_PM_eff
    out7["rPET_to_PM"] = rPET_to_PM_eff
    out7["rPET_openloop_total"] = rPET_open_eff
    out7["rPET_open"] = rPET_open_eff

    delta_rPET = rPET_to_PM_eff - rPET_to_PM_base
    P_to_PM_total_raw_base = _safe_sum(baseline_out7.get("P_to_PM_total_raw", 0.0), default=0.0)
    P_to_PM_total_raw_new = max(P_to_PM_total_raw_base - delta_rPET, 0.0)

    out7["P_to_PM_total_raw"] = P_to_PM_total_raw_new
    out7["PrimPET_to_PM_raw"] = P_to_PM_total_raw_new

    for k in ["Prim_raw", "PrimPET_production", "PrimPET_prod", "PrimaryPET_production"]:
        if k in out7:
            out7[k] = P_to_PM_total_raw_new

    if enforce_no_inflation and fixed_before:
        for k, v in fixed_before.items():
            out7[k] = v

    if Mylog:
        Mylog.info(
            "### HYBRID RELATIVE (keep Produktherstellung constant): "
            f"P_to_PM_total_raw_base={P_to_PM_total_raw_base:.3f} -> new={P_to_PM_total_raw_new:.3f} | "
            f"rPET_to_PM_base={rPET_to_PM_base:.3f} -> eff={rPET_to_PM_eff:.3f} | "
            f"delta_rPET={delta_rPET:.3f} | "
            f"rPET_supply={rPET_supply:.3f} | open_min(10%)={rPET_open_min:.3f} | open={rPET_open_eff:.3f}"
        )

    return out7

import json
from pathlib import Path

baseline_path = Path(PROJECT_DIR) / "results" / "baseline_out7.json"
with open(baseline_path, "r", encoding="utf-8") as f:
    out7_base = json.load(f)

Mylog.info(
    "### Loaded baseline snapshot: "
    f"PM_domestic={out7_base.get('PM_domestic_after_export'):.6f} | "
    f"Prim={out7_base.get('PrimPET_to_PM_raw'):.6f} | "
    f"rPET_to_PM={out7_base.get('rPET_to_PM_total'):.6f}"
)
# ============================================================
# STEP 7 – Full process chain (Fraktion-based recycling stream)
# ============================================================
def run_step7(ParameterDict, PET_MFA_System, ModelClassification, Mylog,
                  verbose=False, run_type="single", par_suffix="S1a1"):

    import numpy as np

    # -----------------------------
    # helpers
    # -----------------------------
    def _pv(PD, name: str):
        if name not in PD:
            raise KeyError(f"Missing parameter '{name}'")
        par = PD[name]
        if isinstance(par, dict) and "Values" in par:
            return par["Values"]
        if hasattr(par, "Values"):
            return par.Values
        return par

    def _p(name_base: str) -> str:
        if par_suffix:
            cand = f"{name_base}_{par_suffix}"
            if cand in ParameterDict:
                return cand
        return name_base

    def _scalar(x, default=0.0):
        try:
            if x is None:
                return float(default)
            if isinstance(x, (float, int, np.floating, np.integer)):
                return float(x)
            a = np.asarray(x, dtype=float)
            if a.size == 0:
                return float(default)
            return float(np.nansum(a))
        except Exception:
            return float(default)

    def _cls_items(key: str):
        return list(ModelClassification[key].Items)

    def _item_name(x):
        return x.Name if hasattr(x, "Name") else str(x)

    def _find_index_contains(cls_key: str, target: str):
        names = [_item_name(it) for it in _cls_items(cls_key)]
        low = [n.lower() for n in names]
        t2 = target.lower()
        for k, nm in enumerate(low):
            if t2 in nm:
                return k
        raise KeyError(f"Could not find '{target}' in '{cls_key}'. Available={names}")

    def _ensure_e_mass(A):
        A = np.asarray(A, dtype=float)
        if A.ndim in (3, 4):
            return A[..., None]
        return A

    def _ensure_e_matrix(S):
        S = np.asarray(S, dtype=float)
        if S.ndim == 4:
            return S[..., None]
        return S

    def _apply_matrix_rowonly(M_in, S, tol=1e-12, return_tensor=False):
        row_sum = S.sum(axis=3)               # (t,r,i,1)
        active = row_sum > tol

        M_active = np.where(active, M_in, 0.0)
        routed = np.einsum("trie,trije->trje", M_active, S)  # (t,r,j,1)

        M_inactive = M_in - M_active
        M_out = routed + M_inactive

        F = None
        if return_tensor:
            F = M_active[:, :, :, None, :] * S
            diag = np.zeros_like(F)
            for kk in range(M_in.shape[2]):
                diag[:, :, kk, kk, :] = M_inactive[:, :, kk, :]
            F = F + diag
        return (M_out, F) if return_tensor else M_out

    def _sum_flow(F, i_from, i_to):
        A = np.asarray(F)
        return float(A[:, :, i_from, i_to, 0].sum())

    # -----------------------------
    # STRICT routing matrix (rowsum==1 for active rows)
    # -----------------------------
    def _get_S_strict(name: str, shape_expected):
        S = _ensure_e_matrix(_pv(ParameterDict, name))
        if S.shape != shape_expected:
            raise ValueError(f"{name} expected {shape_expected} got {S.shape}")
        S = np.clip(S, 0, 1)

        rs = S.sum(axis=3)  # (t,r,i,1)
        used = rs > 1e-12
        bad = used & (np.abs(rs - 1) > 1e-8)
        if np.any(bad):
            first = np.argwhere(bad)[0].tolist()
            raise ValueError(f"{name}: rowsum != 1 for an active row. first_bad={first}")
        return S

    # -----------------------------
    # PARTIAL process output shares
    # -----------------------------
    def _get_S_partial(name: str, shape_expected):
        S = _ensure_e_matrix(_pv(ParameterDict, name))
        if S.shape != shape_expected:
            raise ValueError(f"{name} expected {shape_expected} got {S.shape}")
        S = np.clip(S, 0, 1)

        rs = S.sum(axis=3)  # (t,r,i,1)
        too_big = rs > (1.0 + 1e-8)
        if np.any(too_big):
            first = np.argwhere(too_big)[0].tolist()
            Mylog.warning(f"### WARNING {name}: rowsum > 1 for some rows. Normalizing those rows. first_bad={first}")
            rs_safe = np.where(rs > 1e-12, rs, 1.0)
            S = S / rs_safe[:, :, :, None, :]
        return S

    out = {"ok": True}

    # -----------------------------
    # 7.0 Production + product trade -> available supply for PM (raw accounting)
    # -----------------------------
    M_prod = np.asarray(_pv(ParameterDict, "p_mass_production_inputs_S1a1"), dtype=float)  # (t,r,p,q)
    if M_prod.ndim != 4:
        raise TypeError(f"p_mass_production_inputs_S1a1 must be (t,r,p,q). Got {M_prod.shape}")

    T, Rr, P, Q = M_prod.shape

    M_imp_pq = np.zeros((T, Rr, P, Q), dtype=float)
    M_exp_pq = np.zeros((T, Rr, P, Q), dtype=float)

    if "p_mass_p_imports_S1a1" in ParameterDict:
        M_imp = np.asarray(_pv(ParameterDict, "p_mass_p_imports_S1a1"), dtype=float)
        for k in range(min(P, Q)):
            M_imp_pq[:, :, k, k] = np.clip(M_imp[:, :, k], 0, None)

    if "p_mass_p_exports_S1a1" in ParameterDict:
        M_exp = np.asarray(_pv(ParameterDict, "p_mass_p_exports_S1a1"), dtype=float)
        for k in range(min(P, Q)):
            M_exp_pq[:, :, k, k] = np.clip(M_exp[:, :, k], 0, None)

    out["product_import"] = float(M_imp_pq.sum())
    out["product_export"] = float(M_exp_pq.sum())

    M_mfg_available = np.clip(M_prod + M_imp_pq - M_exp_pq, 0, None)

    # Supply components (raw, baseline)
    p_primary   = 0
    p_rpet_raw  = 2
    p_chemsec   = 3

    F_to_PM_trpq = np.zeros((T, Rr, P, Q), dtype=float)
    for ps in [p_primary, p_rpet_raw, p_chemsec]:
        if ps < P:
            F_to_PM_trpq[:, :, ps, :] = M_mfg_available[:, :, ps, :]

    out["P_to_PM_total_raw"] = float(F_to_PM_trpq.sum())
    out["PrimPET_to_PM_raw"] = float(F_to_PM_trpq[:, :, p_primary, :].sum())
    out["rPET_to_PM_raw"]    = float(F_to_PM_trpq[:, :, p_rpet_raw, :].sum())
    out["chemsec_to_PM_raw"] = float(F_to_PM_trpq[:, :, p_chemsec, :].sum())

    # -----------------------------
    # 2) PM -> Use (market output)  [THIS defines "Verbrauch (Markt)" in Sankey]
    # -----------------------------
    P_to_U_raw = np.asarray(_pv(ParameterDict, "p_mass_production_to_use_S1a1"), dtype=float)
    if P_to_U_raw.ndim == 5:
        P_to_U_raw = P_to_U_raw[..., 0]
    if P_to_U_raw.ndim != 4:
        raise TypeError(f"p_mass_production_to_use_S1a1 must be 4D. Got {P_to_U_raw.shape}")

    tmp_u = np.asarray(_pv(ParameterDict, "p_mass_use_to_collection_S1a1"), dtype=float)
    if tmp_u.ndim == 5:
        tmp_u = tmp_u[..., 0]
    U_expected = tmp_u.shape[2]

    t2, r2, a2, b2 = P_to_U_raw.shape
    if b2 == U_expected:
        P_to_U_trpu = P_to_U_raw
    elif a2 == U_expected:
        Mylog.warning("p_mass_production_to_use_S1a1 is (t,r,u,p). Swapping to (t,r,p,u).")
        P_to_U_trpu = np.swapaxes(P_to_U_raw, 2, 3)
    else:
        raise TypeError(f"Cannot infer u-axis for p_mass_production_to_use_S1a1. shape={P_to_U_raw.shape}, expected U={U_expected}")

    Use_in_tru = P_to_U_trpu.sum(axis=2)  # (t,r,u)
    out["PM_to_U_total"] = float(Use_in_tru.sum())

    # -----------------------------
    # 3) Use -> Collection (waste generation)
    # -----------------------------
    U_to_i_raw = np.asarray(_pv(ParameterDict, "p_mass_use_to_collection_S1a1"), dtype=float)
    U_to_i_trui = U_to_i_raw[..., 0] if U_to_i_raw.ndim == 5 else U_to_i_raw
    if U_to_i_trui.ndim != 4:
        raise TypeError(f"p_mass_use_to_collection_S1a1 must be (t,r,u,i). Got {U_to_i_trui.shape}")

    U_to_W_tru = U_to_i_trui.sum(axis=3)

    # align u if needed
    u_in = Use_in_tru.shape[2]
    u_out = U_to_W_tru.shape[2]
    if u_in != u_out:
        Mylog.warning(f"Use-phase mismatch: Use_in u={u_in}, Waste u={u_out}. Adjusting.")
        if u_in > u_out:
            extra = Use_in_tru[..., (u_out - 1):].sum(axis=2, keepdims=True)
            Use_in_tru = np.concatenate([Use_in_tru[..., : (u_out - 1)], extra], axis=2)
        else:
            pad = ((0, 0), (0, 0), (0, u_out - u_in))
            Use_in_tru = np.pad(Use_in_tru, pad, constant_values=0.0)

    DeltaStock_u_tru = Use_in_tru - U_to_W_tru
    out["DeltaStock_u_total"] = float(DeltaStock_u_tru.sum())
    out["SANK_in_i_from_use"] = float(U_to_i_trui.sum())

    # i inflow
    M_u_i_e = _ensure_e_mass(U_to_i_trui)
    Remain_i = M_u_i_e.sum(axis=2)  # (t,r,i,1)

    # -----------------------------
    # indices
    # -----------------------------
    I = len(_cls_items("i"))

    I_COL_G  = _find_index_contains("i", "Getrennte Sammlung")
    I_COL_M  = _find_index_contains("i", "Gemischte Sammlung")
    I_SORT_G = _find_index_contains("i", "Sortierung_getrennt")
    I_SORT_M = _find_index_contains("i", "Sortierung_gemischt")
    I_REC_G  = _find_index_contains("i", "Recycling_getrennt")
    I_REC_M  = _find_index_contains("i", "Recycling_gemischt")

    IENE = _find_index_contains("i", "Energetische Verwertung")
    ILOS = _find_index_contains("i", "Verluste")
    IDIS = _find_index_contains("i", "Beseitigung")

    IMEC_G = _find_index_contains("i", "Mechanisches Recycling (getrennt)")
    IMEC_M = _find_index_contains("i", "Mechanisches Recycling (gemischt)")
    ICHE_G = _find_index_contains("i", "Chemisches Recycling (getrennt)")
    ICHE_M = _find_index_contains("i", "Chemisches Recycling (gemischt)")

    # p indices for yields outputs
    p_rPET = _find_index_contains("p", "rPET")
    try:
        p_chemsec_out = _find_index_contains("p", "sekund")
    except KeyError:
        try:
            p_chemsec_out = _find_index_contains("p", "chem")
        except KeyError:
            p_chemsec_out = 3  # fallback

    # -----------------------------
    # routing matrices
    # -----------------------------
    shape_S = (T, Rr, I, I, 1)
    S_col      = _get_S_strict("p_share_collection_routing_S1a1", shape_S)
    S_sort_sep = _get_S_strict("p_share_sorting_separate_S1a1", shape_S)
    S_sort_mix = _get_S_strict("p_share_sorting_mixed_S1a1", shape_S)
    S_rec_proc = _get_S_strict("p_share_recycling_routing_S1a1", shape_S)
    S_proc_out = _get_S_partial("p_share_rec_process_outputs_S1a1", shape_S)

    # -----------------------------
    # collection routing
    # -----------------------------
    M1, F_col = _apply_matrix_rowonly(Remain_i, S_col, return_tensor=True)

    out.update({
        "SANK_col_g": float(np.asarray(Remain_i)[:, :, I_COL_G, 0].sum()),
        "SANK_col_m": float(np.asarray(Remain_i)[:, :, I_COL_M, 0].sum()),

        "SANK_col_g_to_sort_g": _sum_flow(F_col, I_COL_G, I_SORT_G),
        "SANK_col_g_to_sort_m": _sum_flow(F_col, I_COL_G, I_SORT_M),
        "SANK_col_g_to_ene":    _sum_flow(F_col, I_COL_G, IENE),
        "SANK_col_g_to_los":    _sum_flow(F_col, I_COL_G, ILOS),
        "SANK_col_g_to_dis":    _sum_flow(F_col, I_COL_G, IDIS),

        "SANK_col_m_to_sort_g": _sum_flow(F_col, I_COL_M, I_SORT_G),
        "SANK_col_m_to_sort_m": _sum_flow(F_col, I_COL_M, I_SORT_M),
        "SANK_col_m_to_ene":    _sum_flow(F_col, I_COL_M, IENE),
        "SANK_col_m_to_los":    _sum_flow(F_col, I_COL_M, ILOS),
        "SANK_col_m_to_dis":    _sum_flow(F_col, I_COL_M, IDIS),
    })

    # -----------------------------
    # sorting routing
    # -----------------------------
    M2a, F_sort_sep = _apply_matrix_rowonly(M1, S_sort_sep, return_tensor=True)
    M2,  F_sort_mix = _apply_matrix_rowonly(M2a, S_sort_mix, return_tensor=True)

    def _sum_both(Fa, Fb, i_from, i_to):
        return _sum_flow(Fa, i_from, i_to) + _sum_flow(Fb, i_from, i_to)

    out.update({
        "SANK_sort_g_to_rec_g": _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, I_REC_G),
        "SANK_sort_g_to_rec_m": _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, I_REC_M),
        "SANK_sort_g_to_ene":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, IENE),
        "SANK_sort_g_to_los":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, ILOS),
        "SANK_sort_g_to_dis":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, IDIS),

        "SANK_sort_m_to_rec_g": _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, I_REC_G),
        "SANK_sort_m_to_rec_m": _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, I_REC_M),
        "SANK_sort_m_to_ene":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, IENE),
        "SANK_sort_m_to_los":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, ILOS),
        "SANK_sort_m_to_dis":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, IDIS),
    })

    # -----------------------------
    # Fraktion -> Process
    # -----------------------------
    M2_snap = np.array(M2, copy=True)

    M_frac_only = np.zeros_like(M2_snap)
    M_frac_only[:, :, I_REC_G, :] = M2_snap[:, :, I_REC_G, :]
    M_frac_only[:, :, I_REC_M, :] = M2_snap[:, :, I_REC_M, :]

    _, F_frac_to_proc = _apply_matrix_rowonly(M_frac_only, S_rec_proc, return_tensor=True)

    out.update({
        "SANK_rec_g_to_MEC": _sum_flow(F_frac_to_proc, I_REC_G, IMEC_G),
        "SANK_rec_g_to_CHE": _sum_flow(F_frac_to_proc, I_REC_G, ICHE_G),
        "SANK_rec_m_to_MEC": _sum_flow(F_frac_to_proc, I_REC_M, IMEC_M),
        "SANK_rec_m_to_CHE": _sum_flow(F_frac_to_proc, I_REC_M, ICHE_M),
    })

    out["MEC_in_sep"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_G, IMEC_G, 0].sum())
    out["CHE_in_sep"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_G, ICHE_G, 0].sum())
    out["MEC_in_mix"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_M, IMEC_M, 0].sum())
    out["CHE_in_mix"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_M, ICHE_M, 0].sum())

    out["MEC_in"] = float(out["MEC_in_sep"] + out["MEC_in_mix"])
    out["CHE_in"] = float(out["CHE_in_sep"] + out["CHE_in_mix"])

    # -----------------------------
    # exports at i (optional): p_mass_i_exports
    # -----------------------------
    if "p_mass_i_exports_S1a1" in ParameterDict:
        X = np.asarray(_pv(ParameterDict, "p_mass_i_exports_S1a1"), dtype=float)
        vec = X[..., 0] if X.ndim == 4 else X  # (t,r,i)
        M_i_export_e = np.clip(vec[..., None], 0, None)  # (t,r,i,1)
    else:
        M_i_export_e = np.zeros((T, Rr, I, 1), dtype=float)

    out["SANK_MEC_export_sep"] = float(M_i_export_e[:, :, IMEC_G, 0].sum())
    out["SANK_MEC_export_mix"] = float(M_i_export_e[:, :, IMEC_M, 0].sum())
    out["SANK_CHE_export_sep"] = float(M_i_export_e[:, :, ICHE_G, 0].sum())
    out["SANK_CHE_export_mix"] = float(M_i_export_e[:, :, ICHE_M, 0].sum())

    out["SANK_MEC_export"] = float(out["SANK_MEC_export_sep"] + out["SANK_MEC_export_mix"])
    out["SANK_CHE_export"] = float(out["SANK_CHE_export_sep"] + out["SANK_CHE_export_mix"])

    # =========================================================
    # NEW (FIXED): Process outputs are controlled by SHARES, not YIELDS
    # =========================================================

    frac_prod_tr_i = np.asarray(_pv(ParameterDict, "p_share_recycling_fraction_to_process_S1a1"), dtype=float)
    if frac_prod_tr_i.ndim == 4:
        frac_prod_tr_i = frac_prod_tr_i[..., 0]  # (t,r,i)
    if frac_prod_tr_i.ndim != 3 or frac_prod_tr_i.shape[2] != I:
        raise TypeError(
            f"p_share_recycling_fraction_to_process_S1a1 must be (t,r,i). Got {frac_prod_tr_i.shape}, I={I}"
        )
    frac_prod_tr_i = np.clip(frac_prod_tr_i, 0.0, 1.0)

    def _fprod(i_idx):
        return float(frac_prod_tr_i[0, 0, i_idx])

    def _sout(i_from, j_to):
        return float(S_proc_out[0, 0, i_from, j_to, 0])

    def _get_abs_shares_for_process(i_idx):
        fprod = _fprod(i_idx)

        sev  = _sout(i_idx, IENE) if IENE is not None else 0.0
        slos = _sout(i_idx, ILOS) if ILOS is not None else 0.0
        sdis = _sout(i_idx, IDIS) if IDIS is not None else 0.0

        fprod = float(np.clip(fprod, 0.0, 1.0))
        sev   = float(np.clip(sev,   0.0, 1.0))
        slos  = float(np.clip(slos,  0.0, 1.0))
        sdis  = float(np.clip(sdis,  0.0, 1.0))

        ssum = fprod + sev + slos + sdis

        if ssum < 1.0 - 1e-8:
            sdis = sdis + (1.0 - ssum)
            ssum = fprod + sev + slos + sdis

        if abs(ssum - 1.0) > 1e-8:
            if ssum > 1e-12:
                fprod /= ssum
                sev   /= ssum
                slos  /= ssum
                sdis  /= ssum
            else:
                fprod, sev, slos, sdis = 0.0, 0.0, 1.0, 0.0

        return fprod, sev, slos, sdis

    def _apply_process_shares_export_first(proc_in, export_out, i_idx):
        proc_in = float(max(proc_in, 0.0))
        export_out = float(max(export_out, 0.0))
        export_used = min(export_out, proc_in)
        rem = proc_in - export_used

        fprod, sev, slos, sdis = _get_abs_shares_for_process(i_idx)

        prod = rem * fprod
        ev   = rem * sev
        los  = rem * slos
        dis  = rem * sdis

        if prod < 0 and abs(prod) < 1e-9: prod = 0.0
        if ev   < 0 and abs(ev)   < 1e-9: ev   = 0.0
        if los  < 0 and abs(los)  < 1e-9: los  = 0.0
        if dis  < 0 and abs(dis)  < 1e-9: dis  = 0.0

        return float(prod), float(ev), float(los), float(dis), float(rem), float(export_used)

    mec_exp_sep = float(out["SANK_MEC_export_sep"])
    mec_exp_mix = float(out["SANK_MEC_export_mix"])
    che_exp_sep = float(out["SANK_CHE_export_sep"])
    che_exp_mix = float(out["SANK_CHE_export_mix"])

    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["MEC_in_sep"], mec_exp_sep, IMEC_G
    )
    out["rPET_from_MEC_sep"] = prod
    out["EV_from_MEC_sep"], out["LOS_from_MEC_sep"], out["DIS_from_MEC_sep"] = EV, LOS, DIS
    out["REM_after_export_MEC_sep"] = REM
    out["MEC_export_used_sep"] = XUSED

    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["MEC_in_mix"], mec_exp_mix, IMEC_M
    )
    out["rPET_from_MEC_mix"] = prod
    out["EV_from_MEC_mix"], out["LOS_from_MEC_mix"], out["DIS_from_MEC_mix"] = EV, LOS, DIS
    out["REM_after_export_MEC_mix"] = REM
    out["MEC_export_used_mix"] = XUSED

    out["rPET_total"] = float(out["rPET_from_MEC_sep"] + out["rPET_from_MEC_mix"])

    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["CHE_in_sep"], che_exp_sep, ICHE_G
    )
    out["chemsec_from_CHE_sep"] = prod
    out["EV_from_CHE_sep"], out["LOS_from_CHE_sep"], out["DIS_from_CHE_sep"] = EV, LOS, DIS
    out["REM_after_export_CHE_sep"] = REM
    out["CHE_export_used_sep"] = XUSED

    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["CHE_in_mix"], che_exp_mix, ICHE_M
    )
    out["chemsec_from_CHE_mix"] = prod
    out["EV_from_CHE_mix"], out["LOS_from_CHE_mix"], out["DIS_from_CHE_mix"] = EV, LOS, DIS
    out["REM_after_export_CHE_mix"] = REM
    out["CHE_export_used_mix"] = XUSED

    out["chemsec_from_CHE_total"] = float(out["chemsec_from_CHE_sep"] + out["chemsec_from_CHE_mix"])
    out["chemsec_to_PrimPET_production_total"] = float(out["chemsec_from_CHE_total"])

    out["EV_from_MEC_total"]   = float(out["EV_from_MEC_sep"]  + out["EV_from_MEC_mix"])
    out["LOS_from_MEC_total"]  = float(out["LOS_from_MEC_sep"] + out["LOS_from_MEC_mix"])
    out["DIS_from_MEC_total"]  = float(out["DIS_from_MEC_sep"] + out["DIS_from_MEC_mix"])

    out["EV_from_CHE_total"]   = float(out["EV_from_CHE_sep"]  + out["EV_from_CHE_mix"])
    out["LOS_from_CHE_total"]  = float(out["LOS_from_CHE_sep"] + out["LOS_from_CHE_mix"])
    out["DIS_from_CHE_total"]  = float(out["DIS_from_CHE_sep"] + out["DIS_from_CHE_mix"])

    if verbose:
        def _chk(i_idx, label):
            fprod, sev, slos, sdis = _get_abs_shares_for_process(i_idx)
            Mylog.info(
                f"### SHARE CHECK {label}: fprod={fprod:.4f}, EV={sev:.4f}, LOS={slos:.4f}, DIS={sdis:.4f}, sum={fprod+sev+slos+sdis:.4f}"
            )
        _chk(IMEC_G, "MEC_getrennt")
        _chk(IMEC_M, "MEC_gemischt")
        _chk(ICHE_G, "CHE_getrennt")
        _chk(ICHE_M, "CHE_gemischt")

    # =========================================================
    # rPET routing to PM: FIXED MASS demand (bottle + otherpack)
    # =========================================================
    rpet_total = float(out["rPET_total"])

    if "p_mass_rPET_demand_bottle_S1a1" in ParameterDict:
        rpet_dem_bottle = float(max(_scalar(_pv(ParameterDict, "p_mass_rPET_demand_bottle_S1a1"), 0.0), 0.0))
    else:
        rpet_dem_bottle = 0.0
        Mylog.warning("p_mass_rPET_demand_bottle missing_S1a1 -> using 0.0")

    if "p_rPET_demand_otherpack_tr_S1a1" in ParameterDict:
        rpet_dem_other = float(max(_scalar(_pv(ParameterDict, "p_rPET_demand_otherpack_tr_S1a1"), 0.0), 0.0))
    else:
        rpet_dem_other = 0.0
        Mylog.warning("p_rPET_demand_otherpack_tr_S1a1 missing -> using 0.0")

    rpet_to_pm_fixed = float(max(rpet_dem_bottle + rpet_dem_other, 0.0))

    rpet_to_pm = float(min(rpet_total, rpet_to_pm_fixed))
    rpet_open  = float(max(rpet_total - rpet_to_pm, 0.0))
    rpet_def   = float(max(rpet_to_pm_fixed - rpet_total, 0.0))

    out["rPET_demand_bottle_fixed"] = float(rpet_dem_bottle)
    out["rPET_demand_otherpack_fixed"] = float(rpet_dem_other)
    out["rPET_to_PM_fixed_total"] = float(rpet_to_pm_fixed)

    out["rPET_to_PM_total"] = float(rpet_to_pm)
    out["rPET_openloop_total"] = float(rpet_open)
    out["rPET_surplus"] = float(rpet_open)
    out["rPET_deficit_for_PM"] = float(rpet_def)

    if "p_share_recycled_output_to_PM_S1a1" in ParameterDict:
        Mylog.warning(
            "Parameter 'p_share_recycled_output_to_PM_S1a1' is present but is NOT used for rPET->PM. "
            "It is routing/process-related, not PM demand."
        )

    prim_raw   = float(out.get("PrimPET_to_PM_raw", 0.0))
    imp_prod   = float(out.get("product_import", 0.0))
    chemsec_raw_supply = float(out.get("chemsec_to_PM_raw", 0.0))
    rpet_to_pm = float(out.get("rPET_to_PM_total", 0.0))
    exp_prod   = float(out.get("product_export", 0.0))
    PM_market  = float(out.get("PM_to_U_total", 0.0))

    PM_domestic = prim_raw + imp_prod + chemsec_raw_supply + rpet_to_pm - exp_prod
    PM_domestic = max(PM_domestic, 0.0)

    PM_rest = PM_domestic - PM_market
    if PM_rest < 0 and abs(PM_rest) < 1e-6:
        PM_rest = 0.0
    PM_rest = max(PM_rest, 0.0)

    out["PM_domestic_after_export"] = float(PM_domestic)
    out["PM_rest_loss"] = float(PM_rest)

    if out["rPET_total"] > out["MEC_in"] + 1e-6:
        Mylog.warning("### SANITY: rPET_total > MEC_in. Check yields or process inflow accounting.")
    if out["rPET_to_PM_total"] > out["rPET_total"] + 1e-6:
        Mylog.warning("### SANITY: rPET_to_PM_total > rPET_total. Check fixed demand parameters.")

    if verbose:
        Mylog.info(
            "### STEP7 CHECK: PM_domestic=%.3f | PM_market=%.3f | PM_rest=%.3f | Use_to_Waste=%.3f | Stock=%.3f | "
            "rPET_total=%.3f | rPET_to_PM=%.3f | rPET_open=%.3f | rPET_dem_bottle=%.3f | rPET_dem_other=%.3f | "
            "chemsec_to_PrimPET=%.3f | Prim_raw=%.3f | ProdImp=%.3f | ProdExp=%.3f"
            % (
                out["PM_domestic_after_export"],
                out["PM_to_U_total"],
                out["PM_rest_loss"],
                out["SANK_in_i_from_use"],
                out["DeltaStock_u_total"],
                out["rPET_total"],
                out["rPET_to_PM_total"],
                out["rPET_openloop_total"],
                out["rPET_demand_bottle_fixed"],
                out["rPET_demand_otherpack_fixed"],
                out["chemsec_to_PrimPET_production_total"],
                out["PrimPET_to_PM_raw"],
                out["product_import"],
                out["product_export"],
            )
        )

    return out


# ============================================================
# (KEEP) Other displacement helpers remain unchanged, but NOT USED
# ============================================================
def apply_hybrid_displacement_on_out7(out7, Mylog=None,
                                      rPET_target_in_PM_share=0.30,
                                      rPET_eligible_to_PM_share=1.0):
    import numpy as np
    def _sum_prefix(d, prefixes):
        tot = 0.0
        for k, v in d.items():
            if any(k.startswith(p) for p in prefixes):
                try:
                    tot += float(np.nansum(v))
                except Exception:
                    pass
        return tot

    PM_total_raw = out7.get("P_to_PM_total_raw", 0.0)
    PM_total_raw_sum = float(np.nansum(PM_total_raw))

    chemsec_raw = out7.get("chemsec_to_PM_raw", 0.0)
    chemsec_sum = float(np.nansum(chemsec_raw))

    rPET_supply_total = _sum_prefix(out7, prefixes=("rPET_from_MEC", "rPET_from_CHE"))
    if "rPET_total" in out7:
        try:
            rPET_supply_total = float(np.nansum(out7["rPET_total"]))
        except Exception:
            pass

    if not np.isfinite(PM_total_raw_sum) or PM_total_raw_sum < 0:
        PM_total_raw_sum = 0.0
    if not np.isfinite(chemsec_sum) or chemsec_sum < 0:
        chemsec_sum = 0.0
    if not np.isfinite(rPET_supply_total) or rPET_supply_total < 0:
        rPET_supply_total = 0.0

    rPET_need_target = PM_total_raw_sum * float(rPET_target_in_PM_share)
    rPET_eligible_supply = rPET_supply_total * float(rPET_eligible_to_PM_share)
    rPET_to_PM_new = min(rPET_need_target, rPET_eligible_supply)

    Prim_new = max(PM_total_raw_sum - chemsec_sum - rPET_to_PM_new, 0.0)
    rPET_open = max(rPET_supply_total - rPET_to_PM_new, 0.0)

    out7["rPET_to_PM_raw"] = rPET_to_PM_new
    out7["PrimPET_to_PM_raw"] = Prim_new
    out7["rPET_total"] = rPET_supply_total
    out7["rPET_open"] = rPET_open
    out7["rPET_to_PM_total"] = rPET_to_PM_new

    if Mylog:
        Mylog.info(
            f"### HYBRID(MC): PM_total_raw={PM_total_raw_sum:.3f} | "
            f"chemsec={chemsec_sum:.3f} | rPET_supply={rPET_supply_total:.3f} | "
            f"rPET_to_PM={rPET_to_PM_new:.3f} | Prim_new={Prim_new:.3f} | rPET_open={rPET_open:.3f}"
        )
    return out7

def apply_hybrid_displacement_on_out7_notarget(out7, Mylog=None):
    import numpy as np
    def _sum_prefix(d, prefixes):
        tot = 0.0
        for k, v in d.items():
            if any(k.startswith(p) for p in prefixes):
                try:
                    tot += float(np.nansum(v))
                except Exception:
                    pass
        return tot

    PM_total_raw = out7.get("P_to_PM_total_raw", 0.0)
    PM_total_raw_sum = float(np.nansum(PM_total_raw))
    if (not np.isfinite(PM_total_raw_sum)) or PM_total_raw_sum < 0:
        PM_total_raw_sum = 0.0

    rPET_to_PM = out7.get("rPET_to_PM_raw", out7.get("rPET_to_PM_total", 0.0))
    chemsec_raw = out7.get("chemsec_to_PM_raw", 0.0)

    rPET_to_PM_sum = float(np.nansum(rPET_to_PM))
    chemsec_sum = float(np.nansum(chemsec_raw))

    if (not np.isfinite(rPET_to_PM_sum)) or rPET_to_PM_sum < 0:
        rPET_to_PM_sum = 0.0
    if (not np.isfinite(chemsec_sum)) or chemsec_sum < 0:
        chemsec_sum = 0.0

    Prim_new_sum = max(PM_total_raw_sum - chemsec_sum - rPET_to_PM_sum, 0.0)
    out7["PrimPET_to_PM_raw"] = Prim_new_sum

    rPET_supply_total = _sum_prefix(out7, prefixes=("rPET_from_MEC", "rPET_from_CHE"))
    if "rPET_total" in out7:
        try:
            rPET_supply_total = float(np.nansum(out7["rPET_total"]))
        except Exception:
            pass

    if (not np.isfinite(rPET_supply_total)) or rPET_supply_total < 0:
        rPET_supply_total = 0.0

    out7["rPET_total"] = rPET_supply_total
    out7["rPET_open"] = max(rPET_supply_total - rPET_to_PM_sum, 0.0)
    out7["rPET_to_PM_total"] = rPET_to_PM_sum

    if Mylog:
        Mylog.info(
            f"### HYBRID(no-target): PM_total_raw={PM_total_raw_sum:.3f} | "
            f"rPET_to_PM={rPET_to_PM_sum:.3f} | chemsec={chemsec_sum:.3f} | "
            f"Prim_new={Prim_new_sum:.3f} | rPET_open={out7['rPET_open']:.3f}"
        )
    return out7


# ============================================================
# export_step7_sankey_no_new_params (UNCHANGED)
# ============================================================
def export_step7_sankey_no_new_params(
    out7,
    out_dir="outputs",
    filename="sankey_Szenario S1a1_2030_Deutschland.html",
    font_size=16,
    use_short_labels=True,
):
    import numpy as np
    from pathlib import Path
    import plotly.graph_objects as go

    def _safe(x):
        try:
            v = float(x)
            return v if v > 0 else 0.0
        except Exception:
            return 0.0

    def _safe_signed(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    def hex_to_rgba(hex_color: str, a: float = 0.35) -> str:
        h = hex_color.lstrip("#")
        r = int(h[0:2], 16); g = int(h[2:4], 16); b = int(h[4:6], 16)
        return f"rgba({r},{g},{b},{a})"

    def _split_after_export(inflow, export, ev_val, los_val, rpet_val, eps=1e-12):
        inflow = float(max(inflow, 0.0))
        export = float(max(export, 0.0))

        export_used = min(export, inflow)
        remaining = inflow - export_used

        ev_val   = float(max(ev_val, 0.0))
        los_val  = float(max(los_val, 0.0))
        rpet_val = float(max(rpet_val, 0.0))

        if inflow <= eps:
            return (0.0, 0.0, 0.0, 0.0, 0.0)

        s_ev  = ev_val / inflow
        s_los = los_val / inflow
        s_rp  = rpet_val / inflow

        s_ev  = min(max(s_ev,  0.0), 1.0)
        s_los = min(max(s_los, 0.0), 1.0)
        s_rp  = min(max(s_rp,  0.0), 1.0)

        s_sum = s_ev + s_los + s_rp
        if s_sum <= eps:
            return (export_used, 0.0, 0.0, 0.0, remaining)

        s_ev  /= s_sum
        s_los /= s_sum
        s_rp  /= s_sum

        EV  = remaining * s_ev
        LOS = remaining * s_los
        RP  = remaining * s_rp

        residual = max(remaining - (EV + LOS + RP), 0.0)
        if residual < 1e-9:
            residual = 0.0

        return (export_used, EV, LOS, RP, residual)

    if use_short_labels:
        SEP_TAG  = "Pfandflaschen"
        MIX_TAG  = "andere PET-Verp."
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"
    else:
        SEP_TAG  = "PET-Getränkeflaschen aus Pfand"
        MIX_TAG  = "andere PET-Verpackungen"
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"

    N_PRIM   = "Produktion von Primär-PET"
    N_IMP    = "Importe (Produkte)"
    N_CHESEC = "Sekundäre Rohstoffe (CHE) / sonstige Inputs"
    N_RPET   = "rPET (Pool)"
    N_PM     = "Produktherstellung"
    N_EXP    = "Exporte (Produkte)"
    N_PMREST = "Restabgänge (Produktherstellung) – Export/Bestand/Sonstiges"
    N_MARKET = "Verbrauch (Markt)"
    N_SAMM   = "Sammlung Fraktion"
    N_STOCK  = "Bestand + Unbekannter Anteil"

    N_COL_SEP = f"{SAMM_TAG}_{SEP_TAG}"
    N_COL_MIX = f"{SAMM_TAG}_{MIX_TAG}"

    N_SORT_SEP = f"{SORT_TAG}_{SEP_TAG}"
    N_SORT_MIX = f"{SORT_TAG}_{MIX_TAG}"

    N_RECFR_SEP = f"{REC_TAG}_{SEP_TAG}"
    N_RECFR_MIX = f"{REC_TAG}_{MIX_TAG}"

    N_EV_COL_SEP  = f"EV ({SAMM_TAG}_{SEP_TAG})"
    N_LOS_COL_SEP = f"Verluste ({SAMM_TAG}_{SEP_TAG})"
    N_DIS_COL_SEP = f"Beseitigung ({SAMM_TAG}_{SEP_TAG})"

    N_EV_COL_MIX  = f"EV ({SAMM_TAG}_{MIX_TAG})"
    N_LOS_COL_MIX = f"Verluste ({SAMM_TAG}_{MIX_TAG})"
    N_DIS_COL_MIX = f"Beseitigung ({SAMM_TAG}_{MIX_TAG})"

    N_EV_SORT_SEP  = f"EV ({SORT_TAG}_{SEP_TAG})"
    N_LOS_SORT_SEP = f"Verluste ({SORT_TAG}_{SEP_TAG})"
    N_DIS_SORT_SEP = f"Beseitigung ({SORT_TAG}_{SEP_TAG})"

    N_EV_SORT_MIX  = f"EV ({SORT_TAG}_{MIX_TAG})"
    N_LOS_SORT_MIX = f"Verluste ({SORT_TAG}_{MIX_TAG})"
    N_DIS_SORT_MIX = f"Beseitigung ({SORT_TAG}_{MIX_TAG})"

    N_MEC_SEP = f"Mechanisches Recycling (MEC, {SEP_TAG})"
    N_MEC_MIX = f"Mechanisches Recycling (MEC, {MIX_TAG})"
    N_CHE_SEP = f"Chemisches Recycling (CHE, {SEP_TAG})"
    N_CHE_MIX = f"Chemisches Recycling (CHE, {MIX_TAG})"

    N_EV_MEC_SEP  = f"EV (MEC, {SEP_TAG})"
    N_LOS_MEC_SEP = f"Verluste (MEC, {SEP_TAG})"
    N_MEC_OTH_SEP = f"MEC-Rückstände / Sonstiges ({SEP_TAG})"

    N_EV_MEC_MIX  = f"EV (MEC, {MIX_TAG})"
    N_LOS_MEC_MIX = f"Verluste (MEC, {MIX_TAG})"
    N_MEC_OTH_MIX = f"MEC-Rückstände / Sonstiges ({MIX_TAG})"

    N_EV_CHE_SEP   = f"EV (CHE, {SEP_TAG})"
    N_LOS_CHE_SEP  = f"Verluste (CHE, {SEP_TAG})"
    N_CHE_PROD_SEP = f"CHE-Produkte / open-loop ({SEP_TAG})"

    N_EV_CHE_MIX   = f"EV (CHE, {MIX_TAG})"
    N_LOS_CHE_MIX  = f"Verluste (CHE, {MIX_TAG})"
    N_CHE_PROD_MIX = f"CHE-Produkte / open-loop ({MIX_TAG})"

    N_RPET_OPEN  = "rPET open-loop"
    N_REC_EXPORT = "Exporte (Recyclingprodukte)"

    PM_market   = _safe(out7.get("PM_to_U_total", 0.0))
    waste_total = _safe(out7.get("SANK_in_i_from_use", 0.0))
    stock_delta = _safe_signed(PM_market - waste_total)

    imp_prod = _safe(out7.get("product_import", 0.0))
    exp_prod = _safe(out7.get("product_export", 0.0))

    prim_raw = _safe(out7.get("PrimPET_to_PM_raw", 0.0))
    chemsec  = _safe(out7.get("chemsec_to_PM", out7.get("chemsec_to_PM_raw", 0.0)))
    rPET_to_PM_raw = _safe(out7.get("rPET_to_PM_total", 0.0))

    PM_domestic = _safe(out7.get(
        "PM_domestic_after_export",
        max(prim_raw + imp_prod + chemsec + rPET_to_PM_raw - exp_prod, 0.0)
    ))
    PM_rest = _safe(out7.get(
        "PM_rest_loss",
        max(PM_domestic - PM_market, 0.0)
    ))

    col_sep = _safe(out7.get("SANK_col_g", 0.0))
    col_mix = _safe(out7.get("SANK_col_m", 0.0))

    col_sep_to_sort_sep = _safe(out7.get("SANK_col_g_to_sort_g", 0.0))
    col_sep_to_sort_mix = _safe(out7.get("SANK_col_g_to_sort_m", 0.0))
    col_sep_to_ene      = _safe(out7.get("SANK_col_g_to_ene", 0.0))
    col_sep_to_los      = _safe(out7.get("SANK_col_g_to_los", 0.0))
    col_sep_to_dis      = _safe(out7.get("SANK_col_g_to_dis", 0.0))

    col_mix_to_sort_sep = _safe(out7.get("SANK_col_m_to_sort_g", 0.0))
    col_mix_to_sort_mix = _safe(out7.get("SANK_col_m_to_sort_m", 0.0))
    col_mix_to_ene      = _safe(out7.get("SANK_col_m_to_ene", 0.0))
    col_mix_to_los      = _safe(out7.get("SANK_col_m_to_los", 0.0))
    col_mix_to_dis      = _safe(out7.get("SANK_col_m_to_dis", 0.0))

    sort_sep_to_rec_sep = _safe(out7.get("SANK_sort_g_to_rec_g", 0.0))
    sort_sep_to_rec_mix = _safe(out7.get("SANK_sort_g_to_rec_m", 0.0))
    sort_sep_to_ene     = _safe(out7.get("SANK_sort_g_to_ene", 0.0))
    sort_sep_to_los     = _safe(out7.get("SANK_sort_g_to_los", 0.0))
    sort_sep_to_dis     = _safe(out7.get("SANK_sort_g_to_dis", 0.0))

    sort_mix_to_rec_sep = _safe(out7.get("SANK_sort_m_to_rec_g", 0.0))
    sort_mix_to_rec_mix = _safe(out7.get("SANK_sort_m_to_rec_m", 0.0))
    sort_mix_to_ene     = _safe(out7.get("SANK_sort_m_to_ene", 0.0))
    sort_mix_to_los     = _safe(out7.get("SANK_sort_m_to_los", 0.0))
    sort_mix_to_dis     = _safe(out7.get("SANK_sort_m_to_dis", 0.0))

    rec_sep_to_MEC = _safe(out7.get("SANK_rec_g_to_MEC", 0.0))
    rec_sep_to_CHE = _safe(out7.get("SANK_rec_g_to_CHE", 0.0))
    rec_mix_to_MEC = _safe(out7.get("SANK_rec_m_to_MEC", 0.0))
    rec_mix_to_CHE = _safe(out7.get("SANK_rec_m_to_CHE", 0.0))

    MEC_in_sep = _safe(out7.get("MEC_in_sep", rec_sep_to_MEC))
    MEC_in_mix = _safe(out7.get("MEC_in_mix", rec_mix_to_MEC))
    CHE_in_sep = _safe(out7.get("CHE_in_sep", rec_sep_to_CHE))
    CHE_in_mix = _safe(out7.get("CHE_in_mix", rec_mix_to_CHE))

    MEC_export_sep = _safe(out7.get("SANK_MEC_export_sep", 0.0))
    MEC_export_mix = _safe(out7.get("SANK_MEC_export_mix", 0.0))
    CHE_export_sep = _safe(out7.get("SANK_CHE_export_sep", 0.0))
    CHE_export_mix = _safe(out7.get("SANK_CHE_export_mix", 0.0))

    EV_MEC_sep_raw  = _safe(out7.get("EV_from_MEC_sep", 0.0))
    LOS_MEC_sep_raw = _safe(out7.get("LOS_from_MEC_sep", 0.0))
    EV_MEC_mix_raw  = _safe(out7.get("EV_from_MEC_mix", 0.0))
    LOS_MEC_mix_raw = _safe(out7.get("LOS_from_MEC_mix", 0.0))

    EV_CHE_sep_raw  = _safe(out7.get("EV_from_CHE_sep", 0.0))
    LOS_CHE_sep_raw = _safe(out7.get("LOS_from_CHE_sep", 0.0))
    EV_CHE_mix_raw  = _safe(out7.get("EV_from_CHE_mix", 0.0))
    LOS_CHE_mix_raw = _safe(out7.get("LOS_from_CHE_mix", 0.0))

    rPET_from_MEC_sep_raw = _safe(out7.get("rPET_from_MEC_sep", 0.0))
    rPET_from_MEC_mix_raw = _safe(out7.get("rPET_from_MEC_mix", 0.0))
    rPET_from_CHE_sep_raw = _safe(out7.get("rPET_from_CHE_sep", 0.0))
    rPET_from_CHE_mix_raw = _safe(out7.get("rPET_from_CHE_mix", 0.0))

    MEC_export_sep_u, EV_MEC_sep, LOS_MEC_sep, rPET_from_MEC_sep, MEC_other_sep = _split_after_export(
        MEC_in_sep, MEC_export_sep, EV_MEC_sep_raw, LOS_MEC_sep_raw, rPET_from_MEC_sep_raw
    )
    MEC_export_mix_u, EV_MEC_mix, LOS_MEC_mix, rPET_from_MEC_mix, MEC_other_mix = _split_after_export(
        MEC_in_mix, MEC_export_mix, EV_MEC_mix_raw, LOS_MEC_mix_raw, rPET_from_MEC_mix_raw
    )

    CHE_export_sep_u = min(CHE_export_sep, CHE_in_sep)
    CHE_export_mix_u = min(CHE_export_mix, CHE_in_mix)

    rem_che_sep = max(CHE_in_sep - CHE_export_sep_u, 0.0)
    rem_che_mix = max(CHE_in_mix - CHE_export_mix_u, 0.0)

    EV_CHE_sep = min(EV_CHE_sep_raw, rem_che_sep)
    EV_CHE_mix = min(EV_CHE_mix_raw, rem_che_mix)

    LOS_CHE_sep = 0.0
    LOS_CHE_mix = 0.0

    rPET_from_CHE_sep = 0.0
    rPET_from_CHE_mix = 0.0

    CHE_products_sep = max(rem_che_sep - EV_CHE_sep, 0.0)
    CHE_products_mix = max(rem_che_mix - EV_CHE_mix, 0.0)

    rPET_total = float(rPET_from_MEC_sep + rPET_from_MEC_mix + rPET_from_CHE_sep + rPET_from_CHE_mix)
    rPET_to_PM = min(float(rPET_to_PM_raw), float(rPET_total))
    rPET_open  = max(float(rPET_total) - float(rPET_to_PM), 0.0)

    nodes = [
        N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM, N_EXP, N_PMREST, N_MARKET, N_SAMM, N_STOCK,
        N_COL_SEP, N_COL_MIX,
        N_SORT_SEP, N_SORT_MIX,
        N_RECFR_SEP, N_RECFR_MIX,
        N_EV_COL_SEP, N_LOS_COL_SEP, N_DIS_COL_SEP,
        N_EV_COL_MIX, N_LOS_COL_MIX, N_DIS_COL_MIX,
        N_EV_SORT_SEP, N_LOS_SORT_SEP, N_DIS_SORT_SEP,
        N_EV_SORT_MIX, N_LOS_SORT_MIX, N_DIS_SORT_MIX,
        N_MEC_SEP, N_MEC_MIX, N_CHE_SEP, N_CHE_MIX,
        N_EV_MEC_SEP, N_LOS_MEC_SEP, N_MEC_OTH_SEP,
        N_EV_MEC_MIX, N_LOS_MEC_MIX, N_MEC_OTH_MIX,
        N_EV_CHE_SEP, N_LOS_CHE_SEP, N_CHE_PROD_SEP,
        N_EV_CHE_MIX, N_LOS_CHE_MIX, N_CHE_PROD_MIX,
        N_RPET_OPEN, N_REC_EXPORT
    ]
    idx = {n: i for i, n in enumerate(nodes)}

    COL = {
        "prim_pm": "#1C4E80",
        "import":  "#5A7FB5",
        "export":  "#2E3440",
        "market":  "#4C566A",
        "sammlung":"#2F8F9D",
        "sort":    "#6B8CAF",
        "recfr":   "#3E6D9C",
        "mec":     "#2C6E9B",
        "che":     "#6A5D8C",
        "ev":      "#1F6F78",
        "loss":    "#111827",
        "dis":     "#374151",
        "rpet":    "#D4A106",
        "stock":   "#D1D5DB",
    }

    def node_color(name: str) -> str:
        if name in (N_PRIM, N_PM):
            return COL["prim_pm"]
        if name == N_IMP:
            return COL["import"]
        if name in (N_EXP, N_REC_EXPORT):
            return COL["export"]
        if name == N_MARKET:
            return COL["market"]
        if name in (N_SAMM, N_COL_SEP, N_COL_MIX):
            return COL["sammlung"]
        if name in (N_SORT_SEP, N_SORT_MIX):
            return COL["sort"]
        if name in (N_RECFR_SEP, N_RECFR_MIX):
            return COL["recfr"]
        if name in (N_MEC_SEP, N_MEC_MIX, N_MEC_OTH_SEP, N_MEC_OTH_MIX):
            return COL["mec"]
        if name in (N_CHE_SEP, N_CHE_MIX, N_CHE_PROD_SEP, N_CHE_PROD_MIX, N_CHESEC):
            return COL["che"]
        if name.startswith("EV"):
            return COL["ev"]
        if name.startswith("Verluste"):
            return COL["loss"]
        if name.startswith("Beseitigung"):
            return COL["dis"]
        if name in (N_RPET, N_RPET_OPEN):
            return COL["rpet"]
        if name in (N_PMREST, N_STOCK):
            return COL["stock"]
        return COL["stock"]

    node_colors = [node_color(n) for n in nodes]

    src, trg, val, link_text, link_color = [], [], [], [], []

    def L(a, b, v):
        v = float(v)
        if v <= 0:
            return
        if (a not in idx) or (b not in idx):
            missing = []
            if a not in idx: missing.append(a)
            if b not in idx: missing.append(b)
            raise KeyError(f"Sankey node name not found: {missing}")

        src.append(idx[a])
        trg.append(idx[b])
        val.append(v)
        link_text.append(f"{v:.3f} kt")
        link_color.append(hex_to_rgba(node_color(b), 0.35))

    L(N_PRIM,   N_PM, prim_raw)
    L(N_IMP,    N_PM, imp_prod)
    L(N_CHESEC, N_PM, chemsec)
    L(N_RPET,   N_PM, rPET_to_PM)

    L(N_PM, N_EXP,    exp_prod)
    L(N_PM, N_MARKET, PM_market)
    L(N_PM, N_PMREST, PM_rest)

    L(N_MARKET, N_SAMM, waste_total)
    if stock_delta > 0:
        L(N_MARKET, N_STOCK, stock_delta)

    L(N_SAMM, N_COL_SEP, col_sep)
    L(N_SAMM, N_COL_MIX, col_mix)

    L(N_COL_SEP, N_SORT_SEP,    col_sep_to_sort_sep)
    L(N_COL_SEP, N_SORT_MIX,    col_sep_to_sort_mix)
    L(N_COL_SEP, N_EV_COL_SEP,  col_sep_to_ene)
    L(N_COL_SEP, N_LOS_COL_SEP, col_sep_to_los)
    L(N_COL_SEP, N_DIS_COL_SEP, col_sep_to_dis)

    L(N_COL_MIX, N_SORT_SEP,    col_mix_to_sort_sep)
    L(N_COL_MIX, N_SORT_MIX,    col_mix_to_sort_mix)
    L(N_COL_MIX, N_EV_COL_MIX,  col_mix_to_ene)
    L(N_COL_MIX, N_LOS_COL_MIX, col_mix_to_los)
    L(N_COL_MIX, N_DIS_COL_MIX, col_mix_to_dis)

    L(N_SORT_SEP,  N_RECFR_SEP,    sort_sep_to_rec_sep)
    L(N_SORT_SEP,  N_RECFR_MIX,    sort_sep_to_rec_mix)
    L(N_SORT_SEP,  N_EV_SORT_SEP,  sort_sep_to_ene)
    L(N_SORT_SEP,  N_LOS_SORT_SEP, sort_sep_to_los)
    L(N_SORT_SEP,  N_DIS_SORT_SEP, sort_sep_to_dis)

    L(N_SORT_MIX,  N_RECFR_SEP,    sort_mix_to_rec_sep)
    L(N_SORT_MIX,  N_RECFR_MIX,    sort_mix_to_rec_mix)
    L(N_SORT_MIX,  N_EV_SORT_MIX,  sort_mix_to_ene)
    L(N_SORT_MIX,  N_LOS_SORT_MIX, sort_mix_to_los)
    L(N_SORT_MIX,  N_DIS_SORT_MIX, sort_mix_to_dis)

    L(N_RECFR_SEP, N_MEC_SEP, rec_sep_to_MEC)
    L(N_RECFR_SEP, N_CHE_SEP, rec_sep_to_CHE)
    L(N_RECFR_MIX, N_MEC_MIX, rec_mix_to_MEC)
    L(N_RECFR_MIX, N_CHE_MIX, rec_mix_to_CHE)

    L(N_MEC_SEP, N_REC_EXPORT,  MEC_export_sep_u)
    L(N_MEC_SEP, N_RPET,        rPET_from_MEC_sep)
    L(N_MEC_SEP, N_EV_MEC_SEP,  EV_MEC_sep)
    L(N_MEC_SEP, N_LOS_MEC_SEP, LOS_MEC_sep)
    L(N_MEC_SEP, N_MEC_OTH_SEP, MEC_other_sep)

    L(N_MEC_MIX, N_REC_EXPORT,  MEC_export_mix_u)
    L(N_MEC_MIX, N_RPET,        rPET_from_MEC_mix)
    L(N_MEC_MIX, N_EV_MEC_MIX,  EV_MEC_mix)
    L(N_MEC_MIX, N_LOS_MEC_MIX, LOS_MEC_mix)
    L(N_MEC_MIX, N_MEC_OTH_MIX, MEC_other_mix)

    L(N_CHE_SEP, N_REC_EXPORT,   CHE_export_sep_u)
    L(N_CHE_SEP, N_RPET,         rPET_from_CHE_sep)
    L(N_CHE_SEP, N_EV_CHE_SEP,   EV_CHE_sep)
    L(N_CHE_SEP, N_LOS_CHE_SEP,  LOS_CHE_sep)
    L(N_CHE_SEP, N_CHE_PROD_SEP, CHE_products_sep)

    L(N_CHE_MIX, N_REC_EXPORT,   CHE_export_mix_u)
    L(N_CHE_MIX, N_RPET,         rPET_from_CHE_mix)
    L(N_CHE_MIX, N_EV_CHE_MIX,   EV_CHE_mix)
    L(N_CHE_MIX, N_LOS_CHE_MIX,  LOS_CHE_mix)
    L(N_CHE_MIX, N_CHE_PROD_MIX, CHE_products_mix)

    L(N_RPET, N_RPET_OPEN, rPET_open)

    Nn = len(nodes)
    infl = np.zeros(Nn); outf = np.zeros(Nn)
    for s, t, v in zip(src, trg, val):
        outf[s] += v
        infl[t] += v

    totals = np.where(outf > 0, outf, infl)
    totals[idx[N_PM]]   = float(PM_domestic)
    totals[idx[N_RPET]] = float(rPET_total)

    labels = []
    for i, name in enumerate(nodes):
        if name == N_RPET:
            labels.append(
                f"{name}\n({totals[i]:.3f} kt)\n[rPET_total={rPET_total:.3f} | toPM={rPET_to_PM:.3f} | open={rPET_open:.3f}]"
            )
        else:
            labels.append(f"{name}\n({totals[i]:.3f} kt)")

    fig = go.Figure(go.Sankey(
        arrangement="freeform",
        node=dict(
            label=labels,
            pad=18,
            thickness=20,
            color=node_colors,
            line=dict(color="rgba(0,0,0,0.25)", width=0.5),
        ),
        link=dict(
            source=src,
            target=trg,
            value=val,
            color=link_color,
            customdata=link_text,
            hovertemplate="%{source.label} → %{target.label}<br>%{customdata}<extra></extra>",
        ),
    ))
    fig.update_layout(
        title="PET/Polymer Verpackungen – Deutschland 2030, MFA Szenario S1-a1 (kt)",
        font=dict(size=font_size),
    )

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    html_file = out_path / filename
    fig.write_html(str(html_file), config={"displayModeBar": True})

    return str(html_file.resolve())


def export_step7_nodes_links_excel_S1a1(
    out7,
    out_dir="outputs",
    excel_filename="nodes_links_S1a1.xlsx",
    use_short_labels=True,
):
    import numpy as np
    import pandas as pd
    from pathlib import Path

    def _safe(x):
        try:
            v = float(x)
            return v if np.isfinite(v) and v > 0 else 0.0
        except Exception:
            return 0.0

    def _safe_signed(x):
        try:
            v = float(x)
            return v if np.isfinite(v) else 0.0
        except Exception:
            return 0.0

    # ----------------------------
    # Labels (S1a1 = no CHE internal Depoly/Pyro)
    # ----------------------------
    if use_short_labels:
        SEP_TAG  = "Pfandflaschen"
        MIX_TAG  = "andere PET-Verp."
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"
    else:
        SEP_TAG  = "PET-Getränkeflaschen aus Pfand"
        MIX_TAG  = "andere PET-Verpackungen"
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"

    N_PRIM   = "Produktion von Primär-PET"
    N_IMP    = "Importe (Produkte)"
    N_CHESEC = "Sekundär-Rohstoffe"
    N_RPET   = "rPET (Pool)"
    N_PM     = "Produktherstellung"
    N_EXP    = "Exporte (Produkte)"
    N_PMREST = "Restabgänge (Produktherstellung) – Export/Bestand/Sonstiges"
    N_MARKET = "Verbrauch (Markt)"
    N_SAMM   = "Sammlung Fraktion"
    N_STOCK  = "Bestand + Unbekannter Anteil"

    N_COL_SEP = f"{SAMM_TAG}_{SEP_TAG}"
    N_COL_MIX = f"{SAMM_TAG}_{MIX_TAG}"

    N_SORT_SEP = f"{SORT_TAG}_{SEP_TAG}"
    N_SORT_MIX = f"{SORT_TAG}_{MIX_TAG}"

    N_RECFR_SEP = f"{REC_TAG}_{SEP_TAG}"
    N_RECFR_MIX = f"{REC_TAG}_{MIX_TAG}"

    N_EV_COL_SEP  = f"EV ({SAMM_TAG}_{SEP_TAG})"
    N_LOS_COL_SEP = f"Verluste ({SAMM_TAG}_{SEP_TAG})"
    N_DIS_COL_SEP = f"Beseitigung ({SAMM_TAG}_{SEP_TAG})"

    N_EV_COL_MIX  = f"EV ({SAMM_TAG}_{MIX_TAG})"
    N_LOS_COL_MIX = f"Verluste ({SAMM_TAG}_{MIX_TAG})"
    N_DIS_COL_MIX = f"Beseitigung ({SAMM_TAG}_{MIX_TAG})"

    N_EV_SORT_SEP  = f"EV ({SORT_TAG}_{SEP_TAG})"
    N_LOS_SORT_SEP = f"Verluste ({SORT_TAG}_{SEP_TAG})"
    N_DIS_SORT_SEP = f"Beseitigung ({SORT_TAG}_{SEP_TAG})"

    N_EV_SORT_MIX  = f"EV ({SORT_TAG}_{MIX_TAG})"
    N_LOS_SORT_MIX = f"Verluste ({SORT_TAG}_{MIX_TAG})"
    N_DIS_SORT_MIX = f"Beseitigung ({SORT_TAG}_{MIX_TAG})"

    N_MEC_SEP = f"Mechanisches Recycling (MEC, {SEP_TAG})"
    N_MEC_MIX = f"Mechanisches Recycling (MEC, {MIX_TAG})"
    N_CHE_SEP = f"Chemisches Recycling (CHE, {SEP_TAG})"
    N_CHE_MIX = f"Chemisches Recycling (CHE, {MIX_TAG})"

    N_EV_MEC_SEP  = f"EV (MEC, {SEP_TAG})"
    N_LOS_MEC_SEP = f"Verluste (MEC, {SEP_TAG})"
    N_MEC_OTH_SEP = f"MEC-Rückstände / Sonstiges ({SEP_TAG})"

    N_EV_MEC_MIX  = f"EV (MEC, {MIX_TAG})"
    N_LOS_MEC_MIX = f"Verluste (MEC, {MIX_TAG})"
    N_MEC_OTH_MIX = f"MEC-Rückstände / Sonstiges ({MIX_TAG})"

    N_EV_CHE_SEP  = f"EV (CHE, {SEP_TAG})"
    N_LOS_CHE_SEP = f"Verluste (CHE, {SEP_TAG})"
    N_CHE_PROD_SEP = f"CHE-Produkte / open-loop ({SEP_TAG})"

    N_EV_CHE_MIX  = f"EV (CHE, {MIX_TAG})"
    N_LOS_CHE_MIX = f"Verluste (CHE, {MIX_TAG})"
    N_CHE_PROD_MIX = f"CHE-Produkte / open-loop ({MIX_TAG})"

    N_RPET_OPEN  = "rPET open-loop"
    N_REC_EXPORT = "Exporte (Recyclingprodukte)"

    # ----------------------------
    # Values from out7 (same as Sankey base)
    # ----------------------------
    PM_market   = _safe(out7.get("PM_to_U_total", 0.0))
    waste_total = _safe(out7.get("SANK_in_i_from_use", 0.0))
    stock_delta = _safe_signed(PM_market - waste_total)

    imp_prod = _safe(out7.get("product_import", 0.0))
    exp_prod = _safe(out7.get("product_export", 0.0))

    prim_raw = _safe(out7.get("PrimPET_to_PM_raw", 0.0))
    chemsec  = _safe(out7.get("chemsec_to_PM", out7.get("chemsec_to_PM_raw", 0.0)))
    rPET_to_PM = _safe(out7.get("rPET_to_PM_total", 0.0))

    PM_domestic = _safe(out7.get(
        "PM_domestic_after_export",
        max(prim_raw + imp_prod + chemsec + rPET_to_PM - exp_prod, 0.0)
    ))
    PM_rest = _safe(out7.get("PM_rest_loss", max(PM_domestic - PM_market, 0.0)))

    col_sep = _safe(out7.get("SANK_col_g", 0.0))
    col_mix = _safe(out7.get("SANK_col_m", 0.0))

    col_sep_to_sort_sep = _safe(out7.get("SANK_col_g_to_sort_g", 0.0))
    col_sep_to_sort_mix = _safe(out7.get("SANK_col_g_to_sort_m", 0.0))
    col_sep_to_ene      = _safe(out7.get("SANK_col_g_to_ene", 0.0))
    col_sep_to_los      = _safe(out7.get("SANK_col_g_to_los", 0.0))
    col_sep_to_dis      = _safe(out7.get("SANK_col_g_to_dis", 0.0))

    col_mix_to_sort_sep = _safe(out7.get("SANK_col_m_to_sort_g", 0.0))
    col_mix_to_sort_mix = _safe(out7.get("SANK_col_m_to_sort_m", 0.0))
    col_mix_to_ene      = _safe(out7.get("SANK_col_m_to_ene", 0.0))
    col_mix_to_los      = _safe(out7.get("SANK_col_m_to_los", 0.0))
    col_mix_to_dis      = _safe(out7.get("SANK_col_m_to_dis", 0.0))

    sort_sep_to_rec_sep = _safe(out7.get("SANK_sort_g_to_rec_g", 0.0))
    sort_sep_to_rec_mix = _safe(out7.get("SANK_sort_g_to_rec_m", 0.0))
    sort_sep_to_ene     = _safe(out7.get("SANK_sort_g_to_ene", 0.0))
    sort_sep_to_los     = _safe(out7.get("SANK_sort_g_to_los", 0.0))
    sort_sep_to_dis     = _safe(out7.get("SANK_sort_g_to_dis", 0.0))

    sort_mix_to_rec_sep = _safe(out7.get("SANK_sort_m_to_rec_g", 0.0))
    sort_mix_to_rec_mix = _safe(out7.get("SANK_sort_m_to_rec_m", 0.0))
    sort_mix_to_ene     = _safe(out7.get("SANK_sort_m_to_ene", 0.0))
    sort_mix_to_los     = _safe(out7.get("SANK_sort_m_to_los", 0.0))
    sort_mix_to_dis     = _safe(out7.get("SANK_sort_m_to_dis", 0.0))

    rec_sep_to_MEC = _safe(out7.get("SANK_rec_g_to_MEC", 0.0))
    rec_sep_to_CHE = _safe(out7.get("SANK_rec_g_to_CHE", 0.0))
    rec_mix_to_MEC = _safe(out7.get("SANK_rec_m_to_MEC", 0.0))
    rec_mix_to_CHE = _safe(out7.get("SANK_rec_m_to_CHE", 0.0))

    MEC_in_sep = _safe(out7.get("MEC_in_sep", rec_sep_to_MEC))
    MEC_in_mix = _safe(out7.get("MEC_in_mix", rec_mix_to_MEC))
    CHE_in_sep = _safe(out7.get("CHE_in_sep", rec_sep_to_CHE))
    CHE_in_mix = _safe(out7.get("CHE_in_mix", rec_mix_to_CHE))

    MEC_export_sep = _safe(out7.get("SANK_MEC_export_sep", 0.0))
    MEC_export_mix = _safe(out7.get("SANK_MEC_export_mix", 0.0))
    CHE_export_sep = _safe(out7.get("SANK_CHE_export_sep", 0.0))
    CHE_export_mix = _safe(out7.get("SANK_CHE_export_mix", 0.0))

    # MEC outputs (post Step7)
    EV_MEC_sep   = _safe(out7.get("EV_from_MEC_sep", 0.0))
    LOS_MEC_sep  = _safe(out7.get("LOS_from_MEC_sep", 0.0))
    rPET_MEC_sep = _safe(out7.get("rPET_from_MEC_sep", 0.0))

    EV_MEC_mix   = _safe(out7.get("EV_from_MEC_mix", 0.0))
    LOS_MEC_mix  = _safe(out7.get("LOS_from_MEC_mix", 0.0))
    rPET_MEC_mix = _safe(out7.get("rPET_from_MEC_mix", 0.0))

    REM_MEC_sep = _safe(out7.get("REM_after_export_MEC_sep", max(MEC_in_sep - min(MEC_export_sep, MEC_in_sep), 0.0)))
    REM_MEC_mix = _safe(out7.get("REM_after_export_MEC_mix", max(MEC_in_mix - min(MEC_export_mix, MEC_in_mix), 0.0)))
    MEC_OTH_sep = max(REM_MEC_sep - (EV_MEC_sep + LOS_MEC_sep + rPET_MEC_sep), 0.0)
    MEC_OTH_mix = max(REM_MEC_mix - (EV_MEC_mix + LOS_MEC_mix + rPET_MEC_mix), 0.0)

    MEC_export_sep_u = min(MEC_export_sep, MEC_in_sep)
    MEC_export_mix_u = min(MEC_export_mix, MEC_in_mix)
    CHE_export_sep_u = min(CHE_export_sep, CHE_in_sep)
    CHE_export_mix_u = min(CHE_export_mix, CHE_in_mix)

    # CHE outputs (S1a1 baseline-style keys)
    EV_CHE_sep   = _safe(out7.get("EV_from_CHE_sep", 0.0))
    LOS_CHE_sep  = _safe(out7.get("LOS_from_CHE_sep", 0.0))
    DIS_CHE_sep  = _safe(out7.get("DIS_from_CHE_sep", 0.0))
    PROD_CHE_sep = _safe(out7.get("chemsec_from_CHE_sep", out7.get("chemsec_from_CHE", 0.0)))

    EV_CHE_mix   = _safe(out7.get("EV_from_CHE_mix", 0.0))
    LOS_CHE_mix  = _safe(out7.get("LOS_from_CHE_mix", 0.0))
    DIS_CHE_mix  = _safe(out7.get("DIS_from_CHE_mix", 0.0))
    PROD_CHE_mix = _safe(out7.get("chemsec_from_CHE_mix", out7.get("chemsec_from_CHE", 0.0)))

    rPET_total = _safe(out7.get("rPET_total", (rPET_MEC_sep + rPET_MEC_mix)))
    rPET_open  = _safe(out7.get("rPET_openloop_total", max(rPET_total - rPET_to_PM, 0.0)))

    # ----------------------------
    # Nodes
    # ----------------------------
    nodes = [
        N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM, N_EXP, N_PMREST, N_MARKET, N_SAMM, N_STOCK,
        N_COL_SEP, N_COL_MIX,
        N_SORT_SEP, N_SORT_MIX,
        N_RECFR_SEP, N_RECFR_MIX,
        N_EV_COL_SEP, N_LOS_COL_SEP, N_DIS_COL_SEP,
        N_EV_COL_MIX, N_LOS_COL_MIX, N_DIS_COL_MIX,
        N_EV_SORT_SEP, N_LOS_SORT_SEP, N_DIS_SORT_SEP,
        N_EV_SORT_MIX, N_LOS_SORT_MIX, N_DIS_SORT_MIX,
        N_MEC_SEP, N_MEC_MIX, N_CHE_SEP, N_CHE_MIX,
        N_EV_MEC_SEP, N_LOS_MEC_SEP, N_MEC_OTH_SEP,
        N_EV_MEC_MIX, N_LOS_MEC_MIX, N_MEC_OTH_MIX,
        N_EV_CHE_SEP, N_LOS_CHE_SEP, N_CHE_PROD_SEP,
        N_EV_CHE_MIX, N_LOS_CHE_MIX, N_CHE_PROD_MIX,
        N_RPET_OPEN, N_REC_EXPORT
    ]
    idx = {n: i for i, n in enumerate(nodes)}

    links = []
    def L(a, b, v):
        v = float(v)
        if v <= 0:
            return
        if a not in idx or b not in idx:
            raise KeyError(f"Node missing: {a if a not in idx else b}")
        links.append({"source": a, "target": b, "value": v, "source_id": idx[a], "target_id": idx[b]})

    # ----------------------------
    # Links (S1a1 / baseline-style)
    # ----------------------------
    L(N_PRIM,   N_PM, prim_raw)
    L(N_IMP,    N_PM, imp_prod)
    L(N_CHESEC, N_PM, chemsec)
    L(N_RPET,   N_PM, rPET_to_PM)

    L(N_PM, N_EXP,    exp_prod)
    L(N_PM, N_MARKET, PM_market)
    L(N_PM, N_PMREST, PM_rest)

    L(N_MARKET, N_SAMM, waste_total)
    if stock_delta > 0:
        L(N_MARKET, N_STOCK, stock_delta)

    L(N_SAMM, N_COL_SEP, col_sep)
    L(N_SAMM, N_COL_MIX, col_mix)

    L(N_COL_SEP, N_SORT_SEP,    col_sep_to_sort_sep)
    L(N_COL_SEP, N_SORT_MIX,    col_sep_to_sort_mix)
    L(N_COL_SEP, N_EV_COL_SEP,  col_sep_to_ene)
    L(N_COL_SEP, N_LOS_COL_SEP, col_sep_to_los)
    L(N_COL_SEP, N_DIS_COL_SEP, col_sep_to_dis)

    L(N_COL_MIX, N_SORT_SEP,    col_mix_to_sort_sep)
    L(N_COL_MIX, N_SORT_MIX,    col_mix_to_sort_mix)
    L(N_COL_MIX, N_EV_COL_MIX,  col_mix_to_ene)
    L(N_COL_MIX, N_LOS_COL_MIX, col_mix_to_los)
    L(N_COL_MIX, N_DIS_COL_MIX, col_mix_to_dis)

    L(N_SORT_SEP,  N_RECFR_SEP,    sort_sep_to_rec_sep)
    L(N_SORT_SEP,  N_RECFR_MIX,    sort_sep_to_rec_mix)
    L(N_SORT_SEP,  N_EV_SORT_SEP,  sort_sep_to_ene)
    L(N_SORT_SEP,  N_LOS_SORT_SEP, sort_sep_to_los)
    L(N_SORT_SEP,  N_DIS_SORT_SEP, sort_sep_to_dis)

    L(N_SORT_MIX,  N_RECFR_SEP,    sort_mix_to_rec_sep)
    L(N_SORT_MIX,  N_RECFR_MIX,    sort_mix_to_rec_mix)
    L(N_SORT_MIX,  N_EV_SORT_MIX,  sort_mix_to_ene)
    L(N_SORT_MIX,  N_LOS_SORT_MIX, sort_mix_to_los)
    L(N_SORT_MIX,  N_DIS_SORT_MIX, sort_mix_to_dis)

    L(N_RECFR_SEP, N_MEC_SEP, rec_sep_to_MEC)
    L(N_RECFR_SEP, N_CHE_SEP, rec_sep_to_CHE)
    L(N_RECFR_MIX, N_MEC_MIX, rec_mix_to_MEC)
    L(N_RECFR_MIX, N_CHE_MIX, rec_mix_to_CHE)

    # MEC
    L(N_MEC_SEP, N_REC_EXPORT,  MEC_export_sep_u)
    L(N_MEC_SEP, N_RPET,        rPET_MEC_sep)
    L(N_MEC_SEP, N_EV_MEC_SEP,  EV_MEC_sep)
    L(N_MEC_SEP, N_LOS_MEC_SEP, LOS_MEC_sep)
    L(N_MEC_SEP, N_MEC_OTH_SEP, MEC_OTH_sep)

    L(N_MEC_MIX, N_REC_EXPORT,  MEC_export_mix_u)
    L(N_MEC_MIX, N_RPET,        rPET_MEC_mix)
    L(N_MEC_MIX, N_EV_MEC_MIX,  EV_MEC_mix)
    L(N_MEC_MIX, N_LOS_MEC_MIX, LOS_MEC_mix)
    L(N_MEC_MIX, N_MEC_OTH_MIX, MEC_OTH_mix)

    # CHE (baseline-style aggregated, if present)
    L(N_CHE_SEP, N_REC_EXPORT,   CHE_export_sep_u)
    L(N_CHE_SEP, N_EV_CHE_SEP,   EV_CHE_sep)
    L(N_CHE_SEP, N_LOS_CHE_SEP,  LOS_CHE_sep)
    L(N_CHE_SEP, N_CHE_PROD_SEP, PROD_CHE_sep)
    if DIS_CHE_sep > 0:
        L(N_CHE_SEP, f"Beseitigung (CHE, {SEP_TAG})", DIS_CHE_sep)
    L(N_CHE_MIX, N_REC_EXPORT,   CHE_export_mix_u)
    L(N_CHE_MIX, N_EV_CHE_MIX,   EV_CHE_mix)
    L(N_CHE_MIX, N_LOS_CHE_MIX,  LOS_CHE_mix)
    L(N_CHE_MIX, N_CHE_PROD_MIX, PROD_CHE_mix)
    if DIS_CHE_mix > 0:
        L(N_CHE_MIX, f"Beseitigung (CHE, {MIX_TAG})", DIS_CHE_mix)

    # rPET open-loop
    L(N_RPET, N_RPET_OPEN, rPET_open)

    # ----------------------------
    # Tables
    # ----------------------------
    df_links = pd.DataFrame(links)
    if df_links.empty:
        df_links = pd.DataFrame(columns=["source", "target", "value", "source_id", "target_id"])

    in_sum = df_links.groupby("target")["value"].sum() if not df_links.empty else pd.Series(dtype=float)
    out_sum = df_links.groupby("source")["value"].sum() if not df_links.empty else pd.Series(dtype=float)

    node_rows = []
    for n in nodes:
        tin = float(in_sum.get(n, 0.0))
        tout = float(out_sum.get(n, 0.0))
        node_rows.append({
            "node_id": idx[n],
            "node_name": n,
            "total_in": tin,
            "total_out": tout,
            "net_in_minus_out": tin - tout,
            "total_max(in,out)": max(tin, tout),
        })

    df_nodes = pd.DataFrame(node_rows).sort_values("node_id")

    df_check = pd.DataFrame([
        {"metric": "sum_links", "value": float(df_links["value"].sum() if not df_links.empty else 0.0)},
        {"metric": "PM_domestic_after_export(out7)", "value": float(PM_domestic)},
        {"metric": "PM_to_U_total(out7)", "value": float(PM_market)},
        {"metric": "rPET_total(out7)", "value": float(rPET_total)},
        {"metric": "rPET_to_PM_total(out7)", "value": float(rPET_to_PM)},
        {"metric": "rPET_openloop_total(out7)", "value": float(rPET_open)},
    ])

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_path / excel_filename

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_nodes.to_excel(writer, sheet_name="nodes", index=False)
        df_links.to_excel(writer, sheet_name="links", index=False)
        df_check.to_excel(writer, sheet_name="check", index=False)

    return str(xlsx_path.resolve())
# ============================================================
# STEP 6.4 Deterministic reference run (mean values) + Sankey
# ============================================================

import time
import numpy as np
import pandas as pd
import warnings

Mylog.info("### STEP6-REF: Running deterministic Step7 (mean values)")

ParameterDict_mean = {}
for name, entry in ParameterDict.items():
    V, _SD = _get_V_SD(entry)
    ParameterDict_mean[name] = np.asarray(V, dtype=float)

# ---- BASELINE FIRST ----
Mylog.info("### STEP6-BASELINE: Running baseline Step7")

import json
from pathlib import Path

baseline_path = Path(PROJECT_DIR) / "results" / "baseline_out7.json"
with open(baseline_path, "r", encoding="utf-8") as f:
    out7_base = json.load(f)

Mylog.info(
    "### Loaded baseline snapshot: "
    f"PM_domestic={out7_base.get('PM_domestic_after_export')} | "
    f"Prim={out7_base.get('PrimPET_to_PM_raw')} | "
    f"rPET_to_PM={out7_base.get('rPET_to_PM_total')}"
)


# ---- SCENARIO/REF ----
t0 = time.time()
out7_ref = run_step7(
    ParameterDict=ParameterDict_mean,
    PET_MFA_System=PET_MFA_System,
    ModelClassification=ModelClassification,
    Mylog=Mylog,
    verbose=True,
    run_type="single"
)

# SINGLE correction (this enforces: Produktherstellung constant, Prim decreases only by delta rPET vs baseline)
out7_ref = apply_hybrid_constraints_out7_relative_to_baseline(
    out7_ref,
    baseline_out7=out7_base,
    Mylog=Mylog,
    enforce_openloop_min_share=0.10,
    enforce_no_inflation=True
)

Mylog.info("### STEP6-REF finished in %.2fs" % (time.time() - t0))

Mylog.info(
    "CHECK LINKS: "
    f"rPET_to_PM_raw={float(out7_ref.get('rPET_to_PM_raw', 0.0)):.6f} | "
    f"rPET_to_PM_total={float(out7_ref.get('rPET_to_PM_total', 0.0)):.6f} | "
    f"rPET_openloop_total={float(out7_ref.get('rPET_openloop_total', out7_ref.get('rPET_open', 0.0))):.6f} | "
    f"PrimPET_to_PM_raw={float(out7_ref.get('PrimPET_to_PM_raw', 0.0)):.6f}"
)
# ============================================================
# HARD CHECK (non-crashing) - PLACE RIGHT BEFORE SANKEY EXPORT
# ============================================================
import numpy as np

def _sf(x):
    try:
        return float(np.nansum(x))
    except Exception:
        try:
            return float(x)
        except Exception:
            return np.nan

Mylog.info("### CHECK BEFORE SANKEY (NON-CRASH) ###")

# 0) existence check
Mylog.info(f"has out7_base? {'out7_base' in globals()}")
Mylog.info(f"has out7_ref?  {'out7_ref'  in globals()}")

# 1) key values
PMb = _sf(out7_base.get("PM_domestic_after_export", np.nan)) if "out7_base" in globals() else np.nan
PMr = _sf(out7_ref.get("PM_domestic_after_export", np.nan))  if "out7_ref"  in globals() else np.nan

Primb = _sf(out7_base.get("PrimPET_to_PM_raw", np.nan)) if "out7_base" in globals() else np.nan
Primr = _sf(out7_ref.get("PrimPET_to_PM_raw", np.nan))  if "out7_ref"  in globals() else np.nan
rPETpm = _sf(out7_ref.get("rPET_to_PM_total", np.nan))   if "out7_ref"  in globals() else np.nan

Mylog.info(f"BASE PM_domestic_after_export = {PMb:.6f}")
Mylog.info(f"REF  PM_domestic_after_export = {PMr:.6f}")
Mylog.info(f"BASE PrimPET_to_PM_raw = {Primb:.6f}")
Mylog.info(f"REF  PrimPET_to_PM_raw = {Primr:.6f}")
Mylog.info(f"REF  rPET_to_PM_total  = {rPETpm:.6f}")

# 2) hard condition (log only)
if np.isfinite(PMb) and np.isfinite(PMr):
    diff = abs(PMr - PMb)
    Mylog.info(f"DIFF(PM_domestic) = {diff:.6f}")
    if diff > 1e-6:
        Mylog.warning("### FAIL: PM_domestic_after_export is NOT fixed to baseline -> postprocess not applied or overwritten.")
else:
    Mylog.warning("### FAIL: PM_domestic values are NaN -> out7_base/out7_ref not defined or key missing.")
Mylog.info(f"BASE rPET_to_PM_total = {float(np.nansum(out7_base.get('rPET_to_PM_total', 0.0))):.6f}")

# --- SAVE SANKEY ---
sankey_path = export_step7_sankey_no_new_params(
    out7_ref,
    out_dir="outputs",
    filename="sankey_Szenario_S1a1_Deutschland2.html",
    font_size=12
)
Mylog.info(f"### SANKEY SAVED TO: {sankey_path}")

excel_path = export_step7_nodes_links_excel_S1a1(
    out7_ref,
    out_dir="outputs",
    excel_filename="nodes_links_S1a1.xlsx",
    use_short_labels=True
)
Mylog.info(f"### S1a1 NODES+LINKS EXCEL SAVED TO: {excel_path}")

# ============================================================
# STEP 6.5 Monte Carlo loop (N=500) - minimal logs (PASTE-READY)
# ============================================================

Mylog.info(f"### STEP6-MC: Running Monte Carlo N={N_MC}")

warnings.filterwarnings("ignore", category=RuntimeWarning)

records = []
t0 = time.time()

for n in range(N_MC):
    PD_s = build_sampled_parameterdict(ParameterDict, rng)

    out7 = run_step7(
        ParameterDict=PD_s,
        PET_MFA_System=PET_MFA_System,
        ModelClassification=ModelClassification,
        Mylog=None,
        verbose=False,
        run_type="single"
    )

    out7 = apply_hybrid_constraints_out7_relative_to_baseline(
        out7,
        baseline_out7=out7_base,
        Mylog=None,
        enforce_openloop_min_share=0.10,
        enforce_no_inflation=True
    )

    rec = {
        "mc_run": n,

        "PM_to_U_total": float(out7.get("PM_to_U_total", np.nan)),
        "Use_to_Waste_total": float(out7.get("Use_to_Waste_total", out7.get("SANK_in_i_from_use", np.nan))),
        "DeltaStock_u_total": float(out7.get("DeltaStock_u_total", np.nan)),

        "product_import": float(out7.get("product_import", np.nan)),
        "product_export": float(out7.get("product_export", np.nan)),

        "PrimPET_to_PM_raw": float(out7.get("PrimPET_to_PM_raw", np.nan)),
        "P_to_PM_total_raw": float(out7.get("P_to_PM_total_raw", np.nan)),

        "PM_domestic_after_export": float(out7.get("PM_domestic_after_export", np.nan)),
        "PM_rest_loss": float(out7.get("PM_rest_loss", np.nan)),

        "SANK_in_i_from_use": float(out7.get("SANK_in_i_from_use", np.nan)),

        "SANK_rec_m_to_CHE": float(out7.get("SANK_rec_m_to_CHE", np.nan)),
        "SANK_rec_g_to_CHE": float(out7.get("SANK_rec_g_to_CHE", np.nan)),

        "MEC_in": float(out7.get("MEC_in", np.nan)),
        "CHE_in": float(out7.get("CHE_in", np.nan)),

        "EV_from_MEC_total": float(out7.get("EV_from_MEC_total", np.nan)),
        "LOS_from_MEC_total": float(out7.get("LOS_from_MEC_total", np.nan)),
        "EV_from_CHE_total": float(out7.get("EV_from_CHE_total", np.nan)),
        "LOS_from_CHE_total": float(out7.get("LOS_from_CHE_total", np.nan)),

        "SANK_MEC_export": float(out7.get("SANK_MEC_export", np.nan)),
        "SANK_CHE_export": float(out7.get("SANK_CHE_export", np.nan)),

        "rPET_total": float(out7.get("rPET_total", np.nan)),
        "rPET_to_PM_total": float(out7.get("rPET_to_PM_total", out7.get("rPET_to_PM_raw", np.nan))),
        "rPET_openloop_total": float(out7.get("rPET_openloop_total", out7.get("rPET_open", np.nan))),
        "rPET_deficit_for_PM": float(out7.get("rPET_deficit_for_PM", np.nan)),
    }
    records.append(rec)

    if LOG_EVERY and ((n + 1) % LOG_EVERY == 0):
        Mylog.info(f"### STEP6-MC progress: {n+1}/{N_MC}")

Mylog.info("### STEP6-MC finished in %.2fs" % (time.time() - t0))


# ============================================================
# STEP 6.6 Save results + percentile summary
# ============================================================

df_mc = pd.DataFrame.from_records(records)

mc_csv = MC_DIR / f"mc_results_N{N_MC}_seed{SEED}.csv"
df_mc.to_csv(mc_csv, index=False)
Mylog.info(f"### STEP6 saved MC raw results: {mc_csv}")

def summarize_percentiles(df, cols, qs=(0.05, 0.50, 0.95)):
    out_rows = []
    for c in cols:
        s = df[c].dropna()
        if len(s) == 0:
            continue
        row = {"metric": c, "n": int(s.shape[0])}
        for q in qs:
            row[f"p{int(q*100):02d}"] = float(s.quantile(q))
        row["mean"] = float(s.mean())
        row["std"]  = float(s.std(ddof=1)) if s.shape[0] > 1 else 0.0
        out_rows.append(row)
    return pd.DataFrame(out_rows)

metrics = [c for c in df_mc.columns if c != "mc_run"]
df_sum = summarize_percentiles(df_mc, metrics)

sum_csv = MC_DIR / f"mc_summary_N{N_MC}_seed{SEED}.csv"
df_sum.to_csv(sum_csv, index=False)
Mylog.info(f"### STEP6 saved MC summary: {sum_csv}")

Mylog.info("### STEP6 summary preview (first 10 rows):")
Mylog.info("\n" + df_sum.head(10).to_string(index=False))

MC_RESULTS = {
    "df_mc": df_mc,
    "df_summary": df_sum,
    "mc_csv": str(mc_csv),
    "sum_csv": str(sum_csv),
    "out7_ref": out7_ref,
    "sankey_path": sankey_path,
}