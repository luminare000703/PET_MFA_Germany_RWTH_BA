# ============================================
# STEP 1 – Initialize
# ============================================

import sys
import os
from pathlib import Path
from copy import deepcopy
import logging as log

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# --- ODYM imports (패키지 설치 기준) ---
import odym.classes as msc
import odym.functions as msf
import odym.dynamic_stock_model as dsm


# --------------------------------------------
# 1.1 Project paths
# --------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
RESULTS_DIR = PROJECT_DIR / "results"
LOG_DIR = PROJECT_DIR / "logs"

RESULTS_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)


# --------------------------------------------
# 1.2 Logging
# --------------------------------------------
log_verbosity = log.DEBUG
log_filename = LOG_DIR / "PET_MFA_Baseline.log"

[Mylog, console_log, file_log] = msf.function_logger(
    str(log_filename),
    str(LOG_DIR),
    log_verbosity,
    log_verbosity
)

Mylog.info("### 1 - Initialize PET MFA ODYM model")
Mylog.info(f"SCRIPT_DIR:  {SCRIPT_DIR}")
Mylog.info(f"PROJECT_DIR: {PROJECT_DIR}")
Mylog.info(f"DATA_DIR:    {DATA_DIR}")
Mylog.info(f"RESULTS_DIR: {RESULTS_DIR}")

# ============================================
# STEP 2 – Load Config file and read control parameters
# ============================================

Mylog.info("### 2 - Load Config file and read model control parameters")

# -------------------------------------------------
# 2.1 Load project-specific config file
# -------------------------------------------------
ProjectSpecs_Name_ConFile = "ODYM_Config_PET_DE_2019.xlsx"
ConfigPath = DATA_DIR / ProjectSpecs_Name_ConFile

if not ConfigPath.exists():
    raise FileNotFoundError(f"Config file not found: {ConfigPath}")

Model_Configfile = openpyxl.load_workbook(str(ConfigPath), data_only=True)

# -------------------------------------------------
# 2.2 Read model setting (baseline / scenario name)
# -------------------------------------------------
# Convention: sheet "Config", cell (4,4)
ConfigSheet = Model_Configfile["Config"]

ScriptConfig = {}
ScriptConfig["Model Setting"] = ConfigSheet.cell(4, 4).value

Mylog.info(f"Model Setting: {ScriptConfig['Model Setting']}")

# Load the corresponding setting sheet
SettingSheetName = "Setting_" + ScriptConfig["Model Setting"]
if SettingSheetName not in Model_Configfile.sheetnames:
    raise KeyError(f"Expected sheet '{SettingSheetName}' not found in config file.")

Model_Configsheet = Model_Configfile[SettingSheetName]

Name_Scenario = Model_Configsheet.cell(4, 4).value
Mylog.info(f"Scenario name: {Name_Scenario}")


# -------------------------------------------------
# 2.3 Read model control parameters – General Info
# -------------------------------------------------
SCix = 0
while Model_Configsheet.cell(SCix + 1, 2).value != "General Info":
    SCix += 1

SCix += 2  # first data row under header

while Model_Configsheet.cell(SCix + 1, 4).value is not None:
    key = Model_Configsheet.cell(SCix + 1, 3).value
    val = Model_Configsheet.cell(SCix + 1, 4).value
    ScriptConfig[key] = val
    SCix += 1


# -------------------------------------------------
# 2.4 Read software version selection
# -------------------------------------------------
SCix = 0
while Model_Configsheet.cell(SCix + 1, 2).value != "Software version selection":
    SCix += 1

SCix += 2  # first data row under header

while Model_Configsheet.cell(SCix + 1, 4).value is not None:
    key = Model_Configsheet.cell(SCix + 1, 3).value
    val = Model_Configsheet.cell(SCix + 1, 4).value
    ScriptConfig[key] = val
    SCix += 1


# -------------------------------------------------
# 2.5 Log final ScriptConfig
# -------------------------------------------------
Mylog.info("ScriptConfig successfully read:")
for k, v in ScriptConfig.items():
    Mylog.info(f"  {k}: {v}")

# ============================================
# STEP 3 – Read classification and config blocks
# ============================================

Mylog.info("### 3 - Read classification and data")

# -------------------------------------------------
# Helper functions
# -------------------------------------------------
def find_row_by_label(ws, label: str, label_col: int = 2, start_row: int = 1, max_scan: int = 2000) -> int:

    r = start_row
    for _ in range(max_scan):
        if ws.cell(r, label_col).value == label:
            return r
        r += 1
    raise ValueError(f"Label '{label}' not found in column {label_col} within {max_scan} rows.")

def safe_liststring_to_int_list(raw):

    if raw is None:
        return []
    if isinstance(raw, str) and raw.strip() == "":
        return []
    if isinstance(raw, (list, tuple)):
        return [int(x) for x in raw]

    s = str(raw).strip()
    # allow '(1,2,3)' style
    if s.startswith("(") and s.endswith(")"):
        s = "[" + s[1:-1] + "]"
    return msf.ListStringToListNumbers(s)


# -------------------------------------------------
# 3.1 Load master classification file
# -------------------------------------------------
if "Version of master classification" not in ScriptConfig or ScriptConfig["Version of master classification"] is None:
    raise KeyError("ScriptConfig['Version of master classification'] is missing. Check Step 2 General Info block.")

ClassPath = DATA_DIR / f"{ScriptConfig['Version of master classification']}.xlsx"
if not ClassPath.exists():
    raise FileNotFoundError(f"Classification file not found: {ClassPath}")

Classfile = openpyxl.load_workbook(str(ClassPath), data_only=True)
if "MAIN_Table" not in Classfile.sheetnames:
    raise KeyError(f"'MAIN_Table' not found in classification file. Sheets: {Classfile.sheetnames}")

Classsheet = Classfile["MAIN_Table"]

# Read MAIN_Table
ci = 1
MasterClassification = {}

max_cols = Classsheet.max_column
while (ci + 1) <= max_cols:
    if Classsheet.cell(1, ci + 1).value is None:
        Mylog.info(f"End of MAIN_Table reached at column {ci}.")
        break

    ThisName = Classsheet.cell(1, ci + 1).value
    ThisDim  = Classsheet.cell(2, ci + 1).value
    ThisID   = Classsheet.cell(4, ci + 1).value
    ThisUUID = Classsheet.cell(5, ci + 1).value

    TheseItems = []
    ri = 10

    first_item = Classsheet.cell(ri + 1, ci + 1).value
    if first_item is not None:
        TheseItems.append(first_item)

    while True:
        ri += 1
        v = Classsheet.cell(ri + 1, ci + 1).value
        if v is None:
            break
        TheseItems.append(v)

    MasterClassification[ThisName] = msc.Classification(
        Name=ThisName,
        Dimension=ThisDim,
        ID=ThisID,
        UUID=ThisUUID,
        Items=TheseItems
    )
    ci += 1

Mylog.info(f"Loaded {len(MasterClassification)} master classifications from {ClassPath.name}")


# -------------------------------------------------
# 3.2 Read Index Table block
# -------------------------------------------------
Mylog.info("Read index table from model config sheet.")
idx_row = find_row_by_label(Model_Configsheet, "Index Table", label_col=2)
r = idx_row + 2  # first data row

IT_Aspects, IT_Description, IT_Dimension = [], [], []
IT_Classification, IT_Selector, IT_IndexLetter = [], [], []

while Model_Configsheet.cell(r, 3).value is not None:
    IT_Aspects.append(Model_Configsheet.cell(r, 3).value)
    IT_Description.append(Model_Configsheet.cell(r, 4).value)
    IT_Dimension.append(Model_Configsheet.cell(r, 5).value)
    IT_Classification.append(Model_Configsheet.cell(r, 6).value)
    IT_Selector.append(Model_Configsheet.cell(r, 7).value)
    IT_IndexLetter.append(Model_Configsheet.cell(r, 8).value)
    r += 1

Mylog.info(f"Index table rows read: {len(IT_Aspects)}")


# -------------------------------------------------
# 3.3 Read Model Parameters block
# -------------------------------------------------
Mylog.info("Read parameter list from model config sheet.")
pl_row = find_row_by_label(Model_Configsheet, "Model Parameters", label_col=2)
r = pl_row + 2  # first data row

PL_Names, PL_Description, PL_Version = [], [], []
PL_IndexStructure, PL_IndexMatch, PL_IndexLayer = [], [], []

while Model_Configsheet.cell(r, 3).value is not None:
    PL_Names.append(Model_Configsheet.cell(r, 3).value)
    PL_Description.append(Model_Configsheet.cell(r, 4).value)
    PL_Version.append(Model_Configsheet.cell(r, 5).value)
    PL_IndexStructure.append(Model_Configsheet.cell(r, 6).value)
    PL_IndexMatch.append(Model_Configsheet.cell(r, 7).value)

    raw_layer = Model_Configsheet.cell(r, 8).value
    PL_IndexLayer.append(safe_liststring_to_int_list(raw_layer))
    r += 1

Mylog.info(f"Parameter list rows read: {len(PL_Names)}")
if len(PL_Names) == 0:
    Mylog.warning("Parameter list rows read: 0  --> Config sheet 'Model Parameters' Ckecking the location of the block")

# -------------------------------------------------
# 3.4 Read Process Group List block
# -------------------------------------------------
Mylog.info("Read process list from model config sheet.")
pr_row = find_row_by_label(Model_Configsheet, "Process Group List", label_col=2)
r = pr_row + 2

PrL_Number, PrL_Name, PrL_Code, PrL_Type = [], [], [], []

while Model_Configsheet.cell(r, 3).value is not None:
    v = Model_Configsheet.cell(r, 3).value
    try:
        PrL_Number.append(int(v))
    except Exception:
        PrL_Number.append(v)

    PrL_Name.append(Model_Configsheet.cell(r, 4).value)
    PrL_Code.append(Model_Configsheet.cell(r, 5).value)
    PrL_Type.append(Model_Configsheet.cell(r, 6).value)
    r += 1

Mylog.info(f"Process group list rows read: {len(PrL_Number)}")


# -------------------------------------------------
# 3.5 Read Model flow control block (append into ScriptConfig)
# -------------------------------------------------
Mylog.info("Read model run control from model config sheet.")
mf_row = find_row_by_label(Model_Configsheet, "Model flow control", label_col=2)
r = mf_row + 2

while Model_Configsheet.cell(r, 3).value is not None:
    key = Model_Configsheet.cell(r, 3).value
    val = Model_Configsheet.cell(r, 4).value
    if key is not None:
        ScriptConfig[key] = val
    r += 1

Mylog.info("Model flow control entries added to ScriptConfig.")


# -------------------------------------------------
# 3.6 Quick sanity logs
# -------------------------------------------------
Mylog.info(f"ModelClassification keys loaded (master): {list(MasterClassification.keys())[:8]} ...")
Mylog.info(f"IT_Aspects: {IT_Aspects}")
Mylog.info(f"First 5 parameters: {PL_Names[:5]}")

# ============================================
# STEP 4 – Build ModelClassification from Index Table (apply selectors)
# ============================================

Mylog.info("### 4 - Define model classifications and select items according to config file.")

# 4.1 ModelClassification dict  (aspect -> selected Classification)
ModelClassification = {}

def normalize_selector(x):

    if x is None:
        return "all"

    # if it's already string:
    if isinstance(x, str):
        s = x.strip()
        if s == "":
            return "all"
        if s.lower() == "all":
            return "all"

        return s

    # numbers: keep as int-like string for later interpretation (index vs value)
    if isinstance(x, (int, float)):
        return str(int(x))

    return str(x).strip()

# --------------------------------

for a, cls_name, selector_raw in zip(IT_Aspects, IT_Classification, IT_Selector):

    if cls_name not in MasterClassification:
        raise KeyError(
            f"Index Table requests classification '{cls_name}' for aspect '{a}', "
            f"but it is not in MasterClassification."
        )

    # deepcopy master → model classification
    ModelClassification[a] = deepcopy(MasterClassification[cls_name])


    selector = normalize_selector(selector_raw)

    items = ModelClassification[a].Items
    n_items = len(items)

    if selector == "all":
        pass

    elif (":" in selector) or ("[" in selector):
        pass

    elif selector.lstrip("-").isdigit():
        num = int(selector)

        if 0 <= num < n_items:
            selector = f"[{num}]"

        elif num in items:
            selector = f"[{items.index(num)}]"

        elif str(num) in [str(it) for it in items]:
            idx = [str(it) for it in items].index(str(num))
            selector = f"[{idx}]"
        else:
            raise ValueError(
                f"Numeric selector '{num}' for aspect '{a}' is neither a valid index "
                f"(0..{n_items - 1}) nor an existing item value. "
                f"Classification='{cls_name}', items={items}"
            )

    else:
        # exact match
        if selector in items:
            selector = f"[{items.index(selector)}]"
        else:
            # case-insensitive exact match
            low = [str(it).strip().lower() for it in items]
            s_low = selector.strip().lower()
            hits = [k for k, it in enumerate(low) if it == s_low]
            if len(hits) == 1:
                selector = f"[{hits[0]}]"
            else:
                # contains match (unique only)
                hits2 = [k for k, it in enumerate(low) if s_low in it]
                if len(hits2) == 1:
                    selector = f"[{hits2[0]}]"
                else:
                    raise ValueError(
                        f"Text selector '{selector}' for aspect '{a}' could not be matched uniquely.\n"
                        f"Classification='{cls_name}', n_items={n_items}\n"
                        f"Items={items}"
                    )
    # --- END NEW ---


    if selector not in ("all",) and (":" not in selector) and ("[" not in selector):
        items = ModelClassification[a].Items

        if selector in items:
            selector = f"[{items.index(selector)}]"
        else:

            hits = [k for k, it in enumerate(items) if str(it).strip().lower() == selector.strip().lower()]
            if len(hits) == 1:
                selector = f"[{hits[0]}]"
            else:
                hits2 = [k for k, it in enumerate(items) if selector.strip().lower() in str(it).strip().lower()]
                if len(hits2) == 1:
                    selector = f"[{hits2[0]}]"
                else:
                    raise ValueError(
                        f"Text selector '{selector}' for aspect '{a}' could not be matched uniquely.\n"
                        f"Classification='{cls_name}', n_items={len(items)}\n"
                        f"First items={items[:min(20, len(items))]}"
                    )
    # ---------------------------------------------------------------

    Mylog.info(
        f"Selector check: aspect={a}, raw={selector_raw} ({type(selector_raw)}), normalized={selector}"
    )

import ast

EvalString = msf.EvalItemSelectString(
    selector,
    len(ModelClassification[a].Items)
)

if EvalString.find(':') > -1:
    RangeStart = int(EvalString.split(':')[0])
    RangeStop  = int(EvalString.split(':')[1])
    ModelClassification[a].Items = ModelClassification[a].Items[RangeStart:RangeStop]

elif EvalString.find('[') > -1:
    idx_list = ast.literal_eval(EvalString)

    n_items = len(ModelClassification[a].Items)
    bad = [i for i in idx_list if (not isinstance(i, int)) or i < 0 or i >= n_items]

    if bad:
        Mylog.error(
            f"[IndexError 예방] Selector out of range!\n"
            f"  aspect={a}\n"
            f"  classification={cls_name}\n"
            f"  selector_raw={selector_raw} ({type(selector_raw)})\n"
            f"  normalized_selector={selector}\n"
            f"  EvalString={EvalString}\n"
            f"  n_items={n_items}\n"
            f"  bad_indices={bad}\n"
            f"  available_index_range=0..{n_items-1}"
        )
        raise IndexError(
            f"Selector for aspect '{a}' selects indices {bad} "
            f"but only {n_items} items exist."
        )

    ModelClassification[a].Items = [
        ModelClassification[a].Items[i] for i in idx_list
    ]

elif EvalString == 'all':
    pass

else:
    raise ValueError(
        f"Invalid selector '{selector}' for aspect '{a}'. "
        f"EvalItemSelectString returned '{EvalString}'."
    )

# --- Step 4 end: index sizes ---
Mylog.info(f"ModelClassification aspects: {list(ModelClassification.keys())}")
IndexSize = {a: len(ModelClassification[a].Items) for a in IT_Aspects}
Mylog.info(f"Index sizes: {IndexSize}")

# ============================================
# STEP 5 – Read parameters from Excel files
# (UNCERTAINTY-SAFE, ODYM-VERSION-ROBUST)
# ============================================

import inspect
import ast
import numpy as np
import pandas as pd

Mylog.info("### 5 - Read parameters from Excel files")

# ODYM uncertainty parsing ON
ParseUncertainty = True


# -------------------------------------------------
# 5.0 Build IndexTable (ODYM legacy requirement)
# -------------------------------------------------
IndexTable = pd.DataFrame({
    "Aspect": IT_Aspects,
    "Description": IT_Description,
    "Dimension": IT_Dimension,
    "Classification": [ModelClassification[a] for a in IT_Aspects],
    "ClassificationName": IT_Classification,
    "Selector": IT_Selector,
    "IndexLetter": IT_Aspects,
    "IndexSize": [len(ModelClassification[a].Items) for a in IT_Aspects],
})
IndexTable_ClassificationNames = IT_Classification

Mylog.info("IndexTable IndexLetter check: " + str(IndexTable["IndexLetter"].tolist()))
Mylog.info("IndexTable IndexSize check: " + str(IndexTable["IndexSize"].tolist()))


# -------------------------------------------------
# 5.1 ReadParameterXLSX signature (robust positional)
# -------------------------------------------------
sig = inspect.signature(msf.ReadParameterXLSX)
param_names = list(sig.parameters.keys())

Mylog.info(f"ReadParameterXLSX signature: {sig}")
Mylog.info(f"ReadParameterXLSX parameters: {param_names}")


# -------------------------------------------------
# 5.2 Helper: IndexMatch / LayerSel / Ix
# -------------------------------------------------
def convert_indexmatch_to_int_list_string(indexmatch_raw, indexstructure_raw):
    if indexmatch_raw is None or str(indexmatch_raw).strip() == "":
        indexmatch_raw = str(indexstructure_raw)

    s = str(indexmatch_raw).strip()

    if s.startswith("[") and s.endswith("]"):
        return s

    if "*" in s:
        parts = [x.strip() for x in s.split("*")]
    elif "," in s:
        parts = [x.strip() for x in s.split(",")]
    else:
        parts = [s]

    idx = [IT_Aspects.index(letter) for letter in parts]
    return str(idx)


def normalize_layersel(ls):
    if ls is None:
        return [[0]]
    if isinstance(ls, list):
        if len(ls) == 0:
            return [[0]]
        if isinstance(ls[0], int):
            return [ls]
        return ls
    return [[int(ls)]]


def build_thisparix_from_indexstructure(indexstructure):
    return [s.strip() for s in str(indexstructure).split("*") if s.strip()]


# -------------------------------------------------
# 5.3 Robust ODYM call (positional-safe)
# -------------------------------------------------
def call_ReadParameterXLSX_positional(
    *,
    p_path,
    ThisPar,
    ThisParIx,
    IndexMatch,
    ThisParLayerSel,
    MasterClassification,
    IndexTable,
    IndexTable_ClassificationNames,
    ScriptConfig,
    Mylog,
    ParseUncertainty
):
    argmap = {
        "ParPath": str(p_path.with_suffix("")),
        "ThisPar": ThisPar,
        "ThisParIx": ThisParIx,
        "IndexMatch": str(IndexMatch),
        "ThisParLayerSel": ThisParLayerSel,
        "ThisParProcMethod": "['none']",
        "MasterClassification": MasterClassification,
        "IndexTable": IndexTable,
        "IndexTable_ClassificationNames": IndexTable_ClassificationNames,
        "ScriptConfig": ScriptConfig,
        "Mylog": Mylog,
        "ParseUncertainty": ParseUncertainty,
    }

    args = []
    for name in param_names:
        if name not in argmap:
            raise TypeError(f"Cannot map ReadParameterXLSX argument '{name}'")
        args.append(argmap[name])

    return msf.ReadParameterXLSX(*args)

import numpy as np
import openpyxl
import re

def _to_float_or_none(x):

    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if s == "" or s.lower() in ("nan", "none"):
        return None

    # percent like "3%" or "15 %"
    m = re.match(r"^(-?\d+(\.\d+)?)\s*%$", s)
    if m:
        return float(m.group(1)) / 100.0

    # number in string
    try:
        return float(s)
    except Exception:
        return None


def _find_header_row(ws, required_cols, max_scan_rows=80, max_scan_cols=50):

    req = [c.strip().lower() for c in required_cols]

    for r in range(1, max_scan_rows + 1):
        row_vals = []
        for c in range(1, max_scan_cols + 1):
            v = ws.cell(r, c).value
            if v is None:
                row_vals.append("")
            else:
                row_vals.append(str(v).strip())

        low = [x.lower() for x in row_vals]
        # must contain all required
        if all(any(rc == lv for lv in low) for rc in req):
            col_map = {}
            for rc in req:
                col_map[rc] = low.index(rc)  # 0-based index in row_vals
            return r, col_map

    return None, None


def _get_sheet_for_values_table(wb):

    candidates = []
    for name in ["Values_Master", "Values", "values_master", "values"]:
        if name in wb.sheetnames:
            candidates.append(name)

    # If not found or mismatch, try to detect
    if not candidates:
        for sn in wb.sheetnames:
            ws = wb[sn]
            r, _ = _find_header_row(ws, required_cols=["value"], max_scan_rows=60, max_scan_cols=60)
            if r is not None:
                candidates.append(sn)

    return candidates[0] if candidates else None


def _index_lookup_for_aspect(ModelClassification, aspect_letter):

    items = list(ModelClassification[aspect_letter].Items)
    lut = {}
    for idx, it in enumerate(items):
        lut[str(it).strip()] = idx
    return lut


def read_stddev_list_parameter(
    *,
    xlsx_path,
    index_structure,
    ModelClassification,
    target_shape,
    expect_absolute=True
):

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    sheet_name = _get_sheet_for_values_table(wb)
    if sheet_name is None:
        return None

    ws = wb[sheet_name]

    # Required columns = aspect letters + 'value' + 'StdDev'
    aspects = [s.strip() for s in str(index_structure).split("*") if s.strip()]
    required = aspects + ["value", "StdDev"]

    header_row, col_map = _find_header_row(ws, required_cols=required, max_scan_rows=120, max_scan_cols=80)
    if header_row is None:
        return None

    # Build lookups for aspects to indices
    aspect_luts = {a: _index_lookup_for_aspect(ModelClassification, a) for a in aspects}

    # Create output array
    SD = np.zeros(target_shape, dtype=float)
    filled = 0

    # Iterate rows below header
    r = header_row + 1
    while True:
        # Stop if the first aspect cell is empty AND value empty (end of table)
        first_val = ws.cell(r, col_map[aspects[0].lower()] + 1).value
        val_cell  = ws.cell(r, col_map["value"] + 1).value
        sd_cell   = ws.cell(r, col_map["stddev"] + 1).value  # normalized key is lower

        if first_val is None and val_cell is None and sd_cell is None:
            break

        # Parse indices
        idxs = []
        ok = True
        for a in aspects:
            col0 = col_map[a.lower()] + 1
            v = ws.cell(r, col0).value
            if v is None:
                ok = False
                break

            key = str(v).strip()
            # Some sheets store numeric codes; normalize to string
            lut = aspect_luts[a]
            if key not in lut:
                key2 = str(int(v)) if isinstance(v, float) and v.is_integer() else key
                if key2 in lut:
                    idxs.append(lut[key2])
                else:
                    ok = False
                    break
            else:
                idxs.append(lut[key])

        if not ok:
            r += 1
            continue

        sd = _to_float_or_none(sd_cell)
        if sd is None:
            r += 1
            continue

        # Assign
        SD[tuple(idxs)] = float(sd)
        filled += 1
        r += 1

    if filled == 0:
        return None

    return SD

# -------------------------------------------------
# 5.4  Extract Values + StdDev from ANY ODYM return
# -------------------------------------------------
def extract_values_and_std(res, par_name=""):
    cover = None
    values = None
    std = None

    def _take(x):
        nonlocal cover, values, std

        if isinstance(x, np.ndarray):
            if values is None:
                values = x
            elif std is None and x.shape == values.shape:
                std = x
            return

        if isinstance(x, dict):
            if "Values" in x:
                values = np.asarray(x["Values"])
                if "StdDev" in x:
                    std = np.asarray(x["StdDev"])
            else:
                cover = x
            return

        if hasattr(x, "Values"):
            values = np.asarray(x.Values)
            if hasattr(x, "StdDev"):
                std = np.asarray(x.StdDev)
            return

    if isinstance(res, tuple):
        for x in res:
            _take(x)
    else:
        _take(res)

    if values is None:
        meta = list(cover.keys()) if isinstance(cover, dict) else None
        raise TypeError(
            f"[STEP5 FAIL] '{par_name}' returned no Values. "
            f"type={type(res)}, cover_keys={meta}"
        )

    return values, std, cover


# -------------------------------------------------
# 5.5 Read all parameters → ParameterDict
# -------------------------------------------------
Mylog.info("### 5.5 - Read parameters (xlsx) and build ParameterDict")

ParameterDict = {}

for k in range(len(PL_Names)):
    ThisPar = PL_Names[k]

    IndexStructure = PL_IndexStructure[k]
    ThisParIx = build_thisparix_from_indexstructure(IndexStructure)

    IndexMatch = convert_indexmatch_to_int_list_string(
        PL_IndexMatch[k], IndexStructure
    )

    ThisParLayerSel = normalize_layersel(PL_IndexLayer[k])

    p_file = PL_Version[k]
    if not p_file.lower().endswith(".xlsx"):
        p_file += ".xlsx"
    p_path = DATA_DIR / p_file

    Mylog.info(
        f"[{k+1}/{len(PL_Names)}] Reading '{ThisPar}' from {p_path.name} "
        f"(Ix={IndexStructure}, Match={IndexMatch}, LayerSel={ThisParLayerSel})"
    )

    Par_raw = call_ReadParameterXLSX_positional(
        p_path=p_path,
        ThisPar=ThisPar,
        ThisParIx=ThisParIx,
        IndexMatch=IndexMatch,
        ThisParLayerSel=ThisParLayerSel,
        MasterClassification=MasterClassification,
        IndexTable=IndexTable,
        IndexTable_ClassificationNames=IndexTable_ClassificationNames,
        ScriptConfig=ScriptConfig,
        Mylog=Mylog,
        ParseUncertainty=ParseUncertainty
    )

    V, SD, COVER = extract_values_and_std(Par_raw, ThisPar)


    if SD is None:
        SD = read_stddev_list_parameter(
            xlsx_path=p_path,
            index_structure=IndexStructure,
            ModelClassification=ModelClassification,
            target_shape=V.shape
        )

    ParameterDict[ThisPar] = {
        "Values": V,
        "StdDev": SD,
        "Cover": COVER
    }

    Mylog.info(
        f"  -> OK: '{ThisPar}' Values={V.shape}, StdDev={'yes' if SD is not None else 'no'}"
    )

    ParameterDict[ThisPar] = {
        "Values": V,
        "StdDev": SD,
        "Cover": COVER
    }

Mylog.info(f"### STEP 5 finished: {len(ParameterDict)} parameters loaded")

# ============================================================
# STEP 6 + STEP 7 (FULL, NO SIMPLIFICATION)
# ============================================================

import numpy as np
import pandas as pd
import time
from pathlib import Path
import plotly.graph_objects as go


# ============================================================
# USER SETTINGS (MC)
# ============================================================
N_MC = 500
SEED = 42
LOG_EVERY = 50
rng = np.random.default_rng(SEED)

MC_DIR = RESULTS_DIR / "mc"
MC_DIR.mkdir(parents=True, exist_ok=True)

Mylog.info(f"### STEP6 settings: N_MC={N_MC}, SEED={SEED}, LOG_EVERY={LOG_EVERY}, MC_DIR={MC_DIR}")


# ============================================================
# STEP 6.0 – Create PET_MFA_System (MFAsystem)
# ============================================================
Mylog.info("### STEP6.0 - Creating PET_MFA_System (MFAsystem)")

t_items = list(ModelClassification["t"].Items)

def _to_int_year(x):
    if isinstance(x, (int, np.integer)):
        return int(x)
    s = str(x).strip()
    try:
        return int(float(s))
    except Exception:
        for i in range(len(s) - 3):
            chunk = s[i:i+4]
            if chunk.isdigit():
                return int(chunk)
        raise ValueError(f"Cannot parse time item: {x}")

Time_Start = _to_int_year(t_items[0])
Time_End   = _to_int_year(t_items[-1])

if "Chemical_Elements" in MasterClassification:
    Elements = MasterClassification["Chemical_Elements"]
else:
    Elements = msc.Classification(
        Name="Chemical_Elements",
        Dimension="Element",
        ID="E",
        UUID="",
        Items=["All"]
    )

ProcessList = []
for n, name in zip(PrL_Number, PrL_Name):
    ProcessList.append(msc.Process(ID=int(n), Name=str(name)))

PET_MFA_System = msc.MFAsystem(
    Name=f"PET_MFA_DE_2019_{ScriptConfig.get('Model Setting','Baseline')}",
    Geogr_Scope="Germany",
    Unit="kt",
    Time_Start=Time_Start,
    Time_End=Time_End,
    IndexTable=IndexTable,
    Elements=Elements,
    ProcessList=ProcessList,
    FlowDict={},
    StockDict={},
    ParameterDict={}
)

Mylog.info("### STEP6.0 - PET_MFA_System created successfully")


# ============================================================
# STEP 6.x – Monte Carlo helper functions
# ============================================================

def _get_V_SD(entry):

    if isinstance(entry, np.ndarray):
        return entry, None
    if isinstance(entry, dict):
        V = entry.get("Values", None)
        SD = entry.get("StdDev", None)
        if V is None:
            raise TypeError("Parameter entry dict without 'Values'")
        return np.asarray(V, dtype=float), (None if SD is None else np.asarray(SD, dtype=float))
    if hasattr(entry, "Values"):
        V = np.asarray(entry.Values, dtype=float)
        SD = np.asarray(entry.StdDev, dtype=float) if hasattr(entry, "StdDev") else None
        return V, SD
    raise TypeError(f"Unsupported parameter entry type: {type(entry)}")


def _clip_nonneg(A):
    return np.clip(np.asarray(A, dtype=float), 0.0, None)

def _clip_01(A):
    return np.clip(np.asarray(A, dtype=float), 0.0, 1.0)

def _safe_normal(mean, sd, rng):
    mean = np.asarray(mean, dtype=float)
    if sd is None:
        return mean.copy()
    sd = np.asarray(sd, dtype=float)
    if sd.size == 0 or np.all(sd == 0):
        return mean.copy()
    return rng.normal(loc=mean, scale=sd)

def _row_normalize_last_axis(A, eps=1e-12):
    A = np.asarray(A, dtype=float)
    s = A.sum(axis=-1, keepdims=True)
    return np.where(s > eps, A / s, A)

def _sample_mass(mean, sd, rng):
    return _clip_nonneg(_safe_normal(mean, sd, rng))

def _sample_prob01(mean, sd, rng):
    return _clip_01(_safe_normal(mean, sd, rng))

def _sample_share(mean, sd, rng):
    x = _clip_nonneg(_safe_normal(mean, sd, rng))
    return _row_normalize_last_axis(x)

def _sample_matrix(mean, sd, rng):
    # matrix (..., i, j) with j last axis: normalize rows along last axis
    x = _clip_nonneg(_safe_normal(mean, sd, rng))
    return _row_normalize_last_axis(x)

def sample_parameter(name, V, SD, rng):

    # MASS
    if name.startswith("p_mass_"):
        return _sample_mass(V, SD, rng)

    # SPECIFIC SHARES (row-normalize last axis)
    if name in ("p_share_use_waste_type"):
        return _sample_share(V, SD, rng)

    # ROUTING MATRICES (row-normalize last axis)
    if name in (
        "p_share_collection_routing",
        "p_share_sorting_separate",
        "p_share_sorting_mixed",
        "p_share_recycling_routing",
        "p_share_recycling_stream_routing",
        "p_share_sorting_to_terminal",
    ):
        return _sample_matrix(V, SD, rng)

    # PROBABILITIES (0..1)
    if name in (
        "p_collection_efficiency_w",
        "p_rPET_eligible_share_packaging_tr",
        "p_share_recycled_output_to_PM",
    ):
        return _sample_prob01(V, SD, rng)

    # YIELDS (0..1, no normalization)
    if name == "p_recycling_process_yields":
        return _sample_prob01(V, SD, rng)

    # DEFAULT for other p_share_*:
    if name.startswith("p_share_"):
        # if it's vector-like (<=3 dims), treat as prob; else as share
        if np.asarray(V).ndim <= 3:
            return _sample_prob01(V, SD, rng)
        return _sample_share(V, SD, rng)

    # FALLBACK: nonnegative
    return _sample_mass(V, SD, rng)

def build_sampled_parameterdict(ParameterDict_base, rng):

    PD = {}
    for name, entry in ParameterDict_base.items():
        V, SD = _get_V_SD(entry)
        PD[name] = sample_parameter(name, V, SD, rng)
    return PD

# ============================================================
# Hybrid injection AFTER Step7 (rPET back to PM)
# ============================================================

def apply_step7_hybrid_injection(out7, Mylog, log=False):

    PM_primary = float(out7.get("PM_to_U_total", 0.0) or 0.0)
    rPET_to_PM = float(out7.get("rPET_to_PM_total", 0.0) or 0.0)
    PM_hybrid = PM_primary + rPET_to_PM

    Use_to_Waste = float(out7.get("SANK_in_i_from_use", 0.0) or 0.0)

    Delta_primary = PM_primary - Use_to_Waste
    Delta_hybrid  = PM_hybrid  - Use_to_Waste

    out7["PM_to_U_total_primary"] = PM_primary
    out7["PM_to_U_total_hybrid"]  = PM_hybrid
    out7["Use_to_Waste_total"]    = Use_to_Waste
    out7["DeltaStock_u_total"]    = Delta_primary
    out7["DeltaStock_u_total_hybrid"] = Delta_hybrid

    if log:
        Mylog.info(
            "### HYBRID METRICS: PM_primary=%.3f | rPET_to_PM=%.3f | PM_hybrid=%.3f | Use_to_Waste=%.3f | Delta_primary=%.3f | Delta_hybrid=%.3f"
            % (PM_primary, rPET_to_PM, PM_hybrid, Use_to_Waste, Delta_primary, Delta_hybrid)
        )
    return out7

# ============================================================
# STEP 7 – Full process chain (Fraktion-based recycling stream)
# ============================================================
def run_step7(ParameterDict, PET_MFA_System, ModelClassification, Mylog,
              verbose=False, run_type="single"):

    import numpy as np

    # -----------------------------
    # helpers
    # -----------------------------
    def _pv(PD, name: str):
        if name not in PD:
            raise KeyError(f"Missing parameter '{name}'")
        par = PD[name]
        if isinstance(par, dict) and "Values" in par:
            return par["Values"]
        if hasattr(par, "Values"):
            return par.Values
        return par

    def _scalar(x, default=0.0):
        """Robust scalar extraction (sum for arrays)."""
        try:
            if x is None:
                return float(default)
            if isinstance(x, (float, int, np.floating, np.integer)):
                return float(x)
            a = np.asarray(x, dtype=float)
            if a.size == 0:
                return float(default)
            return float(np.nansum(a))
        except Exception:
            return float(default)

    def _cls_items(key: str):
        return list(ModelClassification[key].Items)

    def _item_name(x):
        return x.Name if hasattr(x, "Name") else str(x)

    def _find_index_contains(cls_key: str, target: str):
        names = [_item_name(it) for it in _cls_items(cls_key)]
        low = [n.lower() for n in names]
        t2 = target.lower()
        for k, nm in enumerate(low):
            if t2 in nm:
                return k
        raise KeyError(f"Could not find '{target}' in '{cls_key}'. Available={names}")

    def _ensure_e_mass(A):
        A = np.asarray(A, dtype=float)
        if A.ndim in (3, 4):
            return A[..., None]
        return A

    def _ensure_e_matrix(S):
        S = np.asarray(S, dtype=float)
        if S.ndim == 4:
            return S[..., None]
        return S

    def _apply_matrix_rowonly(M_in, S, tol=1e-12, return_tensor=False):

        row_sum = S.sum(axis=3)               # (t,r,i,1)
        active = row_sum > tol

        M_active = np.where(active, M_in, 0.0)
        routed = np.einsum("trie,trije->trje", M_active, S)  # (t,r,j,1)

        M_inactive = M_in - M_active
        M_out = routed + M_inactive

        F = None
        if return_tensor:
            F = M_active[:, :, :, None, :] * S
            diag = np.zeros_like(F)
            for kk in range(M_in.shape[2]):
                diag[:, :, kk, kk, :] = M_inactive[:, :, kk, :]
            F = F + diag
        return (M_out, F) if return_tensor else M_out

    def _sum_flow(F, i_from, i_to):
        A = np.asarray(F)
        return float(A[:, :, i_from, i_to, 0].sum())

    # -----------------------------
    # STRICT routing matrix (rowsum==1 for active rows)
    # -----------------------------
    def _get_S_strict(name: str, shape_expected):
        S = _ensure_e_matrix(_pv(ParameterDict, name))
        if S.shape != shape_expected:
            raise ValueError(f"{name} expected {shape_expected} got {S.shape}")
        S = np.clip(S, 0, 1)

        rs = S.sum(axis=3)  # (t,r,i,1)
        used = rs > 1e-12
        bad = used & (np.abs(rs - 1) > 1e-8)
        if np.any(bad):
            first = np.argwhere(bad)[0].tolist()
            raise ValueError(f"{name}: rowsum != 1 for an active row. first_bad={first}")
        return S

    # -----------------------------
    # PARTIAL process output shares (rowsum can be <1) for residual -> ENE/LOS/DIS
    # -----------------------------
    def _get_S_partial(name: str, shape_expected):
        S = _ensure_e_matrix(_pv(ParameterDict, name))
        if S.shape != shape_expected:
            raise ValueError(f"{name} expected {shape_expected} got {S.shape}")
        S = np.clip(S, 0, 1)

        rs = S.sum(axis=3)  # (t,r,i,1)
        too_big = rs > (1.0 + 1e-8)
        if np.any(too_big):
            first = np.argwhere(too_big)[0].tolist()
            Mylog.warning(f"### WARNING {name}: rowsum > 1 for some rows. Normalizing those rows. first_bad={first}")
            rs_safe = np.where(rs > 1e-12, rs, 1.0)
            S = S / rs_safe[:, :, :, None, :]
        return S

    out = {"ok": True}

    # -----------------------------
    # 7.0 Production + product trade -> available supply for PM (raw accounting)
    # -----------------------------
    M_prod = np.asarray(_pv(ParameterDict, "p_mass_production_inputs"), dtype=float)  # (t,r,p,q)
    if M_prod.ndim != 4:
        raise TypeError(f"p_mass_production_inputs must be (t,r,p,q). Got {M_prod.shape}")

    T, Rr, P, Q = M_prod.shape

    M_imp_pq = np.zeros((T, Rr, P, Q), dtype=float)
    M_exp_pq = np.zeros((T, Rr, P, Q), dtype=float)

    if "p_mass_p_imports" in ParameterDict:
        M_imp = np.asarray(_pv(ParameterDict, "p_mass_p_imports"), dtype=float)
        for k in range(min(P, Q)):
            M_imp_pq[:, :, k, k] = np.clip(M_imp[:, :, k], 0, None)

    if "p_mass_p_exports" in ParameterDict:
        M_exp = np.asarray(_pv(ParameterDict, "p_mass_p_exports"), dtype=float)
        for k in range(min(P, Q)):
            M_exp_pq[:, :, k, k] = np.clip(M_exp[:, :, k], 0, None)

    out["product_import"] = float(M_imp_pq.sum())
    out["product_export"] = float(M_exp_pq.sum())

    M_mfg_available = np.clip(M_prod + M_imp_pq - M_exp_pq, 0, None)

    # Supply components (raw, baseline)
    p_primary   = 0
    p_rpet_raw  = 2
    p_chemsec   = 3

    F_to_PM_trpq = np.zeros((T, Rr, P, Q), dtype=float)
    for ps in [p_primary, p_rpet_raw, p_chemsec]:
        if ps < P:
            F_to_PM_trpq[:, :, ps, :] = M_mfg_available[:, :, ps, :]

    out["P_to_PM_total_raw"] = float(F_to_PM_trpq.sum())
    out["PrimPET_to_PM_raw"] = float(F_to_PM_trpq[:, :, p_primary, :].sum())
    out["rPET_to_PM_raw"]    = float(F_to_PM_trpq[:, :, p_rpet_raw, :].sum())
    out["chemsec_to_PM_raw"] = float(F_to_PM_trpq[:, :, p_chemsec, :].sum())

    # -----------------------------
    # 2) PM -> Use (market output)
    # -----------------------------
    P_to_U_raw = np.asarray(_pv(ParameterDict, "p_mass_production_to_use"), dtype=float)
    if P_to_U_raw.ndim == 5:
        P_to_U_raw = P_to_U_raw[..., 0]
    if P_to_U_raw.ndim != 4:
        raise TypeError(f"p_mass_production_to_use must be 4D. Got {P_to_U_raw.shape}")

    tmp_u = np.asarray(_pv(ParameterDict, "p_mass_use_to_collection"), dtype=float)
    if tmp_u.ndim == 5:
        tmp_u = tmp_u[..., 0]
    U_expected = tmp_u.shape[2]

    t2, r2, a2, b2 = P_to_U_raw.shape
    if b2 == U_expected:
        P_to_U_trpu = P_to_U_raw
    elif a2 == U_expected:
        Mylog.warning("p_mass_production_to_use is (t,r,u,p). Swapping to (t,r,p,u).")
        P_to_U_trpu = np.swapaxes(P_to_U_raw, 2, 3)
    else:
        raise TypeError(f"Cannot infer u-axis for p_mass_production_to_use. shape={P_to_U_raw.shape}, expected U={U_expected}")

    Use_in_tru = P_to_U_trpu.sum(axis=2)  # (t,r,u)
    out["PM_to_U_total"] = float(Use_in_tru.sum())  # MARKET output (do NOT add rPET loop here)

    # -----------------------------
    # 3) Use -> Collection (waste generation)
    # -----------------------------
    U_to_i_raw = np.asarray(_pv(ParameterDict, "p_mass_use_to_collection"), dtype=float)
    U_to_i_trui = U_to_i_raw[..., 0] if U_to_i_raw.ndim == 5 else U_to_i_raw
    if U_to_i_trui.ndim != 4:
        raise TypeError(f"p_mass_use_to_collection must be (t,r,u,i). Got {U_to_i_trui.shape}")

    U_to_W_tru = U_to_i_trui.sum(axis=3)

    # align u if needed
    u_in = Use_in_tru.shape[2]
    u_out = U_to_W_tru.shape[2]
    if u_in != u_out:
        Mylog.warning(f"Use-phase mismatch: Use_in u={u_in}, Waste u={u_out}. Adjusting.")
        if u_in > u_out:
            extra = Use_in_tru[..., (u_out - 1):].sum(axis=2, keepdims=True)
            Use_in_tru = np.concatenate([Use_in_tru[..., : (u_out - 1)], extra], axis=2)
        else:
            pad = ((0, 0), (0, 0), (0, u_out - u_in))
            Use_in_tru = np.pad(Use_in_tru, pad, constant_values=0.0)

    DeltaStock_u_tru = Use_in_tru - U_to_W_tru
    out["DeltaStock_u_total"] = float(DeltaStock_u_tru.sum())
    out["SANK_in_i_from_use"] = float(U_to_i_trui.sum())

    # i inflow
    M_u_i_e = _ensure_e_mass(U_to_i_trui)
    Remain_i = M_u_i_e.sum(axis=2)  # (t,r,i,1)

    # -----------------------------
    # indices
    # -----------------------------
    I = len(_cls_items("i"))

    I_COL_G  = _find_index_contains("i", "Getrennte Sammlung")
    I_COL_M  = _find_index_contains("i", "Gemischte Sammlung")
    I_SORT_G = _find_index_contains("i", "Sortierung_getrennt")
    I_SORT_M = _find_index_contains("i", "Sortierung_gemischt")
    I_REC_G  = _find_index_contains("i", "Recycling_getrennt")
    I_REC_M  = _find_index_contains("i", "Recycling_gemischt")

    IENE = _find_index_contains("i", "Energetische Verwertung")
    ILOS = _find_index_contains("i", "Verluste")
    IDIS = _find_index_contains("i", "Beseitigung")

    IMEC_G = _find_index_contains("i", "Mechanisches Recycling (getrennt)")
    IMEC_M = _find_index_contains("i", "Mechanisches Recycling (gemischt)")
    ICHE_G = _find_index_contains("i", "Chemisches Recycling (getrennt)")
    ICHE_M = _find_index_contains("i", "Chemisches Recycling (gemischt)")

    # p indices for yields outputs
    p_rPET = _find_index_contains("p", "rPET")
    # chem. Sekundärrohstoffe: try common substrings
    try:
        p_chemsec_out = _find_index_contains("p", "sekund")
    except KeyError:
        try:
            p_chemsec_out = _find_index_contains("p", "chem")
        except KeyError:
            p_chemsec_out = 3  # fallback

    # -----------------------------
    # routing matrices
    # -----------------------------
    shape_S = (T, Rr, I, I, 1)
    S_col      = _get_S_strict("p_share_collection_routing", shape_S)
    S_sort_sep = _get_S_strict("p_share_sorting_separate", shape_S)
    S_sort_mix = _get_S_strict("p_share_sorting_mixed", shape_S)
    S_rec_proc = _get_S_strict("p_share_recycling_routing", shape_S)     # Fraktion -> MEC/CHE only
    S_proc_out = _get_S_partial("p_share_rec_process_outputs", shape_S)  # residual -> (ENE/LOS/DIS) partial shares

    # -----------------------------
    # collection routing
    # -----------------------------
    M1, F_col = _apply_matrix_rowonly(Remain_i, S_col, return_tensor=True)

    out.update({
        "SANK_col_g": float(np.asarray(Remain_i)[:, :, I_COL_G, 0].sum()),
        "SANK_col_m": float(np.asarray(Remain_i)[:, :, I_COL_M, 0].sum()),

        "SANK_col_g_to_sort_g": _sum_flow(F_col, I_COL_G, I_SORT_G),
        "SANK_col_g_to_sort_m": _sum_flow(F_col, I_COL_G, I_SORT_M),
        "SANK_col_g_to_ene":    _sum_flow(F_col, I_COL_G, IENE),
        "SANK_col_g_to_los":    _sum_flow(F_col, I_COL_G, ILOS),
        "SANK_col_g_to_dis":    _sum_flow(F_col, I_COL_G, IDIS),

        "SANK_col_m_to_sort_g": _sum_flow(F_col, I_COL_M, I_SORT_G),
        "SANK_col_m_to_sort_m": _sum_flow(F_col, I_COL_M, I_SORT_M),
        "SANK_col_m_to_ene":    _sum_flow(F_col, I_COL_M, IENE),
        "SANK_col_m_to_los":    _sum_flow(F_col, I_COL_M, ILOS),
        "SANK_col_m_to_dis":    _sum_flow(F_col, I_COL_M, IDIS),
    })

    # -----------------------------
    # sorting routing
    # -----------------------------
    M2a, F_sort_sep = _apply_matrix_rowonly(M1, S_sort_sep, return_tensor=True)
    M2,  F_sort_mix = _apply_matrix_rowonly(M2a, S_sort_mix, return_tensor=True)

    def _sum_both(Fa, Fb, i_from, i_to):
        return _sum_flow(Fa, i_from, i_to) + _sum_flow(Fb, i_from, i_to)

    out.update({
        "SANK_sort_g_to_rec_g": _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, I_REC_G),
        "SANK_sort_g_to_rec_m": _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, I_REC_M),
        "SANK_sort_g_to_ene":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, IENE),
        "SANK_sort_g_to_los":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, ILOS),
        "SANK_sort_g_to_dis":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_G, IDIS),

        "SANK_sort_m_to_rec_g": _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, I_REC_G),
        "SANK_sort_m_to_rec_m": _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, I_REC_M),
        "SANK_sort_m_to_ene":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, IENE),
        "SANK_sort_m_to_los":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, ILOS),
        "SANK_sort_m_to_dis":   _sum_both(F_sort_sep, F_sort_mix, I_SORT_M, IDIS),
    })

    # -----------------------------
    # Fraktion -> Process (MEC/CHE only; NO EV/LOS/DIS here)
    # -----------------------------
    M2_snap = np.array(M2, copy=True)

    M_frac_only = np.zeros_like(M2_snap)
    M_frac_only[:, :, I_REC_G, :] = M2_snap[:, :, I_REC_G, :]
    M_frac_only[:, :, I_REC_M, :] = M2_snap[:, :, I_REC_M, :]

    _, F_frac_to_proc = _apply_matrix_rowonly(M_frac_only, S_rec_proc, return_tensor=True)

    out.update({
        "SANK_rec_g_to_MEC": _sum_flow(F_frac_to_proc, I_REC_G, IMEC_G),
        "SANK_rec_g_to_CHE": _sum_flow(F_frac_to_proc, I_REC_G, ICHE_G),
        "SANK_rec_m_to_MEC": _sum_flow(F_frac_to_proc, I_REC_M, IMEC_M),
        "SANK_rec_m_to_CHE": _sum_flow(F_frac_to_proc, I_REC_M, ICHE_M),
    })

    # Process inflows
    out["MEC_in_sep"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_G, IMEC_G, 0].sum())
    out["CHE_in_sep"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_G, ICHE_G, 0].sum())
    out["MEC_in_mix"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_M, IMEC_M, 0].sum())
    out["CHE_in_mix"] = float(np.asarray(F_frac_to_proc)[:, :, I_REC_M, ICHE_M, 0].sum())

    out["MEC_in"] = float(out["MEC_in_sep"] + out["MEC_in_mix"])
    out["CHE_in"] = float(out["CHE_in_sep"] + out["CHE_in_mix"])

    # -----------------------------
    # exports at i (optional): p_mass_i_exports
    # -----------------------------
    if "p_mass_i_exports" in ParameterDict:
        X = np.asarray(_pv(ParameterDict, "p_mass_i_exports"), dtype=float)
        vec = X[..., 0] if X.ndim == 4 else X  # (t,r,i)
        M_i_export_e = np.clip(vec[..., None], 0, None)  # (t,r,i,1)
    else:
        M_i_export_e = np.zeros((T, Rr, I, 1), dtype=float)

    out["SANK_MEC_export_sep"] = float(M_i_export_e[:, :, IMEC_G, 0].sum())
    out["SANK_MEC_export_mix"] = float(M_i_export_e[:, :, IMEC_M, 0].sum())
    out["SANK_CHE_export_sep"] = float(M_i_export_e[:, :, ICHE_G, 0].sum())
    out["SANK_CHE_export_mix"] = float(M_i_export_e[:, :, ICHE_M, 0].sum())

    out["SANK_MEC_export"] = float(out["SANK_MEC_export_sep"] + out["SANK_MEC_export_mix"])
    out["SANK_CHE_export"] = float(out["SANK_CHE_export_sep"] + out["SANK_CHE_export_mix"])


    # -----------------------------
    # read p_share_recycling_fraction_to_process (t,r,i)  -> product share
    # -----------------------------
    frac_prod_tr_i = np.asarray(_pv(ParameterDict, "p_share_recycling_fraction_to_process"), dtype=float)
    if frac_prod_tr_i.ndim == 4:
        frac_prod_tr_i = frac_prod_tr_i[..., 0]  # (t,r,i)
    if frac_prod_tr_i.ndim != 3 or frac_prod_tr_i.shape[2] != I:
        raise TypeError(
            f"p_share_recycling_fraction_to_process must be (t,r,i). Got {frac_prod_tr_i.shape}, I={I}"
        )
    frac_prod_tr_i = np.clip(frac_prod_tr_i, 0.0, 1.0)

    def _fprod(i_idx):
        return float(frac_prod_tr_i[0, 0, i_idx])


    def _sout(i_from, j_to):
        return float(S_proc_out[0, 0, i_from, j_to, 0])

    def _get_abs_shares_for_process(i_idx):
        fprod = _fprod(i_idx)

        sev  = _sout(i_idx, IENE) if IENE is not None else 0.0
        slos = _sout(i_idx, ILOS) if ILOS is not None else 0.0
        sdis = _sout(i_idx, IDIS) if IDIS is not None else 0.0

        # clip
        fprod = float(np.clip(fprod, 0.0, 1.0))
        sev   = float(np.clip(sev,   0.0, 1.0))
        slos  = float(np.clip(slos,  0.0, 1.0))
        sdis  = float(np.clip(sdis,  0.0, 1.0))

        ssum = fprod + sev + slos + sdis

        if ssum < 1.0 - 1e-8:
            sdis = sdis + (1.0 - ssum)
            ssum = fprod + sev + slos + sdis

        if abs(ssum - 1.0) > 1e-8:
            if ssum > 1e-12:
                fprod /= ssum
                sev   /= ssum
                slos  /= ssum
                sdis  /= ssum
            else:
                fprod, sev, slos, sdis = 0.0, 0.0, 1.0, 0.0

        return fprod, sev, slos, sdis

    # -----------------------------
    # Apply: export-first accounting
    # -----------------------------
    def _apply_process_shares_export_first(proc_in, export_out, i_idx):
        proc_in = float(max(proc_in, 0.0))
        export_out = float(max(export_out, 0.0))
        export_used = min(export_out, proc_in)
        rem = proc_in - export_used

        fprod, sev, slos, sdis = _get_abs_shares_for_process(i_idx)

        prod = rem * fprod
        ev   = rem * sev
        los  = rem * slos
        dis  = rem * sdis

        # small numerical cleanup
        if prod < 0 and abs(prod) < 1e-9: prod = 0.0
        if ev   < 0 and abs(ev)   < 1e-9: ev   = 0.0
        if los  < 0 and abs(los)  < 1e-9: los  = 0.0
        if dis  < 0 and abs(dis)  < 1e-9: dis  = 0.0

        return float(prod), float(ev), float(los), float(dis), float(rem), float(export_used)

    # -----------------------------
    # MEC outputs (product = rPET)
    # -----------------------------
    mec_exp_sep = float(out["SANK_MEC_export_sep"])
    mec_exp_mix = float(out["SANK_MEC_export_mix"])
    che_exp_sep = float(out["SANK_CHE_export_sep"])
    che_exp_mix = float(out["SANK_CHE_export_mix"])

    # MEC getrennt
    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["MEC_in_sep"], mec_exp_sep, IMEC_G
    )
    out["rPET_from_MEC_sep"] = prod
    out["EV_from_MEC_sep"], out["LOS_from_MEC_sep"], out["DIS_from_MEC_sep"] = EV, LOS, DIS
    out["REM_after_export_MEC_sep"] = REM
    out["MEC_export_used_sep"] = XUSED

    # MEC gemischt
    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["MEC_in_mix"], mec_exp_mix, IMEC_M
    )
    out["rPET_from_MEC_mix"] = prod
    out["EV_from_MEC_mix"], out["LOS_from_MEC_mix"], out["DIS_from_MEC_mix"] = EV, LOS, DIS
    out["REM_after_export_MEC_mix"] = REM
    out["MEC_export_used_mix"] = XUSED

    out["rPET_total"] = float(out["rPET_from_MEC_sep"] + out["rPET_from_MEC_mix"])

    # -----------------------------
    # CHE outputs (product = chemsec / CHE-products)
    # -----------------------------
    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["CHE_in_sep"], che_exp_sep, ICHE_G
    )
    out["chemsec_from_CHE_sep"] = prod
    out["EV_from_CHE_sep"], out["LOS_from_CHE_sep"], out["DIS_from_CHE_sep"] = EV, LOS, DIS
    out["REM_after_export_CHE_sep"] = REM
    out["CHE_export_used_sep"] = XUSED

    prod, EV, LOS, DIS, REM, XUSED = _apply_process_shares_export_first(
        out["CHE_in_mix"], che_exp_mix, ICHE_M
    )
    out["chemsec_from_CHE_mix"] = prod
    out["EV_from_CHE_mix"], out["LOS_from_CHE_mix"], out["DIS_from_CHE_mix"] = EV, LOS, DIS
    out["REM_after_export_CHE_mix"] = REM
    out["CHE_export_used_mix"] = XUSED

    out["chemsec_from_CHE_total"] = float(out["chemsec_from_CHE_sep"] + out["chemsec_from_CHE_mix"])

    # explicit meaning: CHE output goes upstream to primary PET production (NOT to PM)
    out["chemsec_to_PrimPET_production_total"] = float(out["chemsec_from_CHE_total"])

    # totals
    out["EV_from_MEC_total"]   = float(out["EV_from_MEC_sep"]  + out["EV_from_MEC_mix"])
    out["LOS_from_MEC_total"]  = float(out["LOS_from_MEC_sep"] + out["LOS_from_MEC_mix"])
    out["DIS_from_MEC_total"]  = float(out["DIS_from_MEC_sep"] + out["DIS_from_MEC_mix"])

    out["EV_from_CHE_total"]   = float(out["EV_from_CHE_sep"]  + out["EV_from_CHE_mix"])
    out["LOS_from_CHE_total"]  = float(out["LOS_from_CHE_sep"] + out["LOS_from_CHE_mix"])
    out["DIS_from_CHE_total"]  = float(out["DIS_from_CHE_sep"] + out["DIS_from_CHE_mix"])

    # Optional sanity log (only if verbose)
    if verbose:
        def _chk(i_idx, label):
            fprod, sev, slos, sdis = _get_abs_shares_for_process(i_idx)
            Mylog.info(
                f"### SHARE CHECK {label}: fprod={fprod:.4f}, EV={sev:.4f}, LOS={slos:.4f}, DIS={sdis:.4f}, sum={fprod+sev+slos+sdis:.4f}"
            )
        _chk(IMEC_G, "MEC_getrennt")
        _chk(IMEC_M, "MEC_gemischt")
        _chk(ICHE_G, "CHE_getrennt")
        _chk(ICHE_M, "CHE_gemischt")

    # =========================================================
    # rPET routing to PM: FIXED MASS demand (bottle + otherpack)
    # =========================================================
    rpet_total = float(out["rPET_total"])

    if "p_mass_rPET_demand_bottle" in ParameterDict:
        rpet_dem_bottle = float(max(_scalar(_pv(ParameterDict, "p_mass_rPET_demand_bottle"), 0.0), 0.0))
    else:
        rpet_dem_bottle = 0.0
        Mylog.warning("p_mass_rPET_demand_bottle missing -> using 0.0")

    if "p_rPET_demand_otherpack_tr" in ParameterDict:
        rpet_dem_other = float(max(_scalar(_pv(ParameterDict, "p_rPET_demand_otherpack_tr"), 0.0), 0.0))
    else:
        rpet_dem_other = 0.0
        Mylog.warning("p_rPET_demand_otherpack_tr missing -> using 0.0")

    rpet_to_pm_fixed = float(max(rpet_dem_bottle + rpet_dem_other, 0.0))

    rpet_to_pm = float(min(rpet_total, rpet_to_pm_fixed))
    rpet_open  = float(max(rpet_total - rpet_to_pm, 0.0))
    rpet_def   = float(max(rpet_to_pm_fixed - rpet_total, 0.0))

    out["rPET_demand_bottle_fixed"] = float(rpet_dem_bottle)
    out["rPET_demand_otherpack_fixed"] = float(rpet_dem_other)
    out["rPET_to_PM_fixed_total"] = float(rpet_to_pm_fixed)

    out["rPET_to_PM_total"] = float(rpet_to_pm)
    out["rPET_openloop_total"] = float(rpet_open)
    out["rPET_surplus"] = float(rpet_open)
    out["rPET_deficit_for_PM"] = float(rpet_def)

    if "p_share_recycled_output_to_PM" in ParameterDict:
        Mylog.warning(
            "Parameter 'p_share_recycled_output_to_PM' is present but is NOT used for rPET->PM. "
            "It is routing/process-related, not PM demand."
        )

    # -----------------------------
    # Produktherstellung accounting (PM)
    # -----------------------------
    prim_raw   = float(out.get("PrimPET_to_PM_raw", 0.0))
    imp_prod   = float(out.get("product_import", 0.0))
    chemsec_raw_supply = float(out.get("chemsec_to_PM_raw", 0.0))  # external/raw chemsec supply only
    rpet_to_pm = float(out.get("rPET_to_PM_total", 0.0))
    exp_prod   = float(out.get("product_export", 0.0))
    PM_market  = float(out.get("PM_to_U_total", 0.0))

    PM_domestic = prim_raw + imp_prod + chemsec_raw_supply + rpet_to_pm - exp_prod
    PM_domestic = max(PM_domestic, 0.0)

    PM_rest = PM_domestic - PM_market
    if PM_rest < 0 and abs(PM_rest) < 1e-6:
        PM_rest = 0.0
    PM_rest = max(PM_rest, 0.0)

    out["PM_domestic_after_export"] = float(PM_domestic)
    out["PM_rest_loss"] = float(PM_rest)

    # -----------------------------
    # sanity checks
    # -----------------------------
    if out["rPET_total"] > out["MEC_in"] + 1e-6:
        Mylog.warning("### SANITY: rPET_total > MEC_in. Check yields or process inflow accounting.")
    if out["rPET_to_PM_total"] > out["rPET_total"] + 1e-6:
        Mylog.warning("### SANITY: rPET_to_PM_total > rPET_total. Check fixed demand parameters.")

    if verbose:
        Mylog.info(
            "### STEP7 CHECK: PM_domestic=%.3f | PM_market=%.3f | PM_rest=%.3f | Use_to_Waste=%.3f | Stock=%.3f | "
            "rPET_total=%.3f | rPET_to_PM=%.3f | rPET_open=%.3f | rPET_dem_bottle=%.3f | rPET_dem_other=%.3f | "
            "chemsec_to_PrimPET=%.3f | Prim_raw=%.3f | ProdImp=%.3f | ProdExp=%.3f"
            % (
                out["PM_domestic_after_export"],
                out["PM_to_U_total"],
                out["PM_rest_loss"],
                out["SANK_in_i_from_use"],
                out["DeltaStock_u_total"],
                out["rPET_total"],
                out["rPET_to_PM_total"],
                out["rPET_openloop_total"],
                out["rPET_demand_bottle_fixed"],
                out["rPET_demand_otherpack_fixed"],
                out["chemsec_to_PrimPET_production_total"],
                out["PrimPET_to_PM_raw"],
                out["product_import"],
                out["product_export"],
            )
        )

    return out


def export_step7_sankey_no_new_params(
    out7,
    out_dir="outputs",
    filename="sankey_Baseline_2019_Deutschland.html",
    font_size=16,
    use_short_labels=True,
):
    import numpy as np
    from pathlib import Path
    import plotly.graph_objects as go

    # -----------------------------
    # helpers
    # -----------------------------
    def _safe(x):
        try:
            v = float(x)
            return v if v > 0 else 0.0
        except Exception:
            return 0.0

    def _safe_signed(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    def hex_to_rgba(hex_color: str, a: float = 0.35) -> str:
        h = hex_color.lstrip("#")
        r = int(h[0:2], 16); g = int(h[2:4], 16); b = int(h[4:6], 16)
        return f"rgba({r},{g},{b},{a})"

    def _split_after_export(inflow, export, ev_val, los_val, rpet_val, eps=1e-12):

        inflow = float(max(inflow, 0.0))
        export = float(max(export, 0.0))

        export_used = min(export, inflow)
        remaining = inflow - export_used

        ev_val   = float(max(ev_val, 0.0))
        los_val  = float(max(los_val, 0.0))
        rpet_val = float(max(rpet_val, 0.0))

        if inflow <= eps:
            return (0.0, 0.0, 0.0, 0.0, 0.0)

        s_ev  = ev_val / inflow
        s_los = los_val / inflow
        s_rp  = rpet_val / inflow

        s_ev  = min(max(s_ev,  0.0), 1.0)
        s_los = min(max(s_los, 0.0), 1.0)
        s_rp  = min(max(s_rp,  0.0), 1.0)

        s_sum = s_ev + s_los + s_rp
        if s_sum <= eps:
            return (export_used, 0.0, 0.0, 0.0, remaining)

        # normalize shares over remaining (export removed)
        s_ev  /= s_sum
        s_los /= s_sum
        s_rp  /= s_sum

        EV  = remaining * s_ev
        LOS = remaining * s_los
        RP  = remaining * s_rp

        residual = max(remaining - (EV + LOS + RP), 0.0)
        if residual < 1e-9:
            residual = 0.0

        return (export_used, EV, LOS, RP, residual)

    # -----------------------------
    # Label variants (short vs long)
    # -----------------------------
    if use_short_labels:
        SEP_TAG  = "Pfandflaschen"
        MIX_TAG  = "andere PET-Verp."
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"
    else:
        SEP_TAG  = "PET-Getränkeflaschen aus Pfand"
        MIX_TAG  = "andere PET-Verpackungen"
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"

    # -----------------------------
    # Canonical node names
    # -----------------------------
    N_PRIM   = "Produktion von Primär-PET"
    N_IMP    = "Importe (Produkte)"
    N_CHESEC = "Sekundäre Rohstoffe (CHE) / sonstige Inputs"
    N_RPET   = "rPET (Pool)"
    N_PM     = "Produktherstellung"
    N_EXP    = "Exporte (Produkte)"


    N_PMREST = "Restabgänge (Produktherstellung) – Export/Bestand/Sonstiges"

    N_MARKET = "Verbrauch (Markt)"
    N_SAMM   = "Sammlung Fraktion"
    N_STOCK  = "Bestand + Unbekannter Anteil"

    N_COL_SEP = f"{SAMM_TAG}_{SEP_TAG}"
    N_COL_MIX = f"{SAMM_TAG}_{MIX_TAG}"

    N_SORT_SEP = f"{SORT_TAG}_{SEP_TAG}"
    N_SORT_MIX = f"{SORT_TAG}_{MIX_TAG}"

    N_RECFR_SEP = f"{REC_TAG}_{SEP_TAG}"
    N_RECFR_MIX = f"{REC_TAG}_{MIX_TAG}"

    N_EV_COL_SEP  = f"EV ({SAMM_TAG}_{SEP_TAG})"
    N_LOS_COL_SEP = f"Verluste ({SAMM_TAG}_{SEP_TAG})"
    N_DIS_COL_SEP = f"Beseitigung ({SAMM_TAG}_{SEP_TAG})"

    N_EV_COL_MIX  = f"EV ({SAMM_TAG}_{MIX_TAG})"
    N_LOS_COL_MIX = f"Verluste ({SAMM_TAG}_{MIX_TAG})"
    N_DIS_COL_MIX = f"Beseitigung ({SAMM_TAG}_{MIX_TAG})"

    N_EV_SORT_SEP  = f"EV ({SORT_TAG}_{SEP_TAG})"
    N_LOS_SORT_SEP = f"Verluste ({SORT_TAG}_{SEP_TAG})"
    N_DIS_SORT_SEP = f"Beseitigung ({SORT_TAG}_{SEP_TAG})"

    N_EV_SORT_MIX  = f"EV ({SORT_TAG}_{MIX_TAG})"
    N_LOS_SORT_MIX = f"Verluste ({SORT_TAG}_{MIX_TAG})"
    N_DIS_SORT_MIX = f"Beseitigung ({SORT_TAG}_{MIX_TAG})"

    N_MEC_SEP = f"Mechanisches Recycling (MEC, {SEP_TAG})"
    N_MEC_MIX = f"Mechanisches Recycling (MEC, {MIX_TAG})"
    N_CHE_SEP = f"Chemisches Recycling (CHE, {SEP_TAG})"
    N_CHE_MIX = f"Chemisches Recycling (CHE, {MIX_TAG})"

    N_EV_MEC_SEP  = f"EV (MEC, {SEP_TAG})"
    N_LOS_MEC_SEP = f"Verluste (MEC, {SEP_TAG})"
    N_MEC_OTH_SEP = f"MEC-Rückstände / Sonstiges ({SEP_TAG})"

    N_EV_MEC_MIX  = f"EV (MEC, {MIX_TAG})"
    N_LOS_MEC_MIX = f"Verluste (MEC, {MIX_TAG})"
    N_MEC_OTH_MIX = f"MEC-Rückstände / Sonstiges ({MIX_TAG})"

    N_EV_CHE_SEP   = f"EV (CHE, {SEP_TAG})"
    N_LOS_CHE_SEP  = f"Verluste (CHE, {SEP_TAG})"
    N_CHE_PROD_SEP = f"CHE-Produkte / open-loop ({SEP_TAG})"

    N_EV_CHE_MIX   = f"EV (CHE, {MIX_TAG})"
    N_LOS_CHE_MIX  = f"Verluste (CHE, {MIX_TAG})"
    N_CHE_PROD_MIX = f"CHE-Produkte / open-loop ({MIX_TAG})"

    N_RPET_OPEN  = "rPET open-loop"
    N_REC_EXPORT = "Exporte (Recyclingprodukte)"

    # -----------------------------
    # Pull values from out7
    # -----------------------------
    PM_market   = _safe(out7.get("PM_to_U_total", 0.0))
    waste_total = _safe(out7.get("SANK_in_i_from_use", 0.0))
    stock_delta = _safe_signed(PM_market - waste_total)

    imp_prod = _safe(out7.get("product_import", 0.0))
    exp_prod = _safe(out7.get("product_export", 0.0))

    prim_raw = _safe(out7.get("PrimPET_to_PM_raw", 0.0))

    chemsec  = _safe(out7.get("chemsec_to_PM", out7.get("chemsec_to_PM_raw", 0.0)))

    rPET_to_PM_raw = _safe(out7.get("rPET_to_PM_total", 0.0))

    PM_domestic = _safe(out7.get(
        "PM_domestic_after_export",
        max(prim_raw + imp_prod + chemsec + rPET_to_PM_raw - exp_prod, 0.0)
    ))
    PM_rest = _safe(out7.get(
        "PM_rest_loss",
        max(PM_domestic - PM_market, 0.0)
    ))

    # Collection totals
    col_sep = _safe(out7.get("SANK_col_g", 0.0))
    col_mix = _safe(out7.get("SANK_col_m", 0.0))

    # Collection splits
    col_sep_to_sort_sep = _safe(out7.get("SANK_col_g_to_sort_g", 0.0))
    col_sep_to_sort_mix = _safe(out7.get("SANK_col_g_to_sort_m", 0.0))
    col_sep_to_ene      = _safe(out7.get("SANK_col_g_to_ene", 0.0))
    col_sep_to_los      = _safe(out7.get("SANK_col_g_to_los", 0.0))
    col_sep_to_dis      = _safe(out7.get("SANK_col_g_to_dis", 0.0))

    col_mix_to_sort_sep = _safe(out7.get("SANK_col_m_to_sort_g", 0.0))
    col_mix_to_sort_mix = _safe(out7.get("SANK_col_m_to_sort_m", 0.0))
    col_mix_to_ene      = _safe(out7.get("SANK_col_m_to_ene", 0.0))
    col_mix_to_los      = _safe(out7.get("SANK_col_m_to_los", 0.0))
    col_mix_to_dis      = _safe(out7.get("SANK_col_m_to_dis", 0.0))

    # Sorting splits
    sort_sep_to_rec_sep = _safe(out7.get("SANK_sort_g_to_rec_g", 0.0))
    sort_sep_to_rec_mix = _safe(out7.get("SANK_sort_g_to_rec_m", 0.0))
    sort_sep_to_ene     = _safe(out7.get("SANK_sort_g_to_ene", 0.0))
    sort_sep_to_los     = _safe(out7.get("SANK_sort_g_to_los", 0.0))
    sort_sep_to_dis     = _safe(out7.get("SANK_sort_g_to_dis", 0.0))

    sort_mix_to_rec_sep = _safe(out7.get("SANK_sort_m_to_rec_g", 0.0))
    sort_mix_to_rec_mix = _safe(out7.get("SANK_sort_m_to_rec_m", 0.0))
    sort_mix_to_ene     = _safe(out7.get("SANK_sort_m_to_ene", 0.0))
    sort_mix_to_los     = _safe(out7.get("SANK_sort_m_to_los", 0.0))
    sort_mix_to_dis     = _safe(out7.get("SANK_sort_m_to_dis", 0.0))

    # Recycling fractions -> processes
    rec_sep_to_MEC = _safe(out7.get("SANK_rec_g_to_MEC", 0.0))
    rec_sep_to_CHE = _safe(out7.get("SANK_rec_g_to_CHE", 0.0))
    rec_mix_to_MEC = _safe(out7.get("SANK_rec_m_to_MEC", 0.0))
    rec_mix_to_CHE = _safe(out7.get("SANK_rec_m_to_CHE", 0.0))

    MEC_in_sep = _safe(out7.get("MEC_in_sep", rec_sep_to_MEC))
    MEC_in_mix = _safe(out7.get("MEC_in_mix", rec_mix_to_MEC))
    CHE_in_sep = _safe(out7.get("CHE_in_sep", rec_sep_to_CHE))
    CHE_in_mix = _safe(out7.get("CHE_in_mix", rec_mix_to_CHE))

    MEC_export_sep = _safe(out7.get("SANK_MEC_export_sep", 0.0))
    MEC_export_mix = _safe(out7.get("SANK_MEC_export_mix", 0.0))
    CHE_export_sep = _safe(out7.get("SANK_CHE_export_sep", 0.0))
    CHE_export_mix = _safe(out7.get("SANK_CHE_export_mix", 0.0))

    EV_MEC_sep_raw  = _safe(out7.get("EV_from_MEC_sep", 0.0))
    LOS_MEC_sep_raw = _safe(out7.get("LOS_from_MEC_sep", 0.0))
    EV_MEC_mix_raw  = _safe(out7.get("EV_from_MEC_mix", 0.0))
    LOS_MEC_mix_raw = _safe(out7.get("LOS_from_MEC_mix", 0.0))

    EV_CHE_sep_raw  = _safe(out7.get("EV_from_CHE_sep", 0.0))
    LOS_CHE_sep_raw = _safe(out7.get("LOS_from_CHE_sep", 0.0))
    EV_CHE_mix_raw  = _safe(out7.get("EV_from_CHE_mix", 0.0))
    LOS_CHE_mix_raw = _safe(out7.get("LOS_from_CHE_mix", 0.0))

    rPET_from_MEC_sep_raw = _safe(out7.get("rPET_from_MEC_sep", 0.0))
    rPET_from_MEC_mix_raw = _safe(out7.get("rPET_from_MEC_mix", 0.0))
    rPET_from_CHE_sep_raw = _safe(out7.get("rPET_from_CHE_sep", 0.0))
    rPET_from_CHE_mix_raw = _safe(out7.get("rPET_from_CHE_mix", 0.0))

    MEC_export_sep_u, EV_MEC_sep, LOS_MEC_sep, rPET_from_MEC_sep, MEC_other_sep = _split_after_export(
        MEC_in_sep, MEC_export_sep, EV_MEC_sep_raw, LOS_MEC_sep_raw, rPET_from_MEC_sep_raw
    )
    MEC_export_mix_u, EV_MEC_mix, LOS_MEC_mix, rPET_from_MEC_mix, MEC_other_mix = _split_after_export(
        MEC_in_mix, MEC_export_mix, EV_MEC_mix_raw, LOS_MEC_mix_raw, rPET_from_MEC_mix_raw
    )

    # ============================================================
    # CHE outputs: ONLY EV (parameter-based) + open-loop (rest)
    # ============================================================

    CHE_export_sep_u = min(CHE_export_sep, CHE_in_sep)
    CHE_export_mix_u = min(CHE_export_mix, CHE_in_mix)

    rem_che_sep = max(CHE_in_sep - CHE_export_sep_u, 0.0)
    rem_che_mix = max(CHE_in_mix - CHE_export_mix_u, 0.0)

    EV_CHE_sep = min(EV_CHE_sep_raw, rem_che_sep)
    EV_CHE_mix = min(EV_CHE_mix_raw, rem_che_mix)

    LOS_CHE_sep = 0.0
    LOS_CHE_mix = 0.0

    rPET_from_CHE_sep = 0.0
    rPET_from_CHE_mix = 0.0

    CHE_products_sep = max(rem_che_sep - EV_CHE_sep, 0.0)
    CHE_products_mix = max(rem_che_mix - EV_CHE_mix, 0.0)

    # ---- rPET POOL (POST-EXPORT) ----
    rPET_total = float(rPET_from_MEC_sep + rPET_from_MEC_mix + rPET_from_CHE_sep + rPET_from_CHE_mix)
    rPET_to_PM = min(float(rPET_to_PM_raw), float(rPET_total))

    rPET_open  = max(float(rPET_total) - float(rPET_to_PM), 0.0)

    # -----------------------------
    # Nodes
    # -----------------------------
    nodes = [
        N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM, N_EXP, N_PMREST, N_MARKET, N_SAMM, N_STOCK,
        N_COL_SEP, N_COL_MIX,
        N_SORT_SEP, N_SORT_MIX,
        N_RECFR_SEP, N_RECFR_MIX,
        N_EV_COL_SEP, N_LOS_COL_SEP, N_DIS_COL_SEP,
        N_EV_COL_MIX, N_LOS_COL_MIX, N_DIS_COL_MIX,
        N_EV_SORT_SEP, N_LOS_SORT_SEP, N_DIS_SORT_SEP,
        N_EV_SORT_MIX, N_LOS_SORT_MIX, N_DIS_SORT_MIX,
        N_MEC_SEP, N_MEC_MIX, N_CHE_SEP, N_CHE_MIX,
        N_EV_MEC_SEP, N_LOS_MEC_SEP, N_MEC_OTH_SEP,
        N_EV_MEC_MIX, N_LOS_MEC_MIX, N_MEC_OTH_MIX,
        N_EV_CHE_SEP, N_LOS_CHE_SEP, N_CHE_PROD_SEP,
        N_EV_CHE_MIX, N_LOS_CHE_MIX, N_CHE_PROD_MIX,
        N_RPET_OPEN, N_REC_EXPORT
    ]
    idx = {n: i for i, n in enumerate(nodes)}

    # -----------------------------
    # Category colors
    # -----------------------------
    COL = {
        "prim_pm": "#1C4E80",
        "import":  "#5A7FB5",
        "export":  "#2E3440",
        "market":  "#4C566A",
        "sammlung":"#2F8F9D",
        "sort":    "#6B8CAF",
        "recfr":   "#3E6D9C",
        "mec":     "#2C6E9B",
        "che":     "#6A5D8C",
        "ev":      "#1F6F78",
        "loss":    "#111827",
        "dis":     "#374151",
        "rpet":    "#D4A106",
        "stock":   "#D1D5DB",
    }

    def node_color(name: str) -> str:
        if name in (N_PRIM, N_PM):
            return COL["prim_pm"]
        if name == N_IMP:
            return COL["import"]
        if name in (N_EXP, N_REC_EXPORT):
            return COL["export"]
        if name == N_MARKET:
            return COL["market"]
        if name in (N_SAMM, N_COL_SEP, N_COL_MIX):
            return COL["sammlung"]
        if name in (N_SORT_SEP, N_SORT_MIX):
            return COL["sort"]
        if name in (N_RECFR_SEP, N_RECFR_MIX):
            return COL["recfr"]
        if name in (N_MEC_SEP, N_MEC_MIX, N_MEC_OTH_SEP, N_MEC_OTH_MIX):
            return COL["mec"]
        if name in (N_CHE_SEP, N_CHE_MIX, N_CHE_PROD_SEP, N_CHE_PROD_MIX, N_CHESEC):
            return COL["che"]
        if name.startswith("EV"):
            return COL["ev"]
        if name.startswith("Verluste"):
            return COL["loss"]
        if name.startswith("Beseitigung"):
            return COL["dis"]
        if name in (N_RPET, N_RPET_OPEN):
            return COL["rpet"]
        if name in (N_PMREST, N_STOCK):
            return COL["stock"]
        return COL["stock"]

    node_colors = [node_color(n) for n in nodes]

    # -----------------------------
    # Links
    # -----------------------------
    src, trg, val, link_text, link_color = [], [], [], [], []

    def L(a, b, v):
        v = float(v)
        if v <= 0:
            return
        if (a not in idx) or (b not in idx):
            missing = []
            if a not in idx: missing.append(a)
            if b not in idx: missing.append(b)
            raise KeyError(f"Sankey node name not found: {missing}")

        src.append(idx[a])
        trg.append(idx[b])
        val.append(v)
        link_text.append(f"{v:.3f} kt")
        link_color.append(hex_to_rgba(node_color(b), 0.35))

    # Inflows to Produktherstellung
    L(N_PRIM,   N_PM, prim_raw)
    L(N_IMP,    N_PM, imp_prod)
    L(N_CHESEC, N_PM, chemsec)
    L(N_RPET,   N_PM, rPET_to_PM)

    # Outflows from Produktherstellung
    L(N_PM, N_EXP,    exp_prod)
    L(N_PM, N_MARKET, PM_market)
    L(N_PM, N_PMREST, PM_rest)

    # Market -> waste + stock
    L(N_MARKET, N_SAMM, waste_total)
    if stock_delta > 0:
        L(N_MARKET, N_STOCK, stock_delta)

    # Waste -> collection fractions
    L(N_SAMM, N_COL_SEP, col_sep)
    L(N_SAMM, N_COL_MIX, col_mix)

    # Collection splits (sep)
    L(N_COL_SEP, N_SORT_SEP,    col_sep_to_sort_sep)
    L(N_COL_SEP, N_SORT_MIX,    col_sep_to_sort_mix)
    L(N_COL_SEP, N_EV_COL_SEP,  col_sep_to_ene)
    L(N_COL_SEP, N_LOS_COL_SEP, col_sep_to_los)
    L(N_COL_SEP, N_DIS_COL_SEP, col_sep_to_dis)

    # Collection splits (mix)
    L(N_COL_MIX, N_SORT_SEP,    col_mix_to_sort_sep)
    L(N_COL_MIX, N_SORT_MIX,    col_mix_to_sort_mix)
    L(N_COL_MIX, N_EV_COL_MIX,  col_mix_to_ene)
    L(N_COL_MIX, N_LOS_COL_MIX, col_mix_to_los)
    L(N_COL_MIX, N_DIS_COL_MIX, col_mix_to_dis)

    # Sorting splits (sep sorting node)
    L(N_SORT_SEP,  N_RECFR_SEP,    sort_sep_to_rec_sep)
    L(N_SORT_SEP,  N_RECFR_MIX,    sort_sep_to_rec_mix)
    L(N_SORT_SEP,  N_EV_SORT_SEP,  sort_sep_to_ene)
    L(N_SORT_SEP,  N_LOS_SORT_SEP, sort_sep_to_los)
    L(N_SORT_SEP,  N_DIS_SORT_SEP, sort_sep_to_dis)

    # Sorting splits (mix sorting node)
    L(N_SORT_MIX,  N_RECFR_SEP,    sort_mix_to_rec_sep)
    L(N_SORT_MIX,  N_RECFR_MIX,    sort_mix_to_rec_mix)
    L(N_SORT_MIX,  N_EV_SORT_MIX,  sort_mix_to_ene)
    L(N_SORT_MIX,  N_LOS_SORT_MIX, sort_mix_to_los)
    L(N_SORT_MIX,  N_DIS_SORT_MIX, sort_mix_to_dis)

    # Recycling fractions -> processes
    L(N_RECFR_SEP, N_MEC_SEP, rec_sep_to_MEC)
    L(N_RECFR_SEP, N_CHE_SEP, rec_sep_to_CHE)
    L(N_RECFR_MIX, N_MEC_MIX, rec_mix_to_MEC)
    L(N_RECFR_MIX, N_CHE_MIX, rec_mix_to_CHE)

    # MEC outputs (balanced)
    L(N_MEC_SEP, N_REC_EXPORT,  MEC_export_sep_u)
    L(N_MEC_SEP, N_RPET,        rPET_from_MEC_sep)
    L(N_MEC_SEP, N_EV_MEC_SEP,  EV_MEC_sep)
    L(N_MEC_SEP, N_LOS_MEC_SEP, LOS_MEC_sep)
    L(N_MEC_SEP, N_MEC_OTH_SEP, MEC_other_sep)

    L(N_MEC_MIX, N_REC_EXPORT,  MEC_export_mix_u)
    L(N_MEC_MIX, N_RPET,        rPET_from_MEC_mix)
    L(N_MEC_MIX, N_EV_MEC_MIX,  EV_MEC_mix)
    L(N_MEC_MIX, N_LOS_MEC_MIX, LOS_MEC_mix)
    L(N_MEC_MIX, N_MEC_OTH_MIX, MEC_other_mix)

    # CHE outputs (balanced)
    L(N_CHE_SEP, N_REC_EXPORT,   CHE_export_sep_u)
    L(N_CHE_SEP, N_RPET,         rPET_from_CHE_sep)
    L(N_CHE_SEP, N_EV_CHE_SEP,   EV_CHE_sep)
    L(N_CHE_SEP, N_LOS_CHE_SEP,  LOS_CHE_sep)
    L(N_CHE_SEP, N_CHE_PROD_SEP, CHE_products_sep)

    L(N_CHE_MIX, N_REC_EXPORT,   CHE_export_mix_u)
    L(N_CHE_MIX, N_RPET,         rPET_from_CHE_mix)
    L(N_CHE_MIX, N_EV_CHE_MIX,   EV_CHE_mix)
    L(N_CHE_MIX, N_LOS_CHE_MIX,  LOS_CHE_mix)
    L(N_CHE_MIX, N_CHE_PROD_MIX, CHE_products_mix)

    # rPET pool -> open-loop (by definition)
    L(N_RPET, N_RPET_OPEN, rPET_open)

    # -----------------------------
    # Node totals + labels
    # -----------------------------
    Nn = len(nodes)
    infl = np.zeros(Nn); outf = np.zeros(Nn)
    for s, t, v in zip(src, trg, val):
        outf[s] += v
        infl[t] += v

    totals = np.where(outf > 0, outf, infl)
    totals[idx[N_PM]]   = float(PM_domestic)
    totals[idx[N_RPET]] = float(rPET_total)

    labels = []
    for i, name in enumerate(nodes):
        if name == N_RPET:
            labels.append(
                f"{name}\n({totals[i]:.3f} kt)\n[rPET_total={rPET_total:.3f} | toPM={rPET_to_PM:.3f} | open={rPET_open:.3f}]"
            )
        else:
            labels.append(f"{name}\n({totals[i]:.3f} kt)")

    # -----------------------------
    # FIGURE
    # -----------------------------
    fig = go.Figure(go.Sankey(
        arrangement="freeform",
        node=dict(
            label=labels,
            pad=18,
            thickness=20,
            color=node_colors,
            line=dict(color="rgba(0,0,0,0.25)", width=0.5),
        ),
        link=dict(
            source=src,
            target=trg,
            value=val,
            color=link_color,
            customdata=link_text,
            hovertemplate="%{source.label} → %{target.label}<br>%{customdata}<extra></extra>",
        ),
    ))
    fig.update_layout(
        title="PET/Polymer Verpackungen – Deutschland 2019, MFA Baseline (kt)",
        font=dict(size=font_size),
    )

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    html_file = out_path / filename
    fig.write_html(str(html_file), config={"displayModeBar": True})

    return str(html_file.resolve())



def export_step7_nodes_links_excel_Baseline(
    out7,
    out_dir="outputs",
    excel_filename="nodes_links_Baseline.xlsx",
    use_short_labels=True,
):
    import numpy as np
    import pandas as pd
    from pathlib import Path

    def _safe(x):
        try:
            v = float(x)
            return v if np.isfinite(v) and v > 0 else 0.0
        except Exception:
            return 0.0

    def _safe_signed(x):
        try:
            v = float(x)
            return v if np.isfinite(v) else 0.0
        except Exception:
            return 0.0

    # ----------------------------
    # Labels
    # ----------------------------
    if use_short_labels:
        SEP_TAG  = "Pfandflaschen"
        MIX_TAG  = "andere PET-Verp."
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"
    else:
        SEP_TAG  = "PET-Getränkeflaschen aus Pfand"
        MIX_TAG  = "andere PET-Verpackungen"
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"

    N_PRIM   = "Produktion von Primär-PET"
    N_IMP    = "Importe (Produkte)"
    N_CHESEC = "Sekundär-Rohstoffe"
    N_RPET   = "rPET (Pool)"
    N_PM     = "Produktherstellung"
    N_EXP    = "Exporte (Produkte)"
    N_PMREST = "Restabgänge (Produktherstellung) – Export/Bestand/Sonstiges"
    N_MARKET = "Verbrauch (Markt)"
    N_SAMM   = "Sammlung Fraktion"
    N_STOCK  = "Bestand + unbekannter Anteil "

    N_COL_SEP = f"{SAMM_TAG}_{SEP_TAG}"
    N_COL_MIX = f"{SAMM_TAG}_{MIX_TAG}"

    N_SORT_SEP = f"{SORT_TAG}_{SEP_TAG}"
    N_SORT_MIX = f"{SORT_TAG}_{MIX_TAG}"

    N_RECFR_SEP = f"{REC_TAG}_{SEP_TAG}"
    N_RECFR_MIX = f"{REC_TAG}_{MIX_TAG}"

    N_EV_COL_SEP  = f"EV ({SAMM_TAG}_{SEP_TAG})"
    N_LOS_COL_SEP = f"Verluste ({SAMM_TAG}_{SEP_TAG})"
    N_DIS_COL_SEP = f"Beseitigung ({SAMM_TAG}_{SEP_TAG})"

    N_EV_COL_MIX  = f"EV ({SAMM_TAG}_{MIX_TAG})"
    N_LOS_COL_MIX = f"Verluste ({SAMM_TAG}_{MIX_TAG})"
    N_DIS_COL_MIX = f"Beseitigung ({SAMM_TAG}_{MIX_TAG})"

    N_EV_SORT_SEP  = f"EV ({SORT_TAG}_{SEP_TAG})"
    N_LOS_SORT_SEP = f"Verluste ({SORT_TAG}_{SEP_TAG})"
    N_DIS_SORT_SEP = f"Beseitigung ({SORT_TAG}_{SEP_TAG})"

    N_EV_SORT_MIX  = f"EV ({SORT_TAG}_{MIX_TAG})"
    N_LOS_SORT_MIX = f"Verluste ({SORT_TAG}_{MIX_TAG})"
    N_DIS_SORT_MIX = f"Beseitigung ({SORT_TAG}_{MIX_TAG})"

    N_MEC_SEP = f"Mechanisches Recycling (MEC, {SEP_TAG})"
    N_MEC_MIX = f"Mechanisches Recycling (MEC, {MIX_TAG})"
    N_CHE_SEP = f"Chemisches Recycling (CHE, {SEP_TAG})"
    N_CHE_MIX = f"Chemisches Recycling (CHE, {MIX_TAG})"

    N_EV_MEC_SEP  = f"EV (MEC, {SEP_TAG})"
    N_LOS_MEC_SEP = f"Verluste (MEC, {SEP_TAG})"
    N_MEC_OTH_SEP = f"MEC-Rückstände / Sonstiges ({SEP_TAG})"

    N_EV_MEC_MIX  = f"EV (MEC, {MIX_TAG})"
    N_LOS_MEC_MIX = f"Verluste (MEC, {MIX_TAG})"
    N_MEC_OTH_MIX = f"MEC-Rückstände / Sonstiges ({MIX_TAG})"

    N_EV_CHE_SEP  = f"EV (CHE, {SEP_TAG})"
    N_LOS_CHE_SEP = f"Verluste (CHE, {SEP_TAG})"
    N_CHE_PROD_SEP = f"CHE-Produkte / open-loop ({SEP_TAG})"

    N_EV_CHE_MIX  = f"EV (CHE, {MIX_TAG})"
    N_LOS_CHE_MIX = f"Verluste (CHE, {MIX_TAG})"
    N_CHE_PROD_MIX = f"CHE-Produkte / open-loop ({MIX_TAG})"

    N_RPET_OPEN  = "rPET open-loop"
    N_REC_EXPORT = "Exporte (Recyclingprodukte)"

    # ----------------------------
    # Values from out7 (same as Sankey base)
    # ----------------------------
    PM_market   = _safe(out7.get("PM_to_U_total", 0.0))
    waste_total = _safe(out7.get("SANK_in_i_from_use", 0.0))
    stock_delta = _safe_signed(PM_market - waste_total)

    imp_prod = _safe(out7.get("product_import", 0.0))
    exp_prod = _safe(out7.get("product_export", 0.0))

    prim_raw = _safe(out7.get("PrimPET_to_PM_raw", 0.0))
    chemsec  = _safe(out7.get("chemsec_to_PM", out7.get("chemsec_to_PM_raw", 0.0)))
    rPET_to_PM = _safe(out7.get("rPET_to_PM_total", 0.0))

    PM_domestic = _safe(out7.get(
        "PM_domestic_after_export",
        max(prim_raw + imp_prod + chemsec + rPET_to_PM - exp_prod, 0.0)
    ))
    PM_rest = _safe(out7.get("PM_rest_loss", max(PM_domestic - PM_market, 0.0)))

    col_sep = _safe(out7.get("SANK_col_g", 0.0))
    col_mix = _safe(out7.get("SANK_col_m", 0.0))

    col_sep_to_sort_sep = _safe(out7.get("SANK_col_g_to_sort_g", 0.0))
    col_sep_to_sort_mix = _safe(out7.get("SANK_col_g_to_sort_m", 0.0))
    col_sep_to_ene      = _safe(out7.get("SANK_col_g_to_ene", 0.0))
    col_sep_to_los      = _safe(out7.get("SANK_col_g_to_los", 0.0))
    col_sep_to_dis      = _safe(out7.get("SANK_col_g_to_dis", 0.0))

    col_mix_to_sort_sep = _safe(out7.get("SANK_col_m_to_sort_g", 0.0))
    col_mix_to_sort_mix = _safe(out7.get("SANK_col_m_to_sort_m", 0.0))
    col_mix_to_ene      = _safe(out7.get("SANK_col_m_to_ene", 0.0))
    col_mix_to_los      = _safe(out7.get("SANK_col_m_to_los", 0.0))
    col_mix_to_dis      = _safe(out7.get("SANK_col_m_to_dis", 0.0))

    sort_sep_to_rec_sep = _safe(out7.get("SANK_sort_g_to_rec_g", 0.0))
    sort_sep_to_rec_mix = _safe(out7.get("SANK_sort_g_to_rec_m", 0.0))
    sort_sep_to_ene     = _safe(out7.get("SANK_sort_g_to_ene", 0.0))
    sort_sep_to_los     = _safe(out7.get("SANK_sort_g_to_los", 0.0))
    sort_sep_to_dis     = _safe(out7.get("SANK_sort_g_to_dis", 0.0))

    sort_mix_to_rec_sep = _safe(out7.get("SANK_sort_m_to_rec_g", 0.0))
    sort_mix_to_rec_mix = _safe(out7.get("SANK_sort_m_to_rec_m", 0.0))
    sort_mix_to_ene     = _safe(out7.get("SANK_sort_m_to_ene", 0.0))
    sort_mix_to_los     = _safe(out7.get("SANK_sort_m_to_los", 0.0))
    sort_mix_to_dis     = _safe(out7.get("SANK_sort_m_to_dis", 0.0))

    rec_sep_to_MEC = _safe(out7.get("SANK_rec_g_to_MEC", 0.0))
    rec_sep_to_CHE = _safe(out7.get("SANK_rec_g_to_CHE", 0.0))
    rec_mix_to_MEC = _safe(out7.get("SANK_rec_m_to_MEC", 0.0))
    rec_mix_to_CHE = _safe(out7.get("SANK_rec_m_to_CHE", 0.0))

    MEC_in_sep = _safe(out7.get("MEC_in_sep", rec_sep_to_MEC))
    MEC_in_mix = _safe(out7.get("MEC_in_mix", rec_mix_to_MEC))
    CHE_in_sep = _safe(out7.get("CHE_in_sep", rec_sep_to_CHE))
    CHE_in_mix = _safe(out7.get("CHE_in_mix", rec_mix_to_CHE))

    MEC_export_sep = _safe(out7.get("SANK_MEC_export_sep", 0.0))
    MEC_export_mix = _safe(out7.get("SANK_MEC_export_mix", 0.0))
    CHE_export_sep = _safe(out7.get("SANK_CHE_export_sep", 0.0))
    CHE_export_mix = _safe(out7.get("SANK_CHE_export_mix", 0.0))

    # MEC outputs (post Step7)
    EV_MEC_sep   = _safe(out7.get("EV_from_MEC_sep", 0.0))
    LOS_MEC_sep  = _safe(out7.get("LOS_from_MEC_sep", 0.0))
    rPET_MEC_sep = _safe(out7.get("rPET_from_MEC_sep", 0.0))

    EV_MEC_mix   = _safe(out7.get("EV_from_MEC_mix", 0.0))
    LOS_MEC_mix  = _safe(out7.get("LOS_from_MEC_mix", 0.0))
    rPET_MEC_mix = _safe(out7.get("rPET_from_MEC_mix", 0.0))

    REM_MEC_sep = _safe(out7.get("REM_after_export_MEC_sep", max(MEC_in_sep - min(MEC_export_sep, MEC_in_sep), 0.0)))
    REM_MEC_mix = _safe(out7.get("REM_after_export_MEC_mix", max(MEC_in_mix - min(MEC_export_mix, MEC_in_mix), 0.0)))
    MEC_OTH_sep = max(REM_MEC_sep - (EV_MEC_sep + LOS_MEC_sep + rPET_MEC_sep), 0.0)
    MEC_OTH_mix = max(REM_MEC_mix - (EV_MEC_mix + LOS_MEC_mix + rPET_MEC_mix), 0.0)

    MEC_export_sep_u = min(MEC_export_sep, MEC_in_sep)
    MEC_export_mix_u = min(MEC_export_mix, MEC_in_mix)
    CHE_export_sep_u = min(CHE_export_sep, CHE_in_sep)
    CHE_export_mix_u = min(CHE_export_mix, CHE_in_mix)

    # CHE outputs (baseline-style keys)
    EV_CHE_sep   = _safe(out7.get("EV_from_CHE_sep", 0.0))
    LOS_CHE_sep  = _safe(out7.get("LOS_from_CHE_sep", 0.0))
    DIS_CHE_sep  = _safe(out7.get("DIS_from_CHE_sep", 0.0))
    PROD_CHE_sep = _safe(out7.get("chemsec_from_CHE_sep", out7.get("chemsec_from_CHE", 0.0)))

    EV_CHE_mix   = _safe(out7.get("EV_from_CHE_mix", 0.0))
    LOS_CHE_mix  = _safe(out7.get("LOS_from_CHE_mix", 0.0))
    DIS_CHE_mix  = _safe(out7.get("DIS_from_CHE_mix", 0.0))
    PROD_CHE_mix = _safe(out7.get("chemsec_from_CHE_mix", out7.get("chemsec_from_CHE", 0.0)))

    rPET_total = _safe(out7.get("rPET_total", (rPET_MEC_sep + rPET_MEC_mix)))
    rPET_open  = _safe(out7.get("rPET_openloop_total", max(rPET_total - rPET_to_PM, 0.0)))

    # ----------------------------
    # Nodes
    # ----------------------------
    nodes = [
        N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM, N_EXP, N_PMREST, N_MARKET, N_SAMM, N_STOCK,
        N_COL_SEP, N_COL_MIX,
        N_SORT_SEP, N_SORT_MIX,
        N_RECFR_SEP, N_RECFR_MIX,
        N_EV_COL_SEP, N_LOS_COL_SEP, N_DIS_COL_SEP,
        N_EV_COL_MIX, N_LOS_COL_MIX, N_DIS_COL_MIX,
        N_EV_SORT_SEP, N_LOS_SORT_SEP, N_DIS_SORT_SEP,
        N_EV_SORT_MIX, N_LOS_SORT_MIX, N_DIS_SORT_MIX,
        N_MEC_SEP, N_MEC_MIX, N_CHE_SEP, N_CHE_MIX,
        N_EV_MEC_SEP, N_LOS_MEC_SEP, N_MEC_OTH_SEP,
        N_EV_MEC_MIX, N_LOS_MEC_MIX, N_MEC_OTH_MIX,
        N_EV_CHE_SEP, N_LOS_CHE_SEP, N_CHE_PROD_SEP,
        N_EV_CHE_MIX, N_LOS_CHE_MIX, N_CHE_PROD_MIX,
        N_RPET_OPEN, N_REC_EXPORT
    ]
    idx = {n: i for i, n in enumerate(nodes)}

    links = []
    def L(a, b, v):
        v = float(v)
        if v <= 0:
            return
        if a not in idx or b not in idx:
            raise KeyError(f"Node missing: {a if a not in idx else b}")
        links.append({"source": a, "target": b, "value": v, "source_id": idx[a], "target_id": idx[b]})

    # ----------------------------
    # Links  / baseline-style
    # ----------------------------
    L(N_PRIM,   N_PM, prim_raw)
    L(N_IMP,    N_PM, imp_prod)
    L(N_CHESEC, N_PM, chemsec)
    L(N_RPET,   N_PM, rPET_to_PM)

    L(N_PM, N_EXP,    exp_prod)
    L(N_PM, N_MARKET, PM_market)
    L(N_PM, N_PMREST, PM_rest)

    L(N_MARKET, N_SAMM, waste_total)
    if stock_delta > 0:
        L(N_MARKET, N_STOCK, stock_delta)

    L(N_SAMM, N_COL_SEP, col_sep)
    L(N_SAMM, N_COL_MIX, col_mix)

    L(N_COL_SEP, N_SORT_SEP,    col_sep_to_sort_sep)
    L(N_COL_SEP, N_SORT_MIX,    col_sep_to_sort_mix)
    L(N_COL_SEP, N_EV_COL_SEP,  col_sep_to_ene)
    L(N_COL_SEP, N_LOS_COL_SEP, col_sep_to_los)
    L(N_COL_SEP, N_DIS_COL_SEP, col_sep_to_dis)

    L(N_COL_MIX, N_SORT_SEP,    col_mix_to_sort_sep)
    L(N_COL_MIX, N_SORT_MIX,    col_mix_to_sort_mix)
    L(N_COL_MIX, N_EV_COL_MIX,  col_mix_to_ene)
    L(N_COL_MIX, N_LOS_COL_MIX, col_mix_to_los)
    L(N_COL_MIX, N_DIS_COL_MIX, col_mix_to_dis)

    L(N_SORT_SEP,  N_RECFR_SEP,    sort_sep_to_rec_sep)
    L(N_SORT_SEP,  N_RECFR_MIX,    sort_sep_to_rec_mix)
    L(N_SORT_SEP,  N_EV_SORT_SEP,  sort_sep_to_ene)
    L(N_SORT_SEP,  N_LOS_SORT_SEP, sort_sep_to_los)
    L(N_SORT_SEP,  N_DIS_SORT_SEP, sort_sep_to_dis)

    L(N_SORT_MIX,  N_RECFR_SEP,    sort_mix_to_rec_sep)
    L(N_SORT_MIX,  N_RECFR_MIX,    sort_mix_to_rec_mix)
    L(N_SORT_MIX,  N_EV_SORT_MIX,  sort_mix_to_ene)
    L(N_SORT_MIX,  N_LOS_SORT_MIX, sort_mix_to_los)
    L(N_SORT_MIX,  N_DIS_SORT_MIX, sort_mix_to_dis)

    L(N_RECFR_SEP, N_MEC_SEP, rec_sep_to_MEC)
    L(N_RECFR_SEP, N_CHE_SEP, rec_sep_to_CHE)
    L(N_RECFR_MIX, N_MEC_MIX, rec_mix_to_MEC)
    L(N_RECFR_MIX, N_CHE_MIX, rec_mix_to_CHE)

    # MEC
    L(N_MEC_SEP, N_REC_EXPORT,  MEC_export_sep_u)
    L(N_MEC_SEP, N_RPET,        rPET_MEC_sep)
    L(N_MEC_SEP, N_EV_MEC_SEP,  EV_MEC_sep)
    L(N_MEC_SEP, N_LOS_MEC_SEP, LOS_MEC_sep)
    L(N_MEC_SEP, N_MEC_OTH_SEP, MEC_OTH_sep)

    L(N_MEC_MIX, N_REC_EXPORT,  MEC_export_mix_u)
    L(N_MEC_MIX, N_RPET,        rPET_MEC_mix)
    L(N_MEC_MIX, N_EV_MEC_MIX,  EV_MEC_mix)
    L(N_MEC_MIX, N_LOS_MEC_MIX, LOS_MEC_mix)
    L(N_MEC_MIX, N_MEC_OTH_MIX, MEC_OTH_mix)

    # CHE (baseline-style aggregated, if present)
    L(N_CHE_SEP, N_REC_EXPORT,   CHE_export_sep_u)
    L(N_CHE_SEP, N_EV_CHE_SEP,   EV_CHE_sep)
    L(N_CHE_SEP, N_LOS_CHE_SEP,  LOS_CHE_sep)
    L(N_CHE_SEP, N_CHE_PROD_SEP, PROD_CHE_sep)
    if DIS_CHE_sep > 0:
        L(N_CHE_SEP, f"Beseitigung (CHE, {SEP_TAG})", DIS_CHE_sep)

    L(N_CHE_MIX, N_REC_EXPORT,   CHE_export_mix_u)
    L(N_CHE_MIX, N_EV_CHE_MIX,   EV_CHE_mix)
    L(N_CHE_MIX, N_LOS_CHE_MIX,  LOS_CHE_mix)
    L(N_CHE_MIX, N_CHE_PROD_MIX, PROD_CHE_mix)
    if DIS_CHE_mix > 0:
        L(N_CHE_MIX, f"Beseitigung (CHE, {MIX_TAG})", DIS_CHE_mix)

    # rPET open-loop
    L(N_RPET, N_RPET_OPEN, rPET_open)

    # ----------------------------
    # Tables
    # ----------------------------
    df_links = pd.DataFrame(links)
    if df_links.empty:
        df_links = pd.DataFrame(columns=["source", "target", "value", "source_id", "target_id"])

    in_sum = df_links.groupby("target")["value"].sum() if not df_links.empty else pd.Series(dtype=float)
    out_sum = df_links.groupby("source")["value"].sum() if not df_links.empty else pd.Series(dtype=float)

    node_rows = []
    for n in nodes:
        tin = float(in_sum.get(n, 0.0))
        tout = float(out_sum.get(n, 0.0))
        node_rows.append({
            "node_id": idx[n],
            "node_name": n,
            "total_in": tin,
            "total_out": tout,
            "net_in_minus_out": tin - tout,
            "total_max(in,out)": max(tin, tout),
        })

    df_nodes = pd.DataFrame(node_rows).sort_values("node_id")

    df_check = pd.DataFrame([
        {"metric": "sum_links", "value": float(df_links["value"].sum() if not df_links.empty else 0.0)},
        {"metric": "PM_domestic_after_export(out7)", "value": float(PM_domestic)},
        {"metric": "PM_to_U_total(out7)", "value": float(PM_market)},
        {"metric": "rPET_total(out7)", "value": float(rPET_total)},
        {"metric": "rPET_to_PM_total(out7)", "value": float(rPET_to_PM)},
        {"metric": "rPET_openloop_total(out7)", "value": float(rPET_open)},
    ])

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_path / excel_filename

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_nodes.to_excel(writer, sheet_name="nodes", index=False)
        df_links.to_excel(writer, sheet_name="links", index=False)
        df_check.to_excel(writer, sheet_name="check", index=False)

    return str(xlsx_path.resolve())

# ============================================================
# Helpers: export rPET tagging rows to Excel
# ============================================================

def _safe_excel_float(x, default=0.0):
    try:
        v = float(x)
        return v
    except Exception:
        return float(default)


def _write_tagging_rows_to_excel(
    rows,
    out_dir="outputs",
    filename="rPET_tagging_results.xlsx",
    scenario_name="Scenario",
    year=2019,
    Mylog=None,
):
    import pandas as pd
    from pathlib import Path

    df = pd.DataFrame(rows)

    if df.empty:
        raise ValueError("No tagging rows available for Excel export.")

    # preferred order
    preferred_cols = [
        "scenario", "year",
        "row_type", "stage", "group",
        "source", "target", "name",
        "total_kt",
        "virgin_kt", "rPET_pfand_kt", "rPET_other_kt", "rPET_total_kt",
        "virgin_share", "rPET_pfand_share", "rPET_other_share", "rPET_total_share",
        "mass_balance_gap_kt",
    ]
    cols = [c for c in preferred_cols if c in df.columns] + [c for c in df.columns if c not in preferred_cols]
    df = df[cols]

    # split sheets
    df_links = df[df["row_type"] == "link"].copy() if "row_type" in df.columns else df.copy()
    df_nodes = df[df["row_type"] == "node"].copy() if "row_type" in df.columns else df.iloc[0:0].copy()

    # simple summary
    summary_rows = [
        {"metric": "scenario", "value": scenario_name},
        {"metric": "year", "value": year},
        {"metric": "n_rows_total", "value": len(df)},
        {"metric": "n_link_rows", "value": len(df_links)},
        {"metric": "n_node_rows", "value": len(df_nodes)},
    ]

    if "total_kt" in df.columns:
        summary_rows.append({"metric": "sum_total_kt_all_rows", "value": float(df["total_kt"].sum())})
    if "rPET_total_kt" in df.columns:
        summary_rows.append({"metric": "sum_rPET_total_kt_all_rows", "value": float(df["rPET_total_kt"].sum())})

    if not df_links.empty:
        # compact pivot
        try:
            pivot = df_links.pivot_table(
                index=["stage", "source", "target"],
                values=["total_kt", "virgin_kt", "rPET_pfand_kt", "rPET_other_kt", "rPET_total_kt"],
                aggfunc="sum"
            ).reset_index()
        except Exception:
            pivot = pd.DataFrame()
    else:
        pivot = pd.DataFrame()

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_path / filename

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="summary", index=False)
        df.to_excel(writer, sheet_name="all_rows", index=False)
        if not df_links.empty:
            df_links.to_excel(writer, sheet_name="link_rows", index=False)
        if not df_nodes.empty:
            df_nodes.to_excel(writer, sheet_name="node_rows", index=False)
        if not pivot.empty:
            pivot.to_excel(writer, sheet_name="pivot_links", index=False)

    if Mylog is not None:
        Mylog.info(f"### rPET TAGGING EXCEL saved: {xlsx_path.resolve()}")

    return str(xlsx_path.resolve())


def _make_tag_row(
    scenario_name,
    year,
    row_type,
    stage,
    group,
    source,
    target,
    name,
    total,
    virgin,
    rpet_pfand,
    rpet_other,
):
    total = _safe_excel_float(total)
    virgin = _safe_excel_float(virgin)
    rpet_pfand = _safe_excel_float(rpet_pfand)
    rpet_other = _safe_excel_float(rpet_other)

    rpet_total = rpet_pfand + rpet_other

    if total > 1e-12:
        virgin_share = virgin / total
        rpet_pfand_share = rpet_pfand / total
        rpet_other_share = rpet_other / total
        rpet_total_share = rpet_total / total
    else:
        virgin_share = 0.0
        rpet_pfand_share = 0.0
        rpet_other_share = 0.0
        rpet_total_share = 0.0

    return {
        "scenario": scenario_name,
        "year": year,
        "row_type": row_type,
        "stage": stage,
        "group": group,
        "source": source,
        "target": target,
        "name": name,
        "total_kt": total,
        "virgin_kt": virgin,
        "rPET_pfand_kt": rpet_pfand,
        "rPET_other_kt": rpet_other,
        "rPET_total_kt": rpet_total,
        "virgin_share": virgin_share,
        "rPET_pfand_share": rpet_pfand_share,
        "rPET_other_share": rpet_other_share,
        "rPET_total_share": rpet_total_share,
        "mass_balance_gap_kt": total - virgin - rpet_pfand - rpet_other,
    }

# ============================================================
# additional diagramm
# ============================================================

def export_step7_sankey_market_split_tags(
        out7,
        ParameterDict_mean,
        ModelClassification,
        Mylog,
        out_dir="outputs",
        filename="sankey_baseline_2019_Deutschland_marketSplitTags2.html",
        font_size=12,
        use_short_labels=True,
        export_excel=True,
        excel_filename="rPET_tagging_baseline_2019.xlsx",
        scenario_name="Baseline",
        year=2019,
):
    import numpy as np
    from pathlib import Path
    import plotly.graph_objects as go

    # -----------------------------
    # helpers
    # -----------------------------
    def _safe(x, default=0.0):
        try:
            v = float(x)
            return v if np.isfinite(v) else float(default)
        except Exception:
            return float(default)

    def _squeeze(a):
        return np.squeeze(np.asarray(a, dtype=float))

    def _item_name(x):
        return x.Name if hasattr(x, "Name") else str(x)

    def _find_index_contains(cls_key: str, target: str):
        names = [_item_name(it) for it in ModelClassification[cls_key].Items]
        low = [n.lower() for n in names]
        t = target.lower()
        for k, nm in enumerate(low):
            if t in nm:
                return k, names
        raise KeyError(f"Could not find '{target}' in '{cls_key}'. Available={names}")

    def _get_trp_value(param_arr, p_idx, default=0.0):
        A = _squeeze(param_arr)
        try:
            if A.ndim == 0:
                return float(A)
            if A.ndim == 1:
                return float(A[p_idx])
            if A.ndim == 2:
                return float(A[0, p_idx])
            return float(A[0, 0, p_idx])
        except Exception:
            return float(default)

    def hex_to_rgba(hex_color: str, a: float = 0.35) -> str:
        h = hex_color.lstrip("#")
        r = int(h[0:2], 16)
        g = int(h[2:4], 16)
        b = int(h[4:6], 16)
        return f"rgba({r},{g},{b},{a})"

    # -----------------------------
    # labels
    # -----------------------------
    if use_short_labels:
        SEP_TAG = "Pfandflaschen"
        MIX_TAG = "andere PET-Verp."
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG = "Recycling Fraktion"
    else:
        SEP_TAG = "PET-Getränkeflaschen aus Pfand"
        MIX_TAG = "andere PET-Verpackungen"
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG = "Recycling Fraktion"

    # upstream (UNSPLIT)
    N_PRIM = "Produktion von Primär-PET"
    N_IMP = "Importe (Produkte)"
    N_CHESEC = "Sekundäre Rohstoffe (CHE) / sonstige Inputs"
    N_RPET = "rPET (Pool)"
    N_PM = "Produktherstellung"
    N_EXP = "Exporte (Produkte)"
    N_PMREST = "Restabgänge (Produktherstellung) – Export/Bestand/Sonstiges"
    N_MARKET = "Verbrauch (Markt)"
    N_STOCK = "Bestand + unbekannter Anteil"
    B_RPET_OPEN = "rPET open-loop"

    # downstream base nodes (TAGGED)
    B_COL_SEP = f"{SAMM_TAG}_{SEP_TAG}"
    B_COL_MIX = f"{SAMM_TAG}_{MIX_TAG}"

    # explicit leak node to show α
    B_LEAK = "Leakage: Pfanddesign → LVP"

    # not collected / residual
    B_UNCOL = "Nicht getrennt erfasst / Restmüll (aus Markt)"

    B_SORT_SEP = f"{SORT_TAG}_{SEP_TAG}"
    B_SORT_MIX = f"{SORT_TAG}_{MIX_TAG}"

    B_RECFR_SEP = f"{REC_TAG}_{SEP_TAG}"
    B_RECFR_MIX = f"{REC_TAG}_{MIX_TAG}"

    B_EV_COL_SEP = f"EV ({SAMM_TAG}_{SEP_TAG})"
    B_LOS_COL_SEP = f"Verluste ({SAMM_TAG}_{SEP_TAG})"
    B_DIS_COL_SEP = f"Beseitigung ({SAMM_TAG}_{SEP_TAG})"

    B_EV_COL_MIX = f"EV ({SAMM_TAG}_{MIX_TAG})"
    B_LOS_COL_MIX = f"Verluste ({SAMM_TAG}_{MIX_TAG})"
    B_DIS_COL_MIX = f"Beseitigung ({SAMM_TAG}_{MIX_TAG})"

    B_EV_SORT_SEP = f"EV ({SORT_TAG}_{SEP_TAG})"
    B_LOS_SORT_SEP = f"Verluste ({SORT_TAG}_{SEP_TAG})"
    B_DIS_SORT_SEP = f"Beseitigung ({SORT_TAG}_{SEP_TAG})"

    B_EV_SORT_MIX = f"EV ({SORT_TAG}_{MIX_TAG})"
    B_LOS_SORT_MIX = f"Verluste ({SORT_TAG}_{MIX_TAG})"
    B_DIS_SORT_MIX = f"Beseitigung ({SORT_TAG}_{MIX_TAG})"

    B_MEC_SEP = f"Mechanisches Recycling (MEC, {SEP_TAG})"
    B_MEC_MIX = f"Mechanisches Recycling (MEC, {MIX_TAG})"
    B_CHE_SEP = f"Chemisches Recycling (CHE, {SEP_TAG})"
    B_CHE_MIX = f"Chemisches Recycling (CHE, {MIX_TAG})"

    B_EV_MEC_SEP = f"EV (MEC, {SEP_TAG})"
    B_LOS_MEC_SEP = f"Verluste (MEC, {SEP_TAG})"
    B_MEC_OTH_SEP = f"MEC-Rückstände / Sonstiges ({SEP_TAG})"

    B_EV_MEC_MIX = f"EV (MEC, {MIX_TAG})"
    B_LOS_MEC_MIX = f"Verluste (MEC, {MIX_TAG})"
    B_MEC_OTH_MIX = f"MEC-Rückstände / Sonstiges ({MIX_TAG})"

    B_EV_CHE_SEP = f"EV (CHE, {SEP_TAG})"
    B_LOS_CHE_SEP = f"Verluste (CHE, {SEP_TAG})"
    B_CHE_PROD_SEP = f"CHE-Produkte / open-loop ({SEP_TAG})"

    B_EV_CHE_MIX = f"EV (CHE, {MIX_TAG})"
    B_LOS_CHE_MIX = f"Verluste (CHE, {MIX_TAG})"
    B_CHE_PROD_MIX = f"CHE-Produkte / open-loop ({MIX_TAG})"

    B_REC_EXPORT = "Exporte (Recyclingprodukte)"

    # -----------------------------
    # read α, β (keep for logging / consistency)
    # -----------------------------
    p_pfand, _ = _find_index_contains("p", "pfandgetränkeflasche")
    p_other, _ = _find_index_contains("p", "andere pet-verpackungen")

    if "p_share_pfand_leak_to_LVP" not in ParameterDict_mean:
        raise KeyError("Missing parameter: p_share_pfand_leak_to_LVP")
    if "p_share_rPET_pfand_to_applications" not in ParameterDict_mean:
        raise KeyError("Missing parameter: p_share_rPET_pfand_to_applications")

    alpha = _get_trp_value(ParameterDict_mean["p_share_pfand_leak_to_LVP"], p_pfand, default=0.0)
    alpha = float(np.clip(alpha, 0.0, 0.5))

    beta_pf = _get_trp_value(ParameterDict_mean["p_share_rPET_pfand_to_applications"], p_pfand, default=np.nan)
    beta_ot = _get_trp_value(ParameterDict_mean["p_share_rPET_pfand_to_applications"], p_other, default=np.nan)
    beta_sum = float(max(beta_pf, 0.0) + max(beta_ot, 0.0)) if (np.isfinite(beta_pf) and np.isfinite(beta_ot)) else 0.0
    if beta_sum > 1e-12:
        beta_pf = float(max(beta_pf, 0.0) / beta_sum)
        beta_ot = float(max(beta_ot, 0.0) / beta_sum)
    else:
        beta_pf, beta_ot = 0.5, 0.5

    # -----------------------------
    # Pull values from out7
    # -----------------------------
    PM_market = _safe(out7.get("PM_to_U_total", 0.0))
    waste_total = _safe(out7.get("SANK_in_i_from_use", 0.0))
    stock_delta = _safe(PM_market - waste_total, 0.0)

    imp_prod = _safe(out7.get("product_import", 0.0))
    exp_prod = _safe(out7.get("product_export", 0.0))
    prim_raw = _safe(out7.get("PrimPET_to_PM_raw", 0.0))
    chemsec = _safe(out7.get("chemsec_to_PM", out7.get("chemsec_to_PM_raw", 0.0)))
    rPET_to_PM_raw = _safe(out7.get("rPET_to_PM_total", 0.0))

    PM_domestic = _safe(
        out7.get(
            "PM_domestic_after_export",
            max(prim_raw + imp_prod + chemsec + rPET_to_PM_raw - exp_prod, 0.0),
        )
    )
    PM_rest = _safe(out7.get("PM_rest_loss", max(PM_domestic - PM_market, 0.0)))

    # observed collection totals from out7
    col_sep_obs = _safe(out7.get("SANK_col_g", 0.0))
    col_mix_obs = _safe(out7.get("SANK_col_m", 0.0))

    # collection totals / uncollected
    col_total = col_sep_obs + col_mix_obs
    uncoll_total = float(max(waste_total - col_total, 0.0))

    # downstream splits
    col_sep_to_sort_sep = _safe(out7.get("SANK_col_g_to_sort_g", 0.0))
    col_sep_to_sort_mix = _safe(out7.get("SANK_col_g_to_sort_m", 0.0))
    col_sep_to_ene = _safe(out7.get("SANK_col_g_to_ene", 0.0))
    col_sep_to_los = _safe(out7.get("SANK_col_g_to_los", 0.0))
    col_sep_to_dis = _safe(out7.get("SANK_col_g_to_dis", 0.0))

    col_mix_to_sort_sep = _safe(out7.get("SANK_col_m_to_sort_g", 0.0))
    col_mix_to_sort_mix = _safe(out7.get("SANK_col_m_to_sort_m", 0.0))
    col_mix_to_ene = _safe(out7.get("SANK_col_m_to_ene", 0.0))
    col_mix_to_los = _safe(out7.get("SANK_col_m_to_los", 0.0))
    col_mix_to_dis = _safe(out7.get("SANK_col_m_to_dis", 0.0))

    sort_sep_to_rec_sep = _safe(out7.get("SANK_sort_g_to_rec_g", 0.0))
    sort_sep_to_rec_mix = _safe(out7.get("SANK_sort_g_to_rec_m", 0.0))
    sort_sep_to_ene = _safe(out7.get("SANK_sort_g_to_ene", 0.0))
    sort_sep_to_los = _safe(out7.get("SANK_sort_g_to_los", 0.0))
    sort_sep_to_dis = _safe(out7.get("SANK_sort_g_to_dis", 0.0))

    sort_mix_to_rec_sep = _safe(out7.get("SANK_sort_m_to_rec_g", 0.0))
    sort_mix_to_rec_mix = _safe(out7.get("SANK_sort_m_to_rec_m", 0.0))
    sort_mix_to_ene = _safe(out7.get("SANK_sort_m_to_ene", 0.0))
    sort_mix_to_los = _safe(out7.get("SANK_sort_m_to_los", 0.0))
    sort_mix_to_dis = _safe(out7.get("SANK_sort_m_to_dis", 0.0))

    rec_sep_to_MEC = _safe(out7.get("SANK_rec_g_to_MEC", 0.0))
    rec_sep_to_CHE = _safe(out7.get("SANK_rec_g_to_CHE", 0.0))
    rec_mix_to_MEC = _safe(out7.get("SANK_rec_m_to_MEC", 0.0))
    rec_mix_to_CHE = _safe(out7.get("SANK_rec_m_to_CHE", 0.0))

    MEC_in_sep = _safe(out7.get("MEC_in_sep", rec_sep_to_MEC))
    MEC_in_mix = _safe(out7.get("MEC_in_mix", rec_mix_to_MEC))
    CHE_in_sep = _safe(out7.get("CHE_in_sep", rec_sep_to_CHE))
    CHE_in_mix = _safe(out7.get("CHE_in_mix", rec_mix_to_CHE))

    MEC_export_sep = _safe(out7.get("SANK_MEC_export_sep", 0.0))
    MEC_export_mix = _safe(out7.get("SANK_MEC_export_mix", 0.0))
    CHE_export_sep = _safe(out7.get("SANK_CHE_export_sep", 0.0))
    CHE_export_mix = _safe(out7.get("SANK_CHE_export_mix", 0.0))

    # MEC outputs
    EV_MEC_sep = _safe(out7.get("EV_from_MEC_sep", 0.0))
    LOS_MEC_sep = _safe(out7.get("LOS_from_MEC_sep", 0.0))
    rPET_MEC_sep = _safe(out7.get("rPET_from_MEC_sep", 0.0))

    EV_MEC_mix = _safe(out7.get("EV_from_MEC_mix", 0.0))
    LOS_MEC_mix = _safe(out7.get("LOS_from_MEC_mix", 0.0))
    rPET_MEC_mix = _safe(out7.get("rPET_from_MEC_mix", 0.0))

    REM_MEC_sep = _safe(out7.get("REM_after_export_MEC_sep", max(MEC_in_sep - min(MEC_export_sep, MEC_in_sep), 0.0)))
    REM_MEC_mix = _safe(out7.get("REM_after_export_MEC_mix", max(MEC_in_mix - min(MEC_export_mix, MEC_in_mix), 0.0)))
    MEC_OTH_sep = max(REM_MEC_sep - (EV_MEC_sep + LOS_MEC_sep + rPET_MEC_sep), 0.0)
    MEC_OTH_mix = max(REM_MEC_mix - (EV_MEC_mix + LOS_MEC_mix + rPET_MEC_mix), 0.0)

    MEC_export_sep_u = min(MEC_export_sep, MEC_in_sep)
    MEC_export_mix_u = min(MEC_export_mix, MEC_in_mix)
    CHE_export_sep_u = min(CHE_export_sep, CHE_in_sep)
    CHE_export_mix_u = min(CHE_export_mix, CHE_in_mix)

    # CHE outputs
    EV_CHE_sep = _safe(out7.get("EV_from_CHE_sep", 0.0))
    EV_CHE_mix = _safe(out7.get("EV_from_CHE_mix", 0.0))
    CHE_prod_sep = _safe(out7.get("chemsec_from_CHE_sep", out7.get("chemsec_from_CHE", 0.0)))
    CHE_prod_mix = _safe(out7.get("chemsec_from_CHE_mix", out7.get("chemsec_from_CHE", 0.0)))

    rPET_total = _safe(out7.get("rPET_total", (rPET_MEC_sep + rPET_MEC_mix)))
    rPET_open = _safe(out7.get("rPET_openloop_total", max(rPET_total - rPET_to_PM_raw, 0.0)))

    # -----------------------------
    # Market embedded composition (scale to waste_total)
    # -----------------------------
    f_market = (PM_market / PM_domestic) if PM_domestic > 1e-12 else 0.0
    rPET_in_market = float(max(rPET_to_PM_raw * f_market, 0.0))
    rPET_in_market = float(min(rPET_in_market, PM_market))
    virgin_in_market = float(max(PM_market - rPET_in_market, 0.0))

    f_waste = (waste_total / PM_market) if PM_market > 1e-12 else 0.0
    virgin_in_waste = virgin_in_market * f_waste
    rPET_in_waste = rPET_in_market * f_waste

    share_pfand_origin_in_pool = (rPET_MEC_sep / rPET_total) if rPET_total > 1e-12 else 0.0
    rPET_pfand_origin_in_waste = rPET_in_waste * share_pfand_origin_in_pool
    rPET_other_origin_in_waste = max(rPET_in_waste - rPET_pfand_origin_in_waste, 0.0)

    # -----------------------------
    # Leak B-Option (visualize α) without double-count
    # - leak amount inferred from α and observed captured Pfand collection
    # -----------------------------
    if (1.0 - alpha) > 1e-12:
        waste_pfand_generated = col_sep_obs / (1.0 - alpha)
    else:
        waste_pfand_generated = col_sep_obs

    leak_pfand_to_LVP = float(max(waste_pfand_generated * alpha, 0.0))
    leak_pfand_to_LVP = float(min(leak_pfand_to_LVP, col_mix_obs))  # safety
    col_mix_other_only = float(max(col_mix_obs - leak_pfand_to_LVP, 0.0))

    # -----------------------------
    # Composition shares in waste (V / rPET_pfand / rPET_other)
    # -----------------------------
    def _shares(V, P, O):
        T = V + P + O
        return (V / T, P / T, O / T) if T > 1e-12 else (0.0, 0.0, 0.0)

    shV, shP, shO = _shares(virgin_in_waste, rPET_pfand_origin_in_waste, rPET_other_origin_in_waste)

    # -----------------------------
    # composition inflows to base nodes (scaled to out7 totals!)
    # -----------------------------
    comp_in = {}
    C_V = "VIRGIN"
    C_P = "rPET_pfand"
    C_O = "rPET_other"

    def _set_comp(node, V, P, O):
        comp_in[node] = {C_V: float(max(V, 0.0)), C_P: float(max(P, 0.0)), C_O: float(max(O, 0.0))}

    # Market -> (Sammlung_pfand) is exactly captured Pfand collection total (col_sep_obs)
    _set_comp(B_COL_SEP, col_sep_obs * shV, col_sep_obs * shP, col_sep_obs * shO)

    # Market -> Leak is exactly inferred leak amount (leak_pfand_to_LVP)
    _set_comp(B_LEAK, leak_pfand_to_LVP * shV, leak_pfand_to_LVP * shP, leak_pfand_to_LVP * shO)

    # Market -> Sammlung_andere is the rest of LVP collection (col_mix_other_only)
    _set_comp(B_COL_MIX, col_mix_other_only * shV, col_mix_other_only * shP, col_mix_other_only * shO)

    # Market -> Uncollected is waste not in separated collection
    _set_comp(B_UNCOL, uncoll_total * shV, uncoll_total * shP, uncoll_total * shO)

    # -----------------------------
    # Build BASE downstream graph (values only)
    # -----------------------------
    base_nodes = [
        B_COL_SEP, B_LEAK, B_COL_MIX, B_UNCOL,
        B_SORT_SEP, B_SORT_MIX,
        B_RECFR_SEP, B_RECFR_MIX,
        B_MEC_SEP, B_MEC_MIX, B_CHE_SEP, B_CHE_MIX,
        B_EV_COL_SEP, B_LOS_COL_SEP, B_DIS_COL_SEP,
        B_EV_COL_MIX, B_LOS_COL_MIX, B_DIS_COL_MIX,
        B_EV_SORT_SEP, B_LOS_SORT_SEP, B_DIS_SORT_SEP,
        B_EV_SORT_MIX, B_LOS_SORT_MIX, B_DIS_SORT_MIX,
        B_EV_MEC_SEP, B_LOS_MEC_SEP, B_MEC_OTH_SEP,
        B_EV_MEC_MIX, B_LOS_MEC_MIX, B_MEC_OTH_MIX,
        B_EV_CHE_SEP, B_LOS_CHE_SEP, B_CHE_PROD_SEP,
        B_EV_CHE_MIX, B_LOS_CHE_MIX, B_CHE_PROD_MIX,
        B_REC_EXPORT
    ]

    base_links = []

    def BL(a, b, v):
        v = float(v)
        if v <= 1e-12:
            return
        base_links.append((a, b, v))

    # Leak routing (B option): Market->Leak is injected; Leak goes into LVP collection
    BL(B_LEAK, B_COL_MIX, leak_pfand_to_LVP)

    # collection splits
    BL(B_COL_SEP, B_SORT_SEP, col_sep_to_sort_sep)
    BL(B_COL_SEP, B_SORT_MIX, col_sep_to_sort_mix)
    BL(B_COL_SEP, B_EV_COL_SEP, col_sep_to_ene)
    BL(B_COL_SEP, B_LOS_COL_SEP, col_sep_to_los)
    BL(B_COL_SEP, B_DIS_COL_SEP, col_sep_to_dis)

    BL(B_COL_MIX, B_SORT_SEP, col_mix_to_sort_sep)
    BL(B_COL_MIX, B_SORT_MIX, col_mix_to_sort_mix)
    BL(B_COL_MIX, B_EV_COL_MIX, col_mix_to_ene)
    BL(B_COL_MIX, B_LOS_COL_MIX, col_mix_to_los)
    BL(B_COL_MIX, B_DIS_COL_MIX, col_mix_to_dis)

    # sorting splits
    BL(B_SORT_SEP, B_RECFR_SEP, sort_sep_to_rec_sep)
    BL(B_SORT_SEP, B_RECFR_MIX, sort_sep_to_rec_mix)
    BL(B_SORT_SEP, B_EV_SORT_SEP, sort_sep_to_ene)
    BL(B_SORT_SEP, B_LOS_SORT_SEP, sort_sep_to_los)
    BL(B_SORT_SEP, B_DIS_SORT_SEP, sort_sep_to_dis)

    BL(B_SORT_MIX, B_RECFR_SEP, sort_mix_to_rec_sep)
    BL(B_SORT_MIX, B_RECFR_MIX, sort_mix_to_rec_mix)
    BL(B_SORT_MIX, B_EV_SORT_MIX, sort_mix_to_ene)
    BL(B_SORT_MIX, B_LOS_SORT_MIX, sort_mix_to_los)
    BL(B_SORT_MIX, B_DIS_SORT_MIX, sort_mix_to_dis)

    # recycling fraction -> processes
    BL(B_RECFR_SEP, B_MEC_SEP, rec_sep_to_MEC)
    BL(B_RECFR_SEP, B_CHE_SEP, rec_sep_to_CHE)
    BL(B_RECFR_MIX, B_MEC_MIX, rec_mix_to_MEC)
    BL(B_RECFR_MIX, B_CHE_MIX, rec_mix_to_CHE)

    # MEC outputs (rPET output to pool will be added later as TAGGED->UNSPLIT)
    BL(B_MEC_SEP, B_REC_EXPORT, MEC_export_sep_u)
    BL(B_MEC_SEP, B_EV_MEC_SEP, EV_MEC_sep)
    BL(B_MEC_SEP, B_LOS_MEC_SEP, LOS_MEC_sep)
    BL(B_MEC_SEP, B_MEC_OTH_SEP, MEC_OTH_sep)

    BL(B_MEC_MIX, B_REC_EXPORT, MEC_export_mix_u)
    BL(B_MEC_MIX, B_EV_MEC_MIX, EV_MEC_mix)
    BL(B_MEC_MIX, B_LOS_MEC_MIX, LOS_MEC_mix)
    BL(B_MEC_MIX, B_MEC_OTH_MIX, MEC_OTH_mix)

    # CHE outputs
    BL(B_CHE_SEP, B_REC_EXPORT, CHE_export_sep_u)
    BL(B_CHE_SEP, B_EV_CHE_SEP, EV_CHE_sep)
    BL(B_CHE_SEP, B_LOS_CHE_SEP, 0.0)
    BL(B_CHE_SEP, B_CHE_PROD_SEP, CHE_prod_sep)

    BL(B_CHE_MIX, B_REC_EXPORT, CHE_export_mix_u)
    BL(B_CHE_MIX, B_EV_CHE_MIX, EV_CHE_mix)
    BL(B_CHE_MIX, B_LOS_CHE_MIX, 0.0)
    BL(B_CHE_MIX, B_CHE_PROD_MIX, CHE_prod_mix)

    # -----------------------------
    # Composition propagation on BASE nodes
    # -----------------------------
    comp_at = {n: {C_V: 0.0, C_P: 0.0, C_O: 0.0} for n in base_nodes}
    for n, d in comp_in.items():
        if n in comp_at:
            comp_at[n] = dict(d)

    incoming = {n: [] for n in base_nodes}
    outgoing = {n: [] for n in base_nodes}
    for a, b, v in base_links:
        if a in outgoing:
            outgoing[a].append((b, v))
        if b in incoming:
            incoming[b].append((a, v))

    def _total_out(n):
        return float(sum(v for _, v in outgoing.get(n, [])))

    for _ in range(20):
        changed = False
        for n in base_nodes:
            if n in comp_in:
                continue
            incs = incoming.get(n, [])
            if not incs:
                continue
            Vnew = Pnew = Onew = 0.0
            for a, v_ab in incs:
                tout = _total_out(a)
                if tout <= 1e-12:
                    continue
                Va = comp_at[a][C_V]
                Pa = comp_at[a][C_P]
                Oa = comp_at[a][C_O]
                Ta = Va + Pa + Oa
                if Ta <= 1e-12:
                    continue
                shv, shp, sho = Va / Ta, Pa / Ta, Oa / Ta
                Vnew += v_ab * shv
                Pnew += v_ab * shp
                Onew += v_ab * sho
            prev = comp_at[n]
            if (abs(prev[C_V] - Vnew) + abs(prev[C_P] - Pnew) + abs(prev[C_O] - Onew)) > 1e-8:
                comp_at[n] = {C_V: Vnew, C_P: Pnew, C_O: Onew}
                changed = True
        if not changed:
            break

    # -----------------------------
    # Build Excel tagging rows
    # -----------------------------
    tagging_rows = []

    def _comp_triplet_from_node(base_node):
        V = _safe_excel_float(comp_at.get(base_node, {}).get(C_V, 0.0))
        P = _safe_excel_float(comp_at.get(base_node, {}).get(C_P, 0.0))
        O = _safe_excel_float(comp_at.get(base_node, {}).get(C_O, 0.0))
        T = V + P + O
        return V, P, O, T

    def _shares_from_node(base_node):
        V, P, O, T = _comp_triplet_from_node(base_node)
        if T <= 1e-12:
            return 0.0, 0.0, 0.0
        return V / T, P / T, O / T

    def _infer_stage(src_name, trg_name):
        s = str(src_name)
        t = str(trg_name)

        if s in (N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM) or t in (N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM):
            return "upstream"
        if s == N_MARKET:
            return "market_split"
        if s.startswith(SAMM_TAG) or t.startswith(SAMM_TAG):
            return "collection"
        if s.startswith(SORT_TAG) or t.startswith(SORT_TAG):
            return "sorting"
        if s.startswith(REC_TAG) or t.startswith(REC_TAG):
            return "recycling_fraction"
        if s.startswith("Mechanisches Recycling") or t.startswith("Mechanisches Recycling"):
            return "mec"
        if s.startswith("Chemisches Recycling") or t.startswith("Chemisches Recycling"):
            return "che"
        if s == B_RPET_OPEN or t == B_RPET_OPEN:
            return "open_loop"
        return "other"

    # 3-1) Node rows for all base downstream nodes
    for bn in base_nodes:
        V, P, O, T = _comp_triplet_from_node(bn)
        tagging_rows.append(
            _make_tag_row(
                scenario_name=scenario_name,
                year=year,
                row_type="node",
                stage=_infer_stage(bn, ""),
                group="base_node",
                source="",
                target="",
                name=bn,
                total=T,
                virgin=V,
                rpet_pfand=P,
                rpet_other=O,
            )
        )

    # 3-2) Market -> tagged injections
    for bn in [B_COL_SEP, B_LEAK, B_COL_MIX, B_UNCOL]:
        V, P, O, T = _comp_triplet_from_node(bn)
        tagging_rows.append(
            _make_tag_row(
                scenario_name=scenario_name,
                year=year,
                row_type="link",
                stage="market_split",
                group="market_injection",
                source=N_MARKET,
                target=bn,
                name=f"{N_MARKET} -> {bn}",
                total=T,
                virgin=V,
                rpet_pfand=P,
                rpet_other=O,
            )
        )

    # 3-3) Downstream base_links with propagated composition
    for a, b, v_ab in base_links:
        V_a, P_a, O_a, T_a = _comp_triplet_from_node(a)
        if T_a <= 1e-12 or v_ab <= 1e-12:
            continue

        shV = V_a / T_a
        shP = P_a / T_a
        shO = O_a / T_a

        tagging_rows.append(
            _make_tag_row(
                scenario_name=scenario_name,
                year=year,
                row_type="link",
                stage=_infer_stage(a, b),
                group="base_link",
                source=a,
                target=b,
                name=f"{a} -> {b}",
                total=v_ab,
                virgin=v_ab * shV,
                rpet_pfand=v_ab * shP,
                rpet_other=v_ab * shO,
            )
        )

    # 3-4) Upstream rows (unsplit + known composition)
    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "upstream", "upstream",
            N_PRIM, N_PM, f"{N_PRIM} -> {N_PM}",
            prim_raw, prim_raw, 0.0, 0.0
        )
    )
    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "upstream", "upstream",
            N_IMP, N_PM, f"{N_IMP} -> {N_PM}",
            imp_prod, imp_prod, 0.0, 0.0
        )
    )
    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "upstream", "upstream",
            N_CHESEC, N_PM, f"{N_CHESEC} -> {N_PM}",
            chemsec, chemsec, 0.0, 0.0
        )
    )
    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "pool_to_pm", "rpet_pool",
            N_RPET, N_PM, f"{N_RPET} -> {N_PM}",
            min(rPET_to_PM_raw, rPET_total if rPET_total > 0 else rPET_to_PM_raw),
            0.0, rPET_MEC_sep if rPET_total <= 1e-12 else min(rPET_to_PM_raw, rPET_total) * (rPET_MEC_sep / rPET_total if rPET_total > 1e-12 else 0.0),
            rPET_MEC_mix if rPET_total <= 1e-12 else min(rPET_to_PM_raw, rPET_total) * (rPET_MEC_mix / rPET_total if rPET_total > 1e-12 else 0.0),
        )
    )

    if rPET_open > 1e-12:
        tagging_rows.append(
            _make_tag_row(
                scenario_name, year, "link", "open_loop", "rpet_open_loop",
                N_RPET, B_RPET_OPEN, f"{N_RPET} -> {B_RPET_OPEN}",
                rPET_open,
                0.0,
                rPET_open * (rPET_MEC_sep / rPET_total if rPET_total > 1e-12 else 0.0),
                rPET_open * (rPET_MEC_mix / rPET_total if rPET_total > 1e-12 else 0.0),
            )
        )

    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "upstream", "pm_out",
            N_PM, N_EXP, f"{N_PM} -> {N_EXP}",
            exp_prod, exp_prod, 0.0, 0.0
        )
    )
    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "upstream", "pm_out",
            N_PM, N_MARKET, f"{N_PM} -> {N_MARKET}",
            PM_market, virgin_in_market, rPET_in_market * share_pfand_origin_in_pool, max(rPET_in_market - rPET_in_market * share_pfand_origin_in_pool, 0.0)
        )
    )
    tagging_rows.append(
        _make_tag_row(
            scenario_name, year, "link", "upstream", "pm_out",
            N_PM, N_PMREST, f"{N_PM} -> {N_PMREST}",
            PM_rest, PM_rest, 0.0, 0.0
        )
    )

    if stock_delta > 1e-12:
        tagging_rows.append(
            _make_tag_row(
                scenario_name, year, "link", "market_stock", "stock",
                N_MARKET, N_STOCK, f"{N_MARKET} -> {N_STOCK}",
                stock_delta,
                stock_delta * (virgin_in_market / PM_market if PM_market > 1e-12 else 0.0),
                stock_delta * ((rPET_in_market * share_pfand_origin_in_pool) / PM_market if PM_market > 1e-12 else 0.0),
                stock_delta * ((max(rPET_in_market - rPET_in_market * share_pfand_origin_in_pool, 0.0)) / PM_market if PM_market > 1e-12 else 0.0),
            )
        )

    # 3-5) MEC -> rPET Pool rows
    sh_mec_sep_V, sh_mec_sep_P, sh_mec_sep_O = _shares_from_node(B_MEC_SEP)
    sh_mec_mix_V, sh_mec_mix_P, sh_mec_mix_O = _shares_from_node(B_MEC_MIX)

    if rPET_MEC_sep > 1e-12:
        tagging_rows.append(
            _make_tag_row(
                scenario_name, year, "link", "mec_to_pool", "rpet_pool",
                B_MEC_SEP, N_RPET, f"{B_MEC_SEP} -> {N_RPET}",
                rPET_MEC_sep,
                rPET_MEC_sep * sh_mec_sep_V,
                rPET_MEC_sep * sh_mec_sep_P,
                rPET_MEC_sep * sh_mec_sep_O,
            )
        )

    if rPET_MEC_mix > 1e-12:
        tagging_rows.append(
            _make_tag_row(
                scenario_name, year, "link", "mec_to_pool", "rpet_pool",
                B_MEC_MIX, N_RPET, f"{B_MEC_MIX} -> {N_RPET}",
                rPET_MEC_mix,
                rPET_MEC_mix * sh_mec_mix_V,
                rPET_MEC_mix * sh_mec_mix_P,
                rPET_MEC_mix * sh_mec_mix_O,
            )
        )

    # -----------------------------
    # Expand BASE nodes into TAGGED nodes
    # -----------------------------
    comps = [C_V, C_P, C_O]

    def tag(n, c):
        return f"{n} | {c}"

    nodes = [
        N_PRIM, N_IMP, N_CHESEC, N_RPET, N_PM, N_EXP, N_PMREST, N_MARKET, N_STOCK, B_RPET_OPEN
    ]
    for bn in base_nodes:
        for c in comps:
            nodes.append(tag(bn, c))

    idx = {n: i for i, n in enumerate(nodes)}

    # -----------------------------
    # Colors (stage + rPET emphasis)
    # -----------------------------
    STAGE = {
        "market2col": "#4C566A",
        "col2sort":   "#2F8F9D",
        "sort2rec":   "#6B8CAF",
        "rec2proc":   "#3E6D9C",
        "mec":        "#2C6E9B",
        "che":        "#6A5D8C",
        "export":     "#2E3440",
        "upstream":   "#1C4E80",
        "stock":      "#D1D5DB",
        "leak":       "#8B5E34",

        # rPET colors
        "rpet_pfand": "#EFD46A",
        "rpet_other": "#EFD46A",
        "rpet_pool":  "#F6D97B",
    }

    def _is_tagged(n: str) -> bool:
        return " | " in n

    def _base_of(n: str) -> str:
        return n.split(" | ")[0] if _is_tagged(n) else n

    def _comp_of(n: str):
        return n.split(" | ")[1] if _is_tagged(n) else None

    def link_stage(a_base: str, b_base: str) -> str:
        # market injection to downstream
        if a_base == N_MARKET and (
            b_base.startswith(SAMM_TAG) or b_base.startswith("Leakage") or b_base.startswith("Nicht getrennt")
        ):
            return "market2col"

        # leak (internal)
        if a_base == B_LEAK and b_base == B_COL_MIX:
            return "leak"

        # Sammlung -> Sortierung
        if a_base.startswith(SAMM_TAG) and b_base.startswith(SORT_TAG):
            return "col2sort"
        # Sortierung -> Recycling Fraktion
        if a_base.startswith(SORT_TAG) and b_base.startswith(REC_TAG):
            return "sort2rec"
        # Recycling Fraktion -> MEC/CHE
        if a_base.startswith(REC_TAG) and (b_base.startswith("Mechanisches Recycling") or b_base.startswith("Chemisches Recycling")):
            return "rec2proc"
        # MEC / CHE stage
        if a_base.startswith("Mechanisches Recycling"):
            return "mec"
        if a_base.startswith("Chemisches Recycling"):
            return "che"
        # exports
        if "Exporte" in b_base:
            return "export"
        return "upstream"

    def link_kind(b_base: str) -> str:
        if b_base.startswith("EV"):
            return "ev"
        if b_base.startswith("Verluste"):
            return "loss"
        if b_base.startswith("Beseitigung"):
            return "dis"
        if b_base == B_REC_EXPORT or b_base.startswith("Exporte"):
            return "export"
        if b_base == N_RPET:
            return "rpetpool"
        return "main"

    def rgba_for_link(a: str, b: str, alpha: float) -> str:
        a_base = _base_of(a)
        b_base = _base_of(b)
        c = _comp_of(a)  # for tagged->tagged flows

        st = link_stage(a_base, b_base)
        kd = link_kind(b_base)

        # base stage color
        base_col = STAGE.get(st, STAGE["upstream"])

        # rPET override
        if c == C_P:
            base_col = STAGE["rpet_pfand"]
        elif c == C_O:
            base_col = STAGE["rpet_other"]

        # explicit rPET pool links emphasized
        if a_base == N_RPET or b_base == N_RPET:
            base_col = STAGE["rpet_pool"]

        # opacity (make rPET loop more visible)
        if kd == "ev":
            a2 = max(0.18, min(alpha, 0.28))
        elif kd == "loss":
            a2 = max(0.14, min(alpha, 0.22))
        elif kd == "dis":
            a2 = max(0.12, min(alpha, 0.20))
        elif kd == "export":
            a2 = max(0.20, min(alpha, 0.30))
        else:
            a2 = max(0.28, min(alpha, 0.45))

        # extra boost for rPET-related links
        if (c in (C_P, C_O)) or (a_base == N_RPET) or (b_base == N_RPET):
            a2 = max(a2, min(0.85, max(alpha, 0.70)))

        return hex_to_rgba(base_col, a2)

    def node_color(n: str) -> str:
        if n == N_RPET:
            return STAGE["rpet_pool"]
        if n in (N_PRIM, N_PM, N_CHESEC, N_IMP):
            return STAGE["upstream"]
        if n == N_MARKET:
            return STAGE["market2col"]
        if n == N_EXP or n.startswith("Exporte"):
            return STAGE["export"]
        if n == N_STOCK:
            return STAGE["stock"]
        if n == B_RPET_OPEN:
            return STAGE["rpet_pool"]

        b = _base_of(n)
        if b.startswith(SAMM_TAG):
            return STAGE["col2sort"]
        if b.startswith("Leakage"):
            return STAGE["leak"]
        if b.startswith("Nicht getrennt"):
            return STAGE["stock"]
        if b.startswith(SORT_TAG):
            return STAGE["sort2rec"]
        if b.startswith(REC_TAG):
            return STAGE["rec2proc"]
        if b.startswith("Mechanisches Recycling"):
            return STAGE["mec"]
        if b.startswith("Chemisches Recycling"):
            return STAGE["che"]
        if b.startswith("EV"):
            return STAGE["col2sort"]
        if b.startswith("Verluste") or b.startswith("Beseitigung"):
            return STAGE["upstream"]
        return STAGE["stock"]

    node_colors = [node_color(n) for n in nodes]

    # -----------------------------
    # Build Sankey links
    # -----------------------------
    src, trg, val, link_color = [], [], [], []

    def L(a, b, v, alpha=0.35):
        v = float(v)
        if v <= 1e-12:
            return
        src.append(idx[a])
        trg.append(idx[b])
        val.append(v)
        link_color.append(rgba_for_link(a, b, alpha))

    # upstream links
    L(N_PRIM, N_PM, prim_raw, alpha=0.28)
    L(N_IMP, N_PM, imp_prod, alpha=0.22)
    L(N_CHESEC, N_PM, chemsec, alpha=0.22)

    # rPET pool links (EMPHASIZE)
    if rPET_open > 1e-12:
        L(N_RPET, B_RPET_OPEN, rPET_open, alpha=0.70)

    # emphasize rPET pool -> production
    L(N_RPET, N_PM, min(rPET_to_PM_raw, rPET_total if rPET_total > 0 else rPET_to_PM_raw), alpha=0.90)

    L(N_PM, N_EXP, exp_prod, alpha=0.22)
    L(N_PM, N_MARKET, PM_market, alpha=0.28)
    L(N_PM, N_PMREST, PM_rest, alpha=0.16)
    if stock_delta > 0:
        L(N_MARKET, N_STOCK, stock_delta, alpha=0.18)

    # Market -> tagged injections
    for c in comps:
        L(N_MARKET, tag(B_COL_SEP, c), comp_at[B_COL_SEP][c], alpha=0.30)
        L(N_MARKET, tag(B_LEAK, c), comp_at[B_LEAK][c], alpha=0.30)
        L(N_MARKET, tag(B_COL_MIX, c), comp_at[B_COL_MIX][c], alpha=0.30)
        L(N_MARKET, tag(B_UNCOL, c), comp_at[B_UNCOL][c], alpha=0.18)

    # base_links -> tagged links
    for a, b, v_ab in base_links:
        Va, Pa, Oa = comp_at[a][C_V], comp_at[a][C_P], comp_at[a][C_O]
        Ta = Va + Pa + Oa
        if Ta <= 1e-12:
            continue
        sh = {C_V: Va / Ta, C_P: Pa / Ta, C_O: Oa / Ta}
        for c in comps:
            L(tag(a, c), tag(b, c), v_ab * sh[c], alpha=0.35)

    def _shares_at(base_node: str):
        V = comp_at[base_node][C_V]
        P = comp_at[base_node][C_P]
        O = comp_at[base_node][C_O]
        T = V + P + O
        if T <= 1e-12:
            return {C_V: 0.0, C_P: 0.0, C_O: 0.0}
        return {C_V: V / T, C_P: P / T, C_O: O / T}

    sh_mec_sep = _shares_at(B_MEC_SEP)
    sh_mec_mix = _shares_at(B_MEC_MIX)

    # MEC -> rPET pool (EMPHASIZE)
    for c in comps:
        if rPET_MEC_sep > 1e-12 and sh_mec_sep[c] > 1e-12:
            L(tag(B_MEC_SEP, c), N_RPET, rPET_MEC_sep * sh_mec_sep[c], alpha=0.95)
        if rPET_MEC_mix > 1e-12 and sh_mec_mix[c] > 1e-12:
            L(tag(B_MEC_MIX, c), N_RPET, rPET_MEC_mix * sh_mec_mix[c], alpha=0.95)

    # -----------------------------
    # Labels with totals
    # -----------------------------
    Nn = len(nodes)
    infl = np.zeros(Nn)
    outf = np.zeros(Nn)
    for s, t, v in zip(src, trg, val):
        outf[s] += v
        infl[t] += v

    def fmt(x):
        return f"{x:.3f} kt"

    labels = []
    for n in nodes:
        if n == N_MARKET:
            labels.append(f"{n}\n({fmt(PM_market)})")
        elif n == N_PM:
            labels.append(f"{n}\n({fmt(PM_domestic)})")
        elif n == N_RPET:
            i = idx[n]
            tot = max(infl[i], outf[i])
            labels.append(f"{n}\n({fmt(tot)})")
        elif " | " in n:
            base, c = n.split(" | ", 1)
            i = idx[n]
            tot = max(infl[i], outf[i])
            labels.append(f"{base}\n[{c}]\n({fmt(tot)})")
        else:
            i = idx[n]
            tot = max(infl[i], outf[i])
            labels.append(f"{n}\n({fmt(tot)})")

    fig = go.Figure(
        go.Sankey(
            arrangement="freeform",
            node=dict(
                label=labels,
                pad=18,
                thickness=24,
                color=node_colors,
                line=dict(color="rgba(0,0,0,0.12)", width=0.4),
            ),
            link=dict(
                source=src,
                target=trg,
                value=val,
                color=link_color,
                hovertemplate="%{source.label} → %{target.label}<br>%{value:.3f} kt<extra></extra>",
            ),
        )
    )

    fig.update_layout(
        title="PET/Polymer Verpackungen – Deutschland 2019, Sankey-Diagramm mit Marktphasen-Compositions-Tagging (kt)",
        font=dict(size=font_size),
    )

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    html_file = out_path / filename
    fig.write_html(str(html_file), config={"displayModeBar": True})

    excel_file = None
    if export_excel:
        excel_file = _write_tagging_rows_to_excel(
            rows=tagging_rows,
            out_dir=out_dir,
            filename=excel_filename,
            scenario_name=scenario_name,
            year=year,
            Mylog=Mylog,
        )

    # -----------------------------
    # Logging / mass checks
    # -----------------------------
    Mylog.info(
        "### MASS CHECK (MARKET->DOWNSTREAM): waste_total=%.3f | col_sep=%.3f | col_mix=%.3f | leak=%.3f | uncol=%.3f"
        % (waste_total, col_sep_obs, col_mix_obs, leak_pfand_to_LVP, uncoll_total)
    )

    Mylog.info(
        "### TAG-SANKEY saved: %s | alpha=%.4f | beta_pf=%.3f beta_ot=%.3f | "
        "PM_market=%.3f waste_total=%.3f | col_sep=%.3f col_mix=%.3f leak=%.3f uncol=%.3f | "
        "rPET_MEC_sep=%.3f rPET_MEC_mix=%.3f rPET_to_PM=%.3f"
        % (
            str(html_file.resolve()),
            alpha,
            beta_pf,
            beta_ot,
            PM_market,
            waste_total,
            col_sep_obs,
            col_mix_obs,
            leak_pfand_to_LVP,
            uncoll_total,
            rPET_MEC_sep,
            rPET_MEC_mix,
            rPET_to_PM_raw,
        )
    )

    if export_excel and excel_file is not None:
        Mylog.info("### TAGGING EXCEL saved: %s" % excel_file)

    return {
        "html_path": str(html_file.resolve()),
        "excel_path": excel_file
    }

def export_step7_sankey_postmarket_subsystem_A(
    out7,
    subsystem="pfand",
    out_dir="outputs",
    filename=None,
    font_size=14,
    use_short_labels=True,
    show_cross_stream=True,
):

    import numpy as np
    from pathlib import Path
    import plotly.graph_objects as go

    def _safe(x, default=0.0):
        try:
            v = float(x)
            return v if np.isfinite(v) else float(default)
        except Exception:
            return float(default)

    def hex_to_rgba(hex_color: str, a: float = 0.35) -> str:
        h = hex_color.lstrip("#")
        r = int(h[0:2], 16); g = int(h[2:4], 16); b = int(h[4:6], 16)
        return f"rgba({r},{g},{b},{a})"

    # -----------------------------
    # labels
    # -----------------------------
    if use_short_labels:
        SEP_TAG  = "Pfandflaschen"
        MIX_TAG  = "andere PET-Verp."
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"
    else:
        SEP_TAG  = "PET-Getränkeflaschen aus Pfand"
        MIX_TAG  = "andere PET-Verpackungen"
        SAMM_TAG = "Sammlung"
        SORT_TAG = "Sortierung"
        REC_TAG  = "Recycling Fraktion"

    if subsystem not in ("pfand", "other"):
        raise ValueError("subsystem must be 'pfand' or 'other'")

    # choose stream keys/labels
    is_pf = (subsystem == "pfand")
    stream_label = SEP_TAG if is_pf else MIX_TAG
    key_col = "g" if is_pf else "m"

    # -----------------------------
    # nodes
    # -----------------------------
    N_MARKET = "Verbrauch (Markt)"
    N_COL    = f"{SAMM_TAG}_{stream_label}"
    N_SORT   = f"{SORT_TAG}_{stream_label}"
    N_RECFR  = f"{REC_TAG}_{stream_label}"

    N_EV_COL  = f"EV ({SAMM_TAG}_{stream_label})"
    N_LOS_COL = f"Verluste ({SAMM_TAG}_{stream_label})"
    N_DIS_COL = f"Beseitigung ({SAMM_TAG}_{stream_label})"

    N_EV_SORT  = f"EV ({SORT_TAG}_{stream_label})"
    N_LOS_SORT = f"Verluste ({SORT_TAG}_{stream_label})"
    N_DIS_SORT = f"Beseitigung ({SORT_TAG}_{stream_label})"

    N_MEC = f"Mechanisches Recycling (MEC, {stream_label})"
    N_CHE = f"Chemisches Recycling (CHE, {stream_label})"

    N_REC_EXPORT = "Exporte (Recyclingprodukte)"
    N_RPET_PROD  = f"rPET-Produkt (MEC, {stream_label})"
    N_CHE_PROD   = f"CHE-Produkte / open-loop ({stream_label})"
    N_EV_MEC     = f"EV (MEC, {stream_label})"
    N_LOS_MEC    = f"Verluste (MEC, {stream_label})"
    N_MEC_OTH    = f"MEC-Rückstände / Sonstiges ({stream_label})"
    N_EV_CHE     = f"EV (CHE, {stream_label})"

    N_SORT_X = f"{SORT_TAG}_(anderer Strom)"
    N_RECFR_X = f"{REC_TAG}_(anderer Strom)"

    nodes = [
        N_MARKET,
        N_COL, N_SORT, N_RECFR,
        N_EV_COL, N_LOS_COL, N_DIS_COL,
        N_EV_SORT, N_LOS_SORT, N_DIS_SORT,
        N_MEC, N_CHE,
        N_EV_MEC, N_LOS_MEC, N_MEC_OTH,
        N_EV_CHE,
        N_REC_EXPORT, N_RPET_PROD, N_CHE_PROD,
    ]
    if show_cross_stream:
        nodes += [N_SORT_X, N_RECFR_X]

    idx = {n: i for i, n in enumerate(nodes)}

    # -----------------------------
    # colors (RWTH-ish, consistent)
    # -----------------------------
    COL = {
        "market":  "#4C566A",
        "sammlung":"#2F8F9D",
        "sort":    "#6B8CAF",
        "recfr":   "#3E6D9C",
        "mec":     "#2C6E9B",
        "che":     "#6A5D8C",
        "export":  "#2E3440",
        "ev":      "#1F6F78",
        "loss":    "#111827",
        "dis":     "#374151",
        "prod":    "#1C4E80",
        "other":   "#D1D5DB",
    }

    def node_color(n: str) -> str:
        if n == N_MARKET:
            return COL["market"]
        if n.startswith(SAMM_TAG):
            return COL["sammlung"]
        if n.startswith(SORT_TAG):
            return COL["sort"]
        if n.startswith(REC_TAG):
            return COL["recfr"]
        if n.startswith("Mechanisches Recycling"):
            return COL["mec"]
        if n.startswith("Chemisches Recycling"):
            return COL["che"]
        if n.startswith("Exporte"):
            return COL["export"]
        if n.startswith("EV"):
            return COL["ev"]
        if n.startswith("Verluste"):
            return COL["loss"]
        if n.startswith("Beseitigung"):
            return COL["dis"]
        if "rPET-Produkt" in n or "CHE-Produkte" in n:
            return COL["prod"]
        if "anderer Strom" in n:
            return COL["other"]
        return COL["other"]

    node_colors = [hex_to_rgba(node_color(n), 0.95) for n in nodes]

    # -----------------------------
    # pull flows from out7
    # -----------------------------
    # collection inflow observed
    col_in = _safe(out7.get("SANK_col_g" if is_pf else "SANK_col_m", 0.0))

    # collection -> sorting + rejects
    if is_pf:
        col_to_sort_same  = _safe(out7.get("SANK_col_g_to_sort_g", 0.0))
        col_to_sort_other = _safe(out7.get("SANK_col_g_to_sort_m", 0.0))
        col_to_ev  = _safe(out7.get("SANK_col_g_to_ene", 0.0))
        col_to_los = _safe(out7.get("SANK_col_g_to_los", 0.0))
        col_to_dis = _safe(out7.get("SANK_col_g_to_dis", 0.0))
    else:
        col_to_sort_same  = _safe(out7.get("SANK_col_m_to_sort_m", 0.0))
        col_to_sort_other = _safe(out7.get("SANK_col_m_to_sort_g", 0.0))
        col_to_ev  = _safe(out7.get("SANK_col_m_to_ene", 0.0))
        col_to_los = _safe(out7.get("SANK_col_m_to_los", 0.0))
        col_to_dis = _safe(out7.get("SANK_col_m_to_dis", 0.0))

    # sorting -> recycling fraction + rejects
    if is_pf:
        sort_to_rec_same  = _safe(out7.get("SANK_sort_g_to_rec_g", 0.0))
        sort_to_rec_other = _safe(out7.get("SANK_sort_g_to_rec_m", 0.0))
        sort_to_ev  = _safe(out7.get("SANK_sort_g_to_ene", 0.0))
        sort_to_los = _safe(out7.get("SANK_sort_g_to_los", 0.0))
        sort_to_dis = _safe(out7.get("SANK_sort_g_to_dis", 0.0))
    else:
        sort_to_rec_same  = _safe(out7.get("SANK_sort_m_to_rec_m", 0.0))
        sort_to_rec_other = _safe(out7.get("SANK_sort_m_to_rec_g", 0.0))
        sort_to_ev  = _safe(out7.get("SANK_sort_m_to_ene", 0.0))
        sort_to_los = _safe(out7.get("SANK_sort_m_to_los", 0.0))
        sort_to_dis = _safe(out7.get("SANK_sort_m_to_dis", 0.0))

    # recycling fraction -> processes
    if is_pf:
        rec_to_MEC = _safe(out7.get("SANK_rec_g_to_MEC", 0.0))
        rec_to_CHE = _safe(out7.get("SANK_rec_g_to_CHE", 0.0))
        MEC_in = _safe(out7.get("MEC_in_sep", rec_to_MEC))
        CHE_in = _safe(out7.get("CHE_in_sep", rec_to_CHE))
        MEC_export = _safe(out7.get("SANK_MEC_export_sep", 0.0))
        CHE_export = _safe(out7.get("SANK_CHE_export_sep", 0.0))
        EV_MEC   = _safe(out7.get("EV_from_MEC_sep", 0.0))
        LOS_MEC  = _safe(out7.get("LOS_from_MEC_sep", 0.0))
        rPET_MEC = _safe(out7.get("rPET_from_MEC_sep", 0.0))
        EV_CHE   = _safe(out7.get("EV_from_CHE_sep", 0.0))
        CHE_prod = _safe(out7.get("chemsec_from_CHE_sep", out7.get("chemsec_from_CHE", 0.0)))
        REM_after = _safe(out7.get("REM_after_export_MEC_sep", max(MEC_in - min(MEC_export, MEC_in), 0.0)))
    else:
        rec_to_MEC = _safe(out7.get("SANK_rec_m_to_MEC", 0.0))
        rec_to_CHE = _safe(out7.get("SANK_rec_m_to_CHE", 0.0))
        MEC_in = _safe(out7.get("MEC_in_mix", rec_to_MEC))
        CHE_in = _safe(out7.get("CHE_in_mix", rec_to_CHE))
        MEC_export = _safe(out7.get("SANK_MEC_export_mix", 0.0))
        CHE_export = _safe(out7.get("SANK_CHE_export_mix", 0.0))
        EV_MEC   = _safe(out7.get("EV_from_MEC_mix", 0.0))
        LOS_MEC  = _safe(out7.get("LOS_from_MEC_mix", 0.0))
        rPET_MEC = _safe(out7.get("rPET_from_MEC_mix", 0.0))
        EV_CHE   = _safe(out7.get("EV_from_CHE_mix", 0.0))
        CHE_prod = _safe(out7.get("chemsec_from_CHE_mix", out7.get("chemsec_from_CHE", 0.0)))
        REM_after = _safe(out7.get("REM_after_export_MEC_mix", max(MEC_in - min(MEC_export, MEC_in), 0.0)))

    MEC_export_u = min(MEC_export, MEC_in)
    CHE_export_u = min(CHE_export, CHE_in)
    MEC_OTH = max(REM_after - (EV_MEC + LOS_MEC + rPET_MEC), 0.0)

    # -----------------------------
    # links
    # -----------------------------
    src, trg, val, lcol = [], [], [], []

    def add(a, b, v, a_link=0.35):
        v = float(v)
        if v <= 1e-12:
            return
        if a not in idx or b not in idx:
            return
        src.append(idx[a]); trg.append(idx[b]); val.append(v)
        lcol.append(hex_to_rgba(node_color(b), a_link))

    # Market -> collection (subsystem inflow)
    add(N_MARKET, N_COL, col_in, a_link=0.35)

    # Collection -> Sorting (+ cross-stream sink)
    add(N_COL, N_SORT, col_to_sort_same, a_link=0.35)
    if show_cross_stream:
        add(N_COL, N_SORT_X, col_to_sort_other, a_link=0.22)

    add(N_COL, N_EV_COL,  col_to_ev,  a_link=0.25)
    add(N_COL, N_LOS_COL, col_to_los, a_link=0.20)
    add(N_COL, N_DIS_COL, col_to_dis, a_link=0.20)

    # Sorting -> Recycling fraction (+ cross-stream sink)
    add(N_SORT, N_RECFR, sort_to_rec_same, a_link=0.35)
    if show_cross_stream:
        add(N_SORT, N_RECFR_X, sort_to_rec_other, a_link=0.22)

    add(N_SORT, N_EV_SORT,  sort_to_ev,  a_link=0.25)
    add(N_SORT, N_LOS_SORT, sort_to_los, a_link=0.20)
    add(N_SORT, N_DIS_SORT, sort_to_dis, a_link=0.20)

    # Recycling fraction -> processes
    add(N_RECFR, N_MEC, rec_to_MEC, a_link=0.35)
    add(N_RECFR, N_CHE, rec_to_CHE, a_link=0.35)

    # MEC outputs
    add(N_MEC, N_REC_EXPORT, MEC_export_u, a_link=0.25)
    add(N_MEC, N_EV_MEC,     EV_MEC,       a_link=0.25)
    add(N_MEC, N_LOS_MEC,    LOS_MEC,      a_link=0.20)
    add(N_MEC, N_MEC_OTH,    MEC_OTH,      a_link=0.20)
    add(N_MEC, N_RPET_PROD,  rPET_MEC,     a_link=0.35)

    # CHE outputs
    add(N_CHE, N_REC_EXPORT, CHE_export_u, a_link=0.25)
    add(N_CHE, N_EV_CHE,     EV_CHE,       a_link=0.25)
    add(N_CHE, N_CHE_PROD,   CHE_prod,     a_link=0.35)

    # -----------------------------
    # labels with totals
    # -----------------------------
    infl = np.zeros(len(nodes)); outf = np.zeros(len(nodes))
    for s, t, v in zip(src, trg, val):
        outf[s] += v
        infl[t] += v

    def fmt(x): return f"{x:.3f} kt"
    labels = []
    for n in nodes:
        i = idx[n]
        tot = max(infl[i], outf[i])
        labels.append(f"{n}\n({fmt(tot)})")

    title = f"PET Verpackungen – Baseline 2019 ({'Pfandflaschen' if is_pf else 'andere PET-Verpackungen'} | Post-market Subsystem, kt)"

    if filename is None:
        filename = f"sankey_baseline_2019_{subsystem}_postmarket_subsystem_A.html"

    fig = go.Figure(go.Sankey(
        arrangement="freeform",
        node=dict(
            label=labels,
            pad=18,
            thickness=24,
            color=node_colors,
            line=dict(color="rgba(0,0,0,0.12)", width=0.4),
        ),
        link=dict(
            source=src,
            target=trg,
            value=val,
            color=lcol,
            hovertemplate="%{source.label} → %{target.label}<br>%{value:.3f} kt<extra></extra>",
        )
    ))
    fig.update_layout(title=title, font=dict(size=font_size))

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    html_file = out_path / filename
    fig.write_html(str(html_file), config={"displayModeBar": True})

    return str(html_file.resolve())
# ============================================================
# STEP 6.4 Deterministic reference run (mean values) + Sankey
# ============================================================

Mylog.info("### STEP6-REF: Running deterministic Step7 (mean values)")

ParameterDict_mean = {}
for name, entry in ParameterDict.items():
    V, _SD = _get_V_SD(entry)
    ParameterDict_mean[name] = np.asarray(V, dtype=float)

t0 = time.time()
out7_ref = run_step7(
    ParameterDict=ParameterDict_mean,
    PET_MFA_System=PET_MFA_System,
    ModelClassification=ModelClassification,
    Mylog=Mylog,
    verbose=True,
    run_type="single"
)
out7_ref = apply_step7_hybrid_injection(out7_ref, Mylog, log=True)

# ============================================================
# Sankey + rPET tagging export
# ============================================================

html_path = export_step7_sankey_market_split_tags(
    out7=out7_ref,
    ParameterDict_mean=ParameterDict_mean,
    ModelClassification=ModelClassification,
    Mylog=Mylog,
    out_dir="outputs",
    filename="sankey_market_split_tags.html",
    font_size=12,
    use_short_labels=True,
    export_excel=True,
    excel_filename="rPET_tagging.xlsx",
    scenario_name="S3-a2",
    year=2019,
)

Mylog.info(f"### Sankey saved: {html_path}")

Mylog.info(
    "### STEP6-REF done in %.2fs | PM_domestic=%.3f | PM_market=%.3f | PM_rest=%.3f | Use_to_Waste=%.3f | rPET_to_PM=%.3f | imports=%.3f exports=%.3f | Prim_raw=%.3f"
    % (
        time.time() - t0,
        float(out7_ref.get("PM_domestic_after_export", 0.0)),
        float(out7_ref.get("PM_to_U_total", 0.0)),
        float(out7_ref.get("PM_rest_loss", 0.0)),
        float(out7_ref.get("Use_to_Waste_total", 0.0)),
        float(out7_ref.get("rPET_to_PM_total", 0.0)),
        float(out7_ref.get("product_import", 0.0)),
        float(out7_ref.get("product_export", 0.0)),
        float(out7_ref.get("PrimPET_to_PM_raw", 0.0)),
    )
)

# --- SAVE SANKEY (baseline only) ---
sankey_path = export_step7_sankey_no_new_params(
    out7_ref,
    out_dir="outputs",
    filename="sankey_baseline_2019_Deutschland_bk.html",
    font_size=12
)
Mylog.info(f"### SANKEY SAVED TO: {sankey_path}")
# --- save excel ---
excel_path = export_step7_nodes_links_excel_Baseline(
    out7_ref,
    out_dir="outputs",
    excel_filename="nodes_links_Baseline_bk.xlsx",
    use_short_labels=True
)
Mylog.info(f"### Baseline NODES+LINKS EXCEL SAVED TO: {excel_path}")

# --- save addtional diagram ---

sankey_path = export_step7_sankey_market_split_tags(
    out7=out7_ref,
    ParameterDict_mean=ParameterDict_mean,
    ModelClassification=ModelClassification,
    Mylog=Mylog,
    out_dir="outputs",
    filename="sankey_baseline_2019_marketSplitTags2.html",
    font_size=12,
    use_short_labels=True,
)
Mylog.info("Saved tag-sankey: %s", sankey_path)

#---save separate diagram---
pfand_path = export_step7_sankey_postmarket_subsystem_A(
    out7_ref,
    subsystem="pfand",
    out_dir="outputs",
    filename="sankey_baseline_2019_pfand_postmarket_A.html",
    font_size=14,
    use_short_labels=True,
    show_cross_stream=True,
)

other_path = export_step7_sankey_postmarket_subsystem_A(
    out7_ref,
    subsystem="other",
    out_dir="outputs",
    filename="sankey_baseline_2019_other_postmarket_A.html",
    font_size=14,
    use_short_labels=True,
    show_cross_stream=True,
)

Mylog.info("Saved PFAND post-market A: %s", pfand_path)
Mylog.info("Saved OTHER post-market A: %s", other_path)
# ============================================================
# STEP 6.5 Monte Carlo loop (N=500) - minimal logs
# ============================================================

Mylog.info(f"### STEP6-MC: Running Monte Carlo N={N_MC}")

records = []
t0 = time.time()

for n in range(N_MC):
    PD_s = build_sampled_parameterdict(ParameterDict, rng)

    out7 = run_step7(
        ParameterDict=PD_s,
        PET_MFA_System=PET_MFA_System,
        ModelClassification=ModelClassification,
        Mylog=Mylog,
        verbose=False,
        run_type="single"
    )
    out7 = apply_step7_hybrid_injection(out7, Mylog, log=False)

    rec = {
        "mc_run": n,

        "PM_to_U_total_primary": float(out7.get("PM_to_U_total_primary", np.nan)),
        "PM_to_U_total_hybrid":  float(out7.get("PM_to_U_total_hybrid", np.nan)),
        "Use_to_Waste_total":    float(out7.get("Use_to_Waste_total", np.nan)),
        "DeltaStock_u_total":    float(out7.get("DeltaStock_u_total", np.nan)),
        "DeltaStock_u_total_hybrid": float(out7.get("DeltaStock_u_total_hybrid", np.nan)),

        "product_import": float(out7.get("product_import", np.nan)),
        "product_export": float(out7.get("product_export", np.nan)),

        "PrimPET_to_PM_raw": float(out7.get("PrimPET_to_PM_raw", np.nan)),

        "PM_domestic_after_export": float(out7.get("PM_domestic_after_export", np.nan)),
        "PM_rest_loss": float(out7.get("PM_rest_loss", np.nan)),

        "SANK_in_i_from_use": float(out7.get("SANK_in_i_from_use", np.nan)),

        "SANK_rec_m_to_CHE": float(out7.get("SANK_rec_m_to_CHE", np.nan)),
        "SANK_rec_g_to_CHE": float(out7.get("SANK_rec_g_to_CHE", np.nan)),

        "MEC_in": float(out7.get("MEC_in", np.nan)),
        "CHE_in": float(out7.get("CHE_in", np.nan)),

        "EV_from_MEC_total": float(out7.get("EV_from_MEC_total", np.nan)),
        "LOS_from_MEC_total": float(out7.get("LOS_from_MEC_total", np.nan)),
        "EV_from_CHE_total": float(out7.get("EV_from_CHE_total", np.nan)),
        "LOS_from_CHE_total": float(out7.get("LOS_from_CHE_total", np.nan)),

        "SANK_MEC_export": float(out7.get("SANK_MEC_export", np.nan)),
        "SANK_CHE_export": float(out7.get("SANK_CHE_export", np.nan)),

        "rPET_total": float(out7.get("rPET_total", np.nan)),
        "rPET_to_PM_total": float(out7.get("rPET_to_PM_total", np.nan)),
        "rPET_openloop_total": float(out7.get("rPET_openloop_total", np.nan)),
    }
    records.append(rec)

    if (n + 1) % LOG_EVERY == 0:
        Mylog.info(f"### STEP6-MC progress: {n+1}/{N_MC}")

Mylog.info("### STEP6-MC finished in %.2fs" % (time.time() - t0))


# ============================================================
# STEP 6.6 Save results + percentile summary
# ============================================================

df_mc = pd.DataFrame.from_records(records)

mc_csv = MC_DIR / f"mc_results_N{N_MC}_seed{SEED}.csv"
df_mc.to_csv(mc_csv, index=False)
Mylog.info(f"### STEP6 saved MC raw results: {mc_csv}")

def summarize_percentiles(df, cols, qs=(0.05, 0.50, 0.95)):
    out_rows = []
    for c in cols:
        s = df[c].dropna()
        if len(s) == 0:
            continue
        row = {"metric": c, "n": int(s.shape[0])}
        for q in qs:
            row[f"p{int(q*100):02d}"] = float(s.quantile(q))
        row["mean"] = float(s.mean())
        row["std"]  = float(s.std(ddof=1)) if s.shape[0] > 1 else 0.0
        out_rows.append(row)
    return pd.DataFrame(out_rows)

metrics = [c for c in df_mc.columns if c != "mc_run"]
df_sum = summarize_percentiles(df_mc, metrics)

sum_csv = MC_DIR / f"mc_summary_N{N_MC}_seed{SEED}.csv"
df_sum.to_csv(sum_csv, index=False)
Mylog.info(f"### STEP6 saved MC summary: {sum_csv}")

Mylog.info("### STEP6 summary preview (first 10 rows):")
Mylog.info("\n" + df_sum.head(10).to_string(index=False))

MC_RESULTS = {
    "df_mc": df_mc,
    "df_summary": df_sum,
    "mc_csv": str(mc_csv),
    "sum_csv": str(sum_csv),
    "out7_ref": out7_ref,
    "sankey_path": sankey_path,
}

import json
import numpy as np
from pathlib import Path

def _to_float(x):
    try:
        return float(np.nansum(x))
    except Exception:
        try:
            return float(x)
        except Exception:
            return None

baseline_export = {
    "PM_domestic_after_export": _to_float(out7_ref.get("PM_domestic_after_export", 0.0)),
    "PM_to_U_total": _to_float(out7_ref.get("PM_to_U_total", 0.0)),
    "PM_rest_loss": _to_float(out7_ref.get("PM_rest_loss", 0.0)),

    "PrimPET_to_PM_raw": _to_float(out7_ref.get("PrimPET_to_PM_raw", 0.0)),
    "rPET_to_PM_total": _to_float(out7_ref.get("rPET_to_PM_total", 0.0)),
    "rPET_total": _to_float(out7_ref.get("rPET_total", 0.0)),

    "product_import": _to_float(out7_ref.get("product_import", 0.0)),
    "product_export": _to_float(out7_ref.get("product_export", 0.0)),
    "chemsec_to_PM_raw": _to_float(out7_ref.get("chemsec_to_PM_raw", out7_ref.get("chemsec_to_PM", 0.0))),
}

out_path = Path(PROJECT_DIR) / "results" / "baseline_out7.json"
out_path.parent.mkdir(parents=True, exist_ok=True)
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(baseline_export, f, ensure_ascii=False, indent=2)

Mylog.info(f"### Saved baseline_out7 snapshot: {out_path}")